import datetime
import os
import re
from typing import Dict, Optional, List
from datetime import timezone
from dotenv import load_dotenv
from fastapi import FastAPI, Depends, HTTPException, Request, Response, status, Query, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import JSONResponse
import logging
import requests
from app.auth import *
from app.database import *
from app.database import fs
from app.database import fs, warehouse_master_collection, task_assignments_collection
from app.models import AuditForm, UserLogin, UserRegister
from fastapi.responses import JSONResponse, StreamingResponse, FileResponse
import io
from docx import Document
from docx.shared import Pt
import base64, re
from docx.shared import Inches
import smtplib
import asyncio
import traceback
import time
from email.message import EmailMessage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import tempfile
from fastapi import UploadFile, Form
from fastapi import FastAPI, Depends, Form, File, UploadFile, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse, RedirectResponse
from datetime import datetime
import io
import os
import smtplib
from email.message import EmailMessage
import base64
import re
import logging
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import smtplib
from email.message import EmailMessage
from fastapi import UploadFile, Form, File, Depends
from typing import Optional
from datetime import datetime, timezone
from fastapi.responses import JSONResponse
import os, io, base64, re
from docx import Document
from docx.shared import Pt, Inches
import bcrypt
import pandas as pd
from bson import ObjectId
from pydantic import BaseModel

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# User name cache to avoid repeated database queries
USER_NAME_CACHE = {}
CACHE_TIMESTAMP = None

def get_user_names_cached(user_emails):
    """Get user names with caching to improve performance."""
    global USER_NAME_CACHE, CACHE_TIMESTAMP
    from datetime import datetime, timedelta
    
    # Refresh cache every 10 minutes or if empty
    if not CACHE_TIMESTAMP or datetime.now(timezone.utc) - CACHE_TIMESTAMP > timedelta(minutes=10) or not USER_NAME_CACHE:
        try:
            all_users = list(users.find({}, {"_id": 0, "email": 1, "name": 1}))
            USER_NAME_CACHE = {u["email"]: u.get("name", u["email"]) for u in all_users if u.get("email")}
            CACHE_TIMESTAMP = datetime.now(timezone.utc)
            logger.info(f"User name cache refreshed: {len(USER_NAME_CACHE)} users")
        except Exception as e:
            logger.error(f"Failed to refresh user name cache: {e}")
            return {email: email for email in user_emails}  # Fallback to emails
    
    return {email: USER_NAME_CACHE.get(email, email) for email in user_emails if email}

load_dotenv()

SERPAPI_KEY = os.getenv("SERPAPI_KEY")

app = FastAPI()

# Get the absolute path to the static folder
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
STATIC_DIR = os.path.join(BASE_DIR, "static")

# Mount static files for serving HTML/CSS/JS
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

# CORS configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Base response format
base_response = {
    "message": "",
    "success": False,
    "data": None,
    "status_code": status.HTTP_400_BAD_REQUEST
}

def validate_password(password: str) -> bool:
    if len(password) < 8:
        return False
    if not re.search(r"[A-Z]", password):
        return False
    if not re.search(r"[a-z]", password):
        return False
    if not re.search(r"\d", password):
        return False
    if not re.search(r"[!@#$%^&*()]", password):
        return False
    return True


# ─────────────────────────────────────────────────────────────────────────────
#  AUTH ENDPOINTS
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/register")
async def register(user: UserRegister):
    logger.info(f"Register attempt for email: {user.email}")
    try:
        if not validate_password(user.password):
            return JSONResponse(
                content={"message": "Invalid password format", "success": False},
                status_code=status.HTTP_400_BAD_REQUEST
            )
        if user.password != user.confirm_password:
            return JSONResponse(
                content={"message": "Passwords do not match", "success": False},
                status_code=status.HTTP_400_BAD_REQUEST
            )
        existing_user = users.find_one({"email": user.email})
        if existing_user:
            return JSONResponse(
                content={"message": "Email already registered", "success": False},
                status_code=status.HTTP_400_BAD_REQUEST
            )
        hashed_password = bcrypt.hashpw(
            user.password.encode("utf-8"), bcrypt.gensalt()
        ).decode("utf-8")
        users.insert_one({
            "name": user.name,
            "email": user.email,
            "password_hash": hashed_password,
            "created_at": datetime.now(timezone.utc),
        })
        return JSONResponse(
            content={"message": "User registered successfully", "success": True, "data": {"email": user.email}},
            status_code=status.HTTP_201_CREATED
        )
    except Exception as e:
        logger.error(f"Error in register: {str(e)}")
        return JSONResponse(
            content={"message": f"Server error: {str(e)}", "success": False},
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@app.post("/api/login")
def login(user: UserLogin):
    logger.info(f"Login attempt for email: {user.email}")
    db_user = users.find_one({"email": user.email})
    if not db_user:
        return JSONResponse({"message": "Invalid email or password", "success": False}, status_code=401)
    password = user.password
    password_hash = db_user.get("password_hash")
    if not isinstance(password, str) or not isinstance(password_hash, str):
        return JSONResponse({"message": "Invalid email or password", "success": False}, status_code=401)
    if not bcrypt.checkpw(password.encode("utf-8"), password_hash.encode("utf-8")):
        return JSONResponse({"message": "Invalid email or password", "success": False}, status_code=401)
    token = create_jwt({"sub": user.email})
    return JSONResponse(
        {"message": "Logged in successfully", "success": True, "data": {"access_token": token}},
        status_code=200
    )


@app.get("/api/me")
async def get_me(emp_id: str = Depends(get_current_user)):
    try:
        user = users.find_one({"email": emp_id})
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        response = base_response.copy()
        response.update({
            "message": "User info retrieved successfully",
            "success": True,
            "data": {"email": user["email"], "name": user.get("name", "Unknown")},
            "status_code": status.HTTP_200_OK
        })
        return JSONResponse(content=response, status_code=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f"Error in get_me: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/logout")
async def logout(emp_id: str = Depends(get_current_user)):
    return JSONResponse(
        content={"message": "Logged out successfully", "success": True, "data": None},
        status_code=200
    )


# ─────────────────────────────────────────────────────────────────────────────
#  AUDIT SECTIONS
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/get-sections")
async def get_sections(emp_id: str = Depends(get_current_user)):
    try:
        today = datetime.now(timezone.utc).date().isoformat()
        # Check temp collection first, then submitted collection
        audit = temp_audit_data_collection.find_one({"user_id": emp_id, "date": today})
        if not audit:
            audit = audit_data_collection.find_one({"user_id": emp_id, "date": today})
        
        section_keys = [
            "general_report", "stock_reconciliation",
            "observations_on_stacking", "observations_on_warehouse_operations",
            "observations_on_warehouse_record_keeping", "observations_on_wh_infrastructure",
            "observations_on_quality_operation", "checklist_wrt_exchange_circular_mentha_oil",
            "checklist_wrt_exchange_circular_metal", "checklist_wrt_exchange_circular_cotton_bales",
            "signature", "photo"
        ]
        completion_status = {
            k: (audit["completion_status"].get(k, False) if audit and "completion_status" in audit else False)
            for k in section_keys
        }
        response = base_response.copy()
        response.update({
            "message": "Sections retrieved successfully",
            "success": True,
            "data": {"completion_status": completion_status},
            "status_code": status.HTTP_200_OK
        })
        return JSONResponse(content=response, status_code=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f"Error in get_sections: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/get-section/{section_name}")
async def get_section(section_name: str, emp_id: str = Depends(get_current_user)):
    try:
        today = datetime.now(timezone.utc).date().isoformat()
        audit = temp_audit_data_collection.find_one({"user_id": emp_id, "date": today})
        section_data = audit["sections"].get(section_name, {}) if audit and "sections" in audit else {}
        response = base_response.copy()
        response.update({
            "message": f"Section {section_name} retrieved successfully",
            "success": True,
            "data": {"section_data": section_data},
            "status_code": status.HTTP_200_OK
        })
        return JSONResponse(content=response, status_code=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f"Error in get_section: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/save-section")
async def save_section(request: Request, emp_id: str = Depends(get_current_user)):
    try:
        body = await request.json()
        section = body.get("section")
        data = body.get("data")
        date = body.get("date")
        if not section or not data or not date:
            return JSONResponse(
                content={"message": "Missing required fields (section, data, date)", "success": False},
                status_code=400
            )

        # --- GridFS: extract photo and store separately ---
        if section == "photo" and "photo" in data and data["photo"].startswith("data:image"):
            try:
                header, b64data = data["photo"].split(",", 1)
                img_bytes = base64.b64decode(b64data)
                file_id = fs.put(
                    img_bytes,
                    filename=f"{emp_id}_{date}_photo.png",
                    content_type="image/png",
                    metadata={"user_id": emp_id, "date": date}
                )
                data["photo"] = None          # don't store base64 in document
                data["photo_file_id"] = str(file_id)  # store GridFS reference
            except Exception as img_err:
                logger.warning(f"GridFS photo store failed, falling back to base64: {img_err}")

        audit = temp_audit_data_collection.find_one({"user_id": emp_id, "date": date})
        if not audit:
            audit = {
                "user_id": emp_id, "date": date, "sections": {},
                "completion_status": {}, "submitted_by": emp_id,
                "submitted_at": datetime.now(timezone.utc)
            }
        audit["sections"][section] = data
        audit["completion_status"][section] = True
        if audit.get("_id"):
            temp_audit_data_collection.update_one(
                {"_id": audit["_id"]},
                {"$set": {"sections": audit["sections"], "completion_status": audit["completion_status"]}}
            )
        else:
            temp_audit_data_collection.insert_one(audit)
        response = base_response.copy()
        response.update({
            "message": f"Section {section} saved successfully",
            "success": True,
            "data": {"completion_status": audit["completion_status"]},
            "status_code": 200
        })
        return JSONResponse(content=response, status_code=200)
    except Exception as e:
        logger.error(f"Error in save_section: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/get-photo/{file_id}")
async def get_photo(file_id: str, emp_id: str = Depends(get_current_user)):
    """Serve a photo stored in GridFS by its file ID."""
    try:
        oid = ObjectId(file_id)
        if not fs.exists(oid):
            return JSONResponse({"message": "Photo not found", "success": False}, status_code=404)
        grid_out = fs.get(oid)
        data = grid_out.read()
        return StreamingResponse(io.BytesIO(data), media_type="image/png")
    except Exception as e:
        logger.error(f"get_photo error: {e}")
        return JSONResponse({"message": str(e), "success": False}, status_code=500)


@app.post("/api/submit-audit")
async def submit_audit(emp_id: str = Depends(get_current_user)):
    try:
        today = datetime.now(timezone.utc).date().isoformat()
        temp_audit = temp_audit_data_collection.find_one({"user_id": emp_id, "date": today})
        if not temp_audit:
            raise HTTPException(status_code=404, detail="No audit data found to submit")
        
        # Check only the expected checklist sections (not stock_count or other keys)
        completion = temp_audit.get("completion_status", {})
        expected = [
            "general_report", "stock_reconciliation",
            "observations_on_stacking", "observations_on_warehouse_operations",
            "observations_on_warehouse_record_keeping", "observations_on_wh_infrastructure",
            "observations_on_quality_operation", "checklist_wrt_exchange_circular_mentha_oil",
            "checklist_wrt_exchange_circular_metal", "checklist_wrt_exchange_circular_cotton_bales",
            "signature", "photo"
        ]
        if not all(completion.get(s, False) for s in expected):
            raise HTTPException(status_code=400, detail="Not all sections are completed")
        
        # Update submitted_at timestamp to the actual submission time
        temp_audit["submitted_at"] = datetime.now(timezone.utc)
        
        # Remove the _id from temp_audit to avoid duplicate key error
        temp_id = temp_audit.pop("_id")
        
        # Insert into permanent collection (MongoDB will generate new _id)
        result = audit_data_collection.insert_one(temp_audit)
        
        # Delete from temp collection using the original _id
        temp_audit_data_collection.delete_one({"_id": temp_id})
        
        response = base_response.copy()
        response.update({
            "message": "Audit submitted successfully", "success": True,
            "data": {"submitted": True, "audit_id": str(result.inserted_id)},
            "status_code": 200
        })
        return JSONResponse(content=response, status_code=200)
    except Exception as e:
        logger.error(f"Error in submit_audit: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/clear-sections")
async def clear_sections(emp_id: str = Depends(get_current_user)):
    try:
        today = datetime.now(timezone.utc).date().isoformat()
        temp_audit_data_collection.delete_one({"user_id": emp_id, "date": today})
        return JSONResponse({"message": "Sections cleared", "success": True}, status_code=200)
    except Exception as e:
        return JSONResponse({"message": str(e), "success": False}, status_code=500)

# ═════════════════════════════════════════════════════════════════════════════
# ADD THIS ENDPOINT TO YOUR main.py (before the existing /api/get-location)
# ═════════════════════════════════════════════════════════════════════════════

@app.get("/api/get-ip-location")
def get_ip_location():
    """
    Fetch user's location via IP address from backend (avoids CORS issues).
    This solves the CORS problem when geolocation API is not available in production.
    """
    services = [
        ("https://freeipapi.com/api/json", lambda d: {"lat": d.get("latitude"), "lon": d.get("longitude")} if d.get("latitude") else None),
        ("https://ipapi.co/json/", lambda d: {"lat": d.get("latitude"), "lon": d.get("longitude")} if d.get("latitude") else None),
        ("https://ip-api.com/json/?fields=lat,lon,status", lambda d: {"lat": d.get("lat"), "lon": d.get("lon")} if d.get("status") == "success" else None),
    ]
    
    for service_url, parse_fn in services:
        try:
            res = requests.get(service_url, timeout=5)
            if res.status_code == 200:
                data = res.json()
                coords = parse_fn(data)
                if coords and coords["lat"] and coords["lon"]:
                    logger.info(f"IP geolocation via {service_url}: {coords}")
                    return {"latitude": coords["lat"], "longitude": coords["lon"], "success": True}
        except Exception as e:
            logger.warning(f"IP geolocation service {service_url} failed: {e}")
    
    logger.error("All IP geolocation services failed")
    return {"success": False, "error": "Could not determine IP location"}

# ─────────────────────────────────────────────────────────────────────────────
#  LOCATION
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/get-location")
def get_location(lat: float = Query(...), lon: float = Query(...)):
    serp_url = f"https://serpapi.com/search?engine=google_maps&q={lat},{lon}&type=search&api_key={SERPAPI_KEY}"
    try:
        serp_res = requests.get(serp_url, timeout=6)
        serp_data = serp_res.json()
        if serp_res.status_code == 200 and "search_metadata" in serp_data:
            place_result = serp_data.get("place_results", {})
            plus_code = place_result.get("plus_code") or "N/A"
            address = place_result.get("title") or "N/A"
            maps_url = serp_data.get("search_metadata", {}).get("google_maps_url")
            if maps_url:
                return {"source": "serpapi", "latitude": lat, "longitude": lon, "plus_code": plus_code, "address": address, "maps_url": maps_url}
        raise Exception("SerpApi failed or incomplete")
    except Exception as e:
        print(f"⚠️ SerpApi failed: {e}")
        osm_url = f"https://nominatim.openstreetmap.org/reverse?format=jsonv2&lat={lat}&lon={lon}"
        osm_res = requests.get(osm_url, headers={"User-Agent": "audit-app"})
        osm_data = osm_res.json()
        address = osm_data.get("display_name", "Address not found")
        plus_code = address
        maps_url = f"https://www.google.com/maps/search/{lat}%2C{lon}?hl=en"
        return {"source": "osm", "latitude": lat, "longitude": lon, "plus_code": plus_code, "address": address, "maps_url": maps_url}


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL GENERATION HELPER
# ─────────────────────────────────────────────────────────────────────────────

def _adjust_ws(ws, widths):
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


async def generate_checklist_excel_bytes(emp_id: str, audit_data: dict) -> bytes:
    """Checklist-only Excel – NO Stock Count sheet."""
    wb = Workbook()
    wb.remove(wb.active)
    sections = audit_data.get("sections", {})

    # General Report
    ws = wb.create_sheet("General Report")
    ws.append(["Field", "Value"])
    gr = sections.get("general_report", {})
    if gr:
        for k, v in gr.items():
            if k == "previous_audits" and isinstance(v, list):
                ws.append(["Previous Audits", ""])
                for i, rec in enumerate(v, start=1):
                    ws.append([f"  Record {i} – Date", str(rec.get("date", ""))])
                    ws.append([f"  Record {i} – Auditor Name", str(rec.get("auditor_name", ""))])
                    ws.append([f"  Record {i} – Auditor Type", str(rec.get("auditor_type", ""))])
                    if rec.get("agency_name"):
                        ws.append([f"  Record {i} – Agency Name", str(rec.get("agency_name", ""))])
            else:
                ws.append([k.replace("_", " ").title(), str(v)])
    else:
        ws.append(["No general report saved.", ""])
    _adjust_ws(ws, [40, 30])

    # Stock Reconciliation
    ws = wb.create_sheet("Stock Reconciliation")
    ws.append(["Commodity Name", "Stock Type", "Qty as per MCXCCL", "Qty as per Registered", "Qty as per Physical", "Difference", "Remarks"])
    stock = sections.get("stock_reconciliation", {}).get("commodities", [])
    if stock:
        for item in stock:
            ws.append([item.get("commodity_name",""), item.get("commodity",""), item.get("qty_mcxccl",""), item.get("qty_registered",""), item.get("qty_physical",""), item.get("difference",""), item.get("remarks","")])
    else:
        ws.append(["No stock data.", "", "", "", "", "", ""])
    _adjust_ws(ws, [20, 20, 20, 20, 20, 20, 30])

    # Question-based sections
    q_sections = [
        ("observations_on_stacking", "Observations on Stacking"),
        ("observations_on_warehouse_operations", "Observations on WH Operations"),
        ("observations_on_warehouse_record_keeping", "Observations on WH Record Keeping"),
        ("observations_on_wh_infrastructure", "Observations on WH Infrastructure"),
        ("observations_on_quality_operation", "Observations on Quality Operation"),
        ("checklist_wrt_exchange_circular_mentha_oil", "Checklist Mentha Oil"),
        ("checklist_wrt_exchange_circular_metal", "Checklist Metals"),
        ("checklist_wrt_exchange_circular_cotton_bales", "Checklist Cotton Bales"),
    ]
    for key, title in q_sections:
        ws = wb.create_sheet(title)
        ws.append(["Question", "Yes/No", "Remarks"])
        qlist = sections.get(key, {}).get("questions", [])
        if qlist:
            for idx, q in enumerate(qlist, start=1):
                ws.append([f"{idx}. {q.get('question', f'Question {idx}').strip()}", q.get("answer","").strip(), q.get("remarks","").strip()])
        else:
            ws.append(["No data saved.", "", ""])
        _adjust_ws(ws, [60, 10, 30])

    # Signature
    ws = wb.create_sheet("Signature")
    sig = sections.get("signature", {}).get("signature")
    if sig:
        try:
            img_data = re.sub("^data:image/.+;base64,", "", sig)
            img_bytes = io.BytesIO(base64.b64decode(img_data))
            img = Image(img_bytes)
            img.width = 250; img.height = 150
            ws["A1"] = "Signature captured during the audit"
            ws.row_dimensions[1].height = 18
            ws.row_dimensions[2].height = 8
            ws.add_image(img, "A3")
        except Exception as e:
            ws["A1"] = f"Unable to embed signature: {e}"
    else:
        ws["A1"] = "Signature not found."
    ws.column_dimensions["A"].width = 60

    # Photo
    # Photo section - support both single photo (legacy) and multiple photos
    ws = wb.create_sheet("Photo")
    photo_section = sections.get("photo", {})
    
    # Check if it's the new multi-photo format
    photos_list = photo_section.get("photos", [])
    
    # Handle legacy single photo format
    if not photos_list and photo_section.get("photo"):
        photos_list = [{
            "photo": photo_section.get("photo"),
            "maps_url": photo_section.get("maps_url", ""),
            "timestamp": datetime.now().isoformat(),
            "location_text": "Legacy photo"
        }]
    
    if photos_list:
        # Multi-photo format
        ws.append(["Photo #", "Timestamp", "Location", "Google Maps Link"])
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 60  # Wide column for images
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 50
        
        for idx, photo_item in enumerate(photos_list, 1):
            timestamp = photo_item.get("timestamp", "N/A")
            location = photo_item.get("location_text", "N/A")
            maps_url = photo_item.get("maps_url", "")
            
            # Add data row
            data_row = idx + 1
            ws[f"A{data_row}"] = f"Photo {idx}"
            ws[f"C{data_row}"] = location
            ws[f"D{data_row}"] = maps_url if maps_url else "N/A"
            
            # Try to embed the image
            try:
                photo_data = photo_item.get("photo", "")
                if photo_data:
                    # Handle base64 encoded image
                    if photo_data.startswith("data:image"):
                        img_data = re.sub("^data:image/.+;base64,", "", photo_data)
                        photo_bytes = base64.b64decode(img_data)
                    else:
                        photo_bytes = base64.b64decode(photo_data)
                    
                    # Create PIL image and resize if needed
                    from PIL import Image as PILImage
                    pil_img = PILImage.open(io.BytesIO(photo_bytes))
                    
                    # Resize to reasonable size for Excel
                    max_size = (400, 300)
                    pil_img.thumbnail(max_size, PILImage.Resampling.LANCZOS)
                    
                    # Save to bytes
                    img_buf = io.BytesIO()
                    pil_img.save(img_buf, format='PNG')
                    img_buf.seek(0)
                    
                    # Create Excel image and add to worksheet
                    excel_img = Image(img_buf)
                    excel_img.width = 400
                    excel_img.height = 300
                    
                    # Position in column B
                    ws.add_image(excel_img, f"B{data_row}")
                    ws.row_dimensions[data_row].height = 225  # Adjust row height for image
                    
            except Exception as e:
                logger.warning(f"Failed to embed photo {idx}: {e}")
                ws[f"B{data_row}"] = f"[Image embedding failed: {str(e)}]"
    else:
        ws.append(["No photos captured"])
        ws.column_dimensions["A"].width = 30

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def generate_stock_count_excel_bytes(audit_data: dict) -> bytes:
    """Stock-count-only Excel – no checklist sheets."""
    sc_data = audit_data.get("stock_count_data", [])
    df = pd.DataFrame(sc_data) if sc_data else pd.DataFrame(columns=["sheet_name", "item_name", "item_code", "qty", "physical_amount", "remarks"])
    columns = ["sheet_name", "item_name", "item_code", "qty", "physical_amount", "remarks"]
    available_cols = [c for c in columns if c in df.columns]
    df = df[available_cols]
    df.rename(columns={"sheet_name": "Sheet Name", "item_name": "Item Name",
                       "item_code": "Item Code", "qty": "Expected Qty",
                       "physical_amount": "Physical Count", "remarks": "Remarks"}, inplace=True)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Stock Count')
    return output.getvalue()


# keep old name as alias so existing callers still work
async def generate_excel_bytes(emp_id: str, audit_data: dict) -> bytes:
    return await generate_checklist_excel_bytes(emp_id, audit_data)



# ─────────────────────────────────────────────────────────────────────────────
#  EXPORT EXCEL
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/export-excel")
async def export_excel(emp_id: str = Depends(get_current_user)):
    try:
        today = datetime.now(timezone.utc).date().isoformat()
        
        # Look specifically for checklist audit (has 'sections' field, not just stock_count)
        audit_data = temp_audit_data_collection.find_one({
            "user_id": emp_id, 
            "date": today,
            "sections": {"$exists": True}
        }, sort=[("submitted_at", -1)])
        
        if not audit_data:
            audit_data = audit_data_collection.find_one({
                "user_id": emp_id, 
                "date": today,
                "sections": {"$exists": True}
            }, sort=[("submitted_at", -1)])
        
        if not audit_data:
            return JSONResponse({"message": "No checklist audit data for today", "success": False}, status_code=404)
        
        completion = audit_data.get("completion_status", {})
        expected = [
            "general_report", "stock_reconciliation",
            "observations_on_stacking", "observations_on_warehouse_operations",
            "observations_on_warehouse_record_keeping", "observations_on_wh_infrastructure",
            "observations_on_quality_operation", "checklist_wrt_exchange_circular_mentha_oil",
            "checklist_wrt_exchange_circular_metal", "checklist_wrt_exchange_circular_cotton_bales",
            "signature", "photo"
        ]
        
        # Check which sections are missing
        missing_sections = [s for s in expected if not completion.get(s, False)]
        if missing_sections:
            logger.error(f"Export failed - Missing sections: {missing_sections}, completion_status: {completion}")
            return JSONResponse({
                "message": f"Complete all sections before exporting. Missing: {', '.join(missing_sections)}", 
                "success": False
            }, status_code=400)
        
        excel_bytes = await generate_excel_bytes(emp_id, audit_data)
        filename = f"audit_{emp_id}_{today}.xlsx"
        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename*=UTF-8\'\'{filename}'}
        )
    except Exception as e:
        logger.error(f"Export-excel error: {e}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  SEND EMAIL
# ─────────────────────────────────────────────────────────────────────────────

def send_smtp_message_sync(msg, mail_username, mail_password):
    last_err = None
    for attempt in range(1, 4):
        try:
            logger.info(f"SMTP send attempt {attempt}/3...")
            with smtplib.SMTP("smtp.gmail.com", 587, timeout=15) as smtp:
                smtp.ehlo()
                smtp.starttls()
                smtp.ehlo()
                smtp.login(mail_username, mail_password)
                smtp.send_message(msg)
            logger.info("SMTP send successful!")
            return True
        except smtplib.SMTPAuthenticationError as e:
            logger.error(f"SMTP auth failed: {e}")
            raise
        except Exception as e:
            last_err = e
            logger.warning(f"SMTP attempt {attempt} failed: {e}. Traceback:\n{traceback.format_exc()}")
            if attempt < 3:
                time.sleep(1.5)
    if last_err:
        raise last_err


def send_email_notification(to_emails: list, subject: str, body: str, attachments: list = None):
    """
    Helper function to send emails with attachments.
    
    Args:
        to_emails: List of recipient email addresses
        subject: Email subject
        body: HTML email body
        attachments: List of tuples (filename, file_bytes, mime_subtype)
    
    Returns:
        tuple: (success: bool, message: str)
    """
    try:
        mail_username = os.getenv("MAIL_USERNAME")
        mail_password = os.getenv("MAIL_PASSWORD")
        
        if not mail_username or not mail_password:
            return False, "Email credentials not configured"
        
        # Create message
        msg = MIMEMultipart()
        msg["From"] = mail_username
        msg["To"] = ", ".join(to_emails)
        msg["Subject"] = subject
        
        # Attach HTML body
        msg.attach(MIMEText(body, "html"))
        
        # Attach files if provided
        if attachments:
            for filename, file_bytes, mime_subtype in attachments:
                msg.add_attachment(
                    file_bytes,
                    maintype="application",
                    subtype=mime_subtype,
                    filename=filename
                )
        
        # Send email
        send_smtp_message_sync(msg, mail_username, mail_password)
        
        return True, "Email sent successfully"
        
    except smtplib.SMTPAuthenticationError as auth_err:
        return False, f"Email authentication failed: {str(auth_err)}"
    except Exception as e:
        logger.error(f"Email send error: {e}")
        return False, f"Failed to send email: {str(e)}"


# ─────────────────────────────────────────────────────────────────────────────
#  SEND EMAIL API ENDPOINT
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/send-email")
async def send_email(
    to_email: str = Form(...),
    email_subject: Optional[str] = Form(default=None),  # Custom subject
    email_body: Optional[str] = Form(default=None),  # Custom body/message
    attachment: Optional[UploadFile] = File(default=None),  # Keep for backward compatibility
    attachments: List[UploadFile] = File(default=[]),  # New: multiple files support
    email_type: str = Form(default="checklist"),   # "checklist" or "stock-count"
    audit_id: Optional[str] = Form(default=None),
    emp_id: str = Depends(get_current_user)
):
    """
    email_type='checklist'  → validates checklist sections, attaches checklist Excel.
    email_type='stock-count' → validates stock count submission, attaches stock-count Excel.
    """
    try:
        # Get user name for email subject
        user = users.find_one({"email": emp_id})
        user_name = user.get("name", "Unknown") if user else "Unknown"
        
        today = datetime.now(timezone.utc).date().isoformat()
        
        # Resolve audit_data based on audit_id if provided, else fallback to today
        if audit_id:
            from bson import ObjectId
            audit_data = audit_data_collection.find_one({"user_id": emp_id, "_id": ObjectId(audit_id)})
            if not audit_data:
                audit_data = temp_audit_data_collection.find_one({"user_id": emp_id, "_id": ObjectId(audit_id)})
        else:
            # When no audit_id, find the most recent audit of the correct type for today
            if email_type == "stock-count":
                # Find audit with stock_count data
                audit_data = audit_data_collection.find_one({
                    "user_id": emp_id, 
                    "date": today,
                    "stock_count_data": {"$exists": True}
                }, sort=[("submitted_at", -1)])
                if not audit_data:
                    audit_data = temp_audit_data_collection.find_one({
                        "user_id": emp_id, 
                        "date": today,
                        "stock_count_data": {"$exists": True}
                    }, sort=[("submitted_at", -1)])
            else:
                # Find audit with checklist sections data
                audit_data = audit_data_collection.find_one({
                    "user_id": emp_id, 
                    "date": today,
                    "sections": {"$exists": True}
                }, sort=[("submitted_at", -1)])
                if not audit_data:
                    audit_data = temp_audit_data_collection.find_one({
                        "user_id": emp_id, 
                        "date": today,
                        "sections": {"$exists": True}
                    }, sort=[("submitted_at", -1)])

        if not audit_data:
            return JSONResponse({"message": "No audit data found to email", "success": False}, status_code=404)

        target_date = audit_data.get("date", today)

        if email_type == "stock-count":
            # Only validate that stock count was submitted
            if not audit_data.get("completion_status", {}).get("stock_count", False):
                return JSONResponse({"message": "Please submit stock count before sending email", "success": False}, status_code=400)
            if not audit_data.get("stock_count_data"):
                return JSONResponse({"message": "No stock count data found", "success": False}, status_code=400)
            # Build stock-count-only Excel
            data = audit_data["stock_count_data"]
            df = pd.DataFrame(data)
            columns = ["sheet_name", "item_name", "item_code", "qty", "physical_amount", "remarks"]
            available_cols = [c for c in columns if c in df.columns]
            df = df[available_cols]
            df.rename(columns={"sheet_name": "Sheet Name", "item_name": "Item Name",
                                "item_code": "Item Code", "qty": "Expected Qty",
                                "physical_amount": "Physical Count", "remarks": "Remarks"}, inplace=True)
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Stock Count')
            excel_bytes = output.getvalue()
            excel_name = f"Stock_Count_{target_date}.xlsx"
            subject = f"Stock Count Report – {target_date} – {user_name}"
            body = f"Dear Manager,\n\nPlease find the Stock Count report attached.\n\nRegards,\n{user_name}\nAudit App"
        else:
            # Checklist — validate all checklist sections
            completion = audit_data.get("completion_status", {})
            expected = [
                "general_report", "stock_reconciliation",
                "observations_on_stacking", "observations_on_warehouse_operations",
                "observations_on_warehouse_record_keeping", "observations_on_wh_infrastructure",
                "observations_on_quality_operation", "checklist_wrt_exchange_circular_mentha_oil",
                "checklist_wrt_exchange_circular_metal", "checklist_wrt_exchange_circular_cotton_bales",
                "signature", "photo"
            ]
            if not all(completion.get(s, False) for s in expected):
                return JSONResponse({"message": "Complete all checklist sections before sending email", "success": False}, status_code=400)
            excel_bytes = await generate_excel_bytes(emp_id, audit_data)
            excel_name = f"Checklist_Audit_{target_date}.xlsx"
            
            # Use custom subject/body if provided, otherwise use defaults
            if not email_subject:
                subject = f"Checklist Audit Report – {target_date} – {user_name}"
            else:
                subject = email_subject
                
            if not email_body:
                body = f"Dear Auditor Manager,\n\nPlease find the Checklist Audit report attached.\n\nRegards,\n{user_name}\nAudit App"
            else:
                body = email_body

        msg = EmailMessage()
        msg["Subject"] = subject if email_subject or audit_id or audit_data else (email_subject or "Documents from Audit App")
        msg["From"] = os.getenv("MAIL_USERNAME")
        msg["To"] = to_email
        msg["Cc"] = emp_id
        msg.set_content(body)

        # Attach user-uploaded files (support both single 'attachment' and multiple 'attachments')
        files_to_attach = []
        
        # Handle single file (backward compatibility)
        if attachment:
            files_to_attach.append(attachment)
        
        # Handle multiple files (new feature)
        if attachments:
            files_to_attach.extend(attachments)
        
        # Attach all user-uploaded files
        for uploaded_file in files_to_attach:
            file_bytes = await uploaded_file.read()
            file_name = uploaded_file.filename
            allowed_extensions = (".pdf", ".xlsx", ".xls", ".xlsb")
            
            if not file_name.lower().endswith(allowed_extensions):
                return JSONResponse({
                    "message": f"File '{file_name}' has invalid type. Only PDF or Excel files are allowed", 
                    "success": False
                }, status_code=400)
            
            # Determine MIME type
            if file_name.lower().endswith('.pdf'):
                subtype = "pdf"
            else:
                subtype = "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            
            msg.add_attachment(
                file_bytes, 
                maintype="application",
                subtype=subtype, 
                filename=file_name
            )

        # Always attach the generated excel sheet (if this is checklist/stock-count email with audit_id)
        if email_type in ['checklist', 'stock-count'] and (audit_id or audit_data):
            msg.add_attachment(
                excel_bytes, 
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=excel_name
            )

        mail_username = os.getenv("MAIL_USERNAME")
        mail_password = os.getenv("MAIL_PASSWORD")
        
        if not mail_username or not mail_password:
            return JSONResponse({
                "message": "Email credentials not configured. Please contact administrator.",
                "success": False
            }, status_code=500)

        try:
            await asyncio.to_thread(send_smtp_message_sync, msg, mail_username, mail_password)
        except smtplib.SMTPAuthenticationError as auth_err:
            logger.error(f"SMTP Authentication failed: {auth_err}")
            return JSONResponse({
                "message": "Email authentication failed. Please verify email credentials in settings.",
                "success": False
            }, status_code=500)
        
        return JSONResponse({"message": "Email sent successfully", "success": True}, status_code=200)
    except Exception as e:
        logger.error(f"Send-email error: {e}")
        return JSONResponse({"message": f"Failed to send email: {str(e)}", "success": False}, status_code=500)

# ─────────────────────────────────────────────────────────────────────────────
#  UPLOAD ITEM MASTER  –  NEW JSON endpoint (called by the wizard)
# ─────────────────────────────────────────────────────────────────────────────

class SheetItemsPayload(BaseModel):
    sheet_name: str
    items: List[dict]   # each: { item_code, item_name, qty }

class UploadItemsJsonPayload(BaseModel):
    sheets: List[SheetItemsPayload]


@app.post("/api/upload-items-json")
async def upload_items_json(
    payload: UploadItemsJsonPayload,
    emp_id: str = Depends(get_current_user)
):
    """
    Accept pre-parsed item data from the frontend wizard.
    Each sheet's items are stored with a sheet_name field so the
    stock-count view can group / filter by sheet.

    All existing items are replaced on each call.
    """
    try:
        logger.info(f"upload-items-json called by {emp_id}, sheets={[s.sheet_name for s in payload.sheets]}")

        all_items = []
        sheet_summary = []

        for sheet in payload.sheets:
            sheet_name = sheet.sheet_name.strip()
            valid_items = []

            for raw in sheet.items:
                item_code = str(raw.get("item_code", "")).strip()
                item_name = str(raw.get("item_name", "")).strip()
                qty       = str(raw.get("qty", "")).strip()
                extra_col = str(raw.get("extra_col", "")).strip()

                # Skip blank / sentinel rows
                if not item_code or not item_name:
                    continue
                if item_code.lower() in ("nan", "none", "item code", ""):
                    continue
                if item_name.lower() in ("nan", "none", "item name", ""):
                    continue

                valid_items.append({
                    "item_code":   item_code,
                    "item_name":   item_name,
                    "qty":         qty,
                    "extra_col":   extra_col,
                    "sheet_name":  sheet_name,
                    "uploaded_by": emp_id,
                    "uploaded_at": datetime.now(timezone.utc),
                })

            all_items.extend(valid_items)
            sheet_summary.append({"sheet": sheet_name, "count": len(valid_items)})
            logger.info(f"  Sheet '{sheet_name}': {len(valid_items)} valid items")

        if not all_items:
            return JSONResponse(
                {"message": "No valid items found in the uploaded data. Check that Item Code and Item Name columns are not empty.", "success": False},
                status_code=400
            )

        # Store this upload in history
        upload_history_collection.insert_one({
            "uploaded_by": emp_id,
            "uploaded_at": datetime.now(timezone.utc),
            "total_items": len(all_items),
            "sheets": sheet_summary
        })

        # Replace only the sheets that were uploaded — leave other sheets untouched
        uploaded_sheet_names = [s.sheet_name.strip() for s in payload.sheets]
        item_master_collection.delete_many({"sheet_name": {"$in": uploaded_sheet_names}})
        if all_items:
            item_master_collection.insert_many(all_items)

        logger.info(f"Inserted {len(all_items)} items from {len(payload.sheets)} sheet(s)")

        return JSONResponse(
            {
                "message": f"Successfully uploaded {len(all_items)} items from {len(payload.sheets)} sheet(s)",
                "success": True,
                "data": {
                    "total_count": len(all_items),
                    "sheets": sheet_summary,
                    "sample_items": [
                        {"item_code": i["item_code"], "item_name": i["item_name"], "sheet_name": i["sheet_name"]}
                        for i in all_items[:5]
                    ]
                }
            },
            status_code=200
        )

    except Exception as e:
        logger.error(f"upload-items-json error: {e}")
        import traceback; logger.error(traceback.format_exc())
        return JSONResponse({"message": f"Failed to upload items: {str(e)}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  LEGACY FILE-UPLOAD ENDPOINT  (kept for backward compatibility)
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/upload-items")
async def upload_items(
    file: UploadFile = File(...),
    emp_id: str = Depends(get_current_user)
):
    """Original file-upload endpoint – still works but wizard uses /api/upload-items-json instead."""
    try:
        fname = file.filename.lower()
        if not any(fname.endswith(ext) for ext in ('.xlsx', '.xls', '.xlsb', '.csv')):
            return JSONResponse({"message": "Only Excel (.xlsx/.xls/.xlsb) or CSV files are allowed", "success": False}, status_code=400)
        contents = await file.read()
        try:
            if fname.endswith('.csv'):
                df = pd.read_csv(io.BytesIO(contents))
            elif fname.endswith('.xlsb'):
                df = pd.read_excel(io.BytesIO(contents), engine='pyxlsb')
            else:
                df = pd.read_excel(io.BytesIO(contents))
        except Exception as parse_error:
            return JSONResponse({"message": f"Could not parse file: {str(parse_error)}", "success": False}, status_code=400)

        item_code_col = next((c for c in df.columns if 'item code' in str(c).lower()), None)
        item_name_col = next((c for c in df.columns if 'item name' in str(c).lower()), None)
        if not item_code_col or not item_name_col:
            return JSONResponse({"message": "Could not find 'Item Code' and 'Item Name' columns.", "success": False}, status_code=400)

        items = []
        for _, row in df.iterrows():
            code = str(row[item_code_col]).strip() if pd.notna(row[item_code_col]) else ""
            name = str(row[item_name_col]).strip() if pd.notna(row[item_name_col]) else ""
            if code and name and code.lower() not in ['nan','none','','item code'] and name.lower() not in ['nan','none','','item name']:
                items.append({"item_code": code, "item_name": name, "qty": "", "sheet_name": "Default", "uploaded_by": emp_id, "uploaded_at": datetime.now(timezone.utc)})

        if not items:
            return JSONResponse({"message": "No valid items found in the file.", "success": False}, status_code=400)

        item_master_collection.delete_many({})
        item_master_collection.insert_many(items)
        return JSONResponse({"message": f"Successfully uploaded {len(items)} items", "success": True, "data": {"count": len(items)}}, status_code=200)

    except Exception as e:
        logger.error(f"Upload items error: {e}")
        return JSONResponse({"message": f"Failed to upload items: {str(e)}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  GET ITEMS  (stock count – now returns sheet_name + qty)
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/get-items")
async def get_items(
    search: str = Query(None),
    page: int = Query(1),
    limit: int = Query(50),
    emp_id: str = Depends(get_current_user)
):
    try:
        query = {}
        if search:
            query = {"$or": [
                {"item_code": {"$regex": search, "$options": "i"}},
                {"item_name": {"$regex": search, "$options": "i"}}
            ]}

        # Calculate pagination
        skip = (page - 1) * limit
        total = item_master_collection.count_documents(query)
        
        # Fetch from master — include sheet_name and qty now with pagination
        items = list(item_master_collection.find(
            query,
            {"_id": 0, "item_code": 1, "item_name": 1, "sheet_name": 1, "qty": 1}
        ).skip(skip).limit(limit))

        # Get existing stock-count data for this user today
        today = datetime.now(timezone.utc).date().isoformat()
        audit = temp_audit_data_collection.find_one({"user_id": emp_id, "date": today})
        stock_count_lookup = {}
        if audit and "stock_count_data" in audit:
            for sc in audit["stock_count_data"]:
                stock_count_lookup[sc["item_code"]] = {
                    "physical_amount": sc.get("physical_amount", ""),
                    "remarks": sc.get("remarks", "")
                }

        # Merge
        for item in items:
            sc = stock_count_lookup.get(item["item_code"], {})
            item["physical_amount"] = sc.get("physical_amount", "")
            item["remarks"] = sc.get("remarks", "")
            item.setdefault("sheet_name", "")
            item.setdefault("qty", "")

        return JSONResponse(
            {
                "message": "Items retrieved successfully", 
                "success": True, 
                "data": {
                    "items": items,
                    "page": page,
                    "limit": limit,
                    "total": total,
                    "has_more": (page * limit) < total
                }
            },
            status_code=200
        )
    except Exception as e:
        logger.error(f"Get items error: {e}")
        return JSONResponse({"message": f"Failed to get items: {str(e)}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  SAVE STOCK COUNT ITEM  (now persists sheet_name)
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/save-stock-count-item")
async def save_stock_count_item(request: Request, emp_id: str = Depends(get_current_user)):
    try:
        body = await request.json()
        item_code      = body.get("item_code")
        item_name      = body.get("item_name")
        sheet_name     = body.get("sheet_name", "")
        physical_amount = body.get("physical_amount", "")
        remarks        = body.get("remarks", "")

        if not item_code or not item_name:
            return JSONResponse({"message": "Item code and name are required", "success": False}, status_code=400)

        # Fetch expected qty from item master
        master = item_master_collection.find_one({"item_code": item_code}, {"_id": 0, "qty": 1})
        qty = master.get("qty", "") if master else ""

        today = datetime.now(timezone.utc).date().isoformat()
        audit = temp_audit_data_collection.find_one({"user_id": emp_id, "date": today})
        if not audit:
            audit = {
                "user_id": emp_id, "date": today, "sections": {},
                "completion_status": {}, "stock_count_data": [],
                "submitted_by": emp_id, "submitted_at": datetime.now(timezone.utc)
            }
        if "stock_count_data" not in audit:
            audit["stock_count_data"] = []

        item_found = False
        for item in audit["stock_count_data"]:
            if item["item_code"] == item_code:
                item.update({"physical_amount": physical_amount, "remarks": remarks, "item_name": item_name, "sheet_name": sheet_name, "qty": qty})
                item_found = True
                break

        if not item_found:
            audit["stock_count_data"].append({
                "item_code": item_code, "item_name": item_name,
                "sheet_name": sheet_name, "qty": qty,
                "physical_amount": physical_amount, "remarks": remarks
            })

        if audit.get("_id"):
            temp_audit_data_collection.update_one(
                {"_id": audit["_id"]},
                {"$set": {"stock_count_data": audit["stock_count_data"]}}
            )
        else:
            temp_audit_data_collection.insert_one(audit)

        return JSONResponse({"message": "Stock count item saved successfully", "success": True}, status_code=200)

    except Exception as e:
        logger.error(f"Save stock count item error: {e}")
        return JSONResponse({"message": f"Failed to save: {str(e)}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  SUBMIT STOCK COUNT
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/submit-stock-count")
async def submit_stock_count(emp_id: str = Depends(get_current_user)):
    try:
        today = datetime.now(timezone.utc).date().isoformat()
        audit = temp_audit_data_collection.find_one({"user_id": emp_id, "date": today})
        if not audit:
            return JSONResponse({"message": "No stock count data found", "success": False}, status_code=404)
        if not audit.get("stock_count_data"):
            return JSONResponse({"message": "Please count at least one item before submitting", "success": False}, status_code=400)
        
        # Mark stock count as complete in temp collection
        temp_audit_data_collection.update_one(
            {"_id": audit["_id"]},
            {"$set": {
                "completion_status.stock_count": True,
                "submitted_at": datetime.now(timezone.utc)
            }}
        )
        
        # Also create/update in audit_data_collection so it appears in Completed tab
        existing_audit = audit_data_collection.find_one({"user_id": emp_id, "date": today})
        if existing_audit:
            # Update existing audit with stock count data
            audit_data_collection.update_one(
                {"_id": existing_audit["_id"]},
                {"$set": {
                    "stock_count_data": audit.get("stock_count_data"),
                    "completion_status.stock_count": True,
                    "submitted_at": datetime.now(timezone.utc)
                }}
            )
            audit_id = str(existing_audit["_id"])
        else:
            # Create new audit record for stock count only
            audit["submitted_at"] = datetime.now(timezone.utc)
            result = audit_data_collection.insert_one(audit.copy())
            audit_id = str(result.inserted_id)
        
        return JSONResponse({"message": "Stock count submitted successfully", "success": True, "data": {"audit_id": audit_id}}, status_code=200)
    except Exception as e:
        logger.error(f"Submit stock count error: {e}")
        return JSONResponse({"message": f"Failed to submit: {str(e)}", "success": False}, status_code=500)
# ─────────────────────────────────────────────────────────────────────────────
#  EXPORT STOCK COUNT EXCEL
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/export-stock-count-excel")
async def export_stock_count_excel(emp_id: str = Depends(get_current_user)):
    try:
        today = datetime.now(timezone.utc).date().isoformat()
        audit = temp_audit_data_collection.find_one({"user_id": emp_id, "date": today})
        if not audit or not audit.get("stock_count_data"):
            return JSONResponse({"message": "No stock count data found", "success": False}, status_code=404)
        
        completion = audit.get("completion_status", {})
        if not completion.get("stock_count", False):
            return JSONResponse({"message": "Please submit stock count before exporting", "success": False}, status_code=400)

        data = audit["stock_count_data"]
        df = pd.DataFrame(data)
        
        columns = ["sheet_name", "item_name", "item_code", "qty", "physical_amount", "remarks"]
        available_cols = [c for c in columns if c in df.columns]
        df = df[available_cols]
        df.rename(columns={
            "sheet_name": "Sheet Name", 
            "item_name": "Item Name", 
            "item_code": "Item Code", 
            "qty": "Expected Qty", 
            "physical_amount": "Physical Count", 
            "remarks": "Remarks"
        }, inplace=True)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Stock Count')
        excel_bytes = output.getvalue()

        filename = f"stock_count_{emp_id}_{today}.xlsx"
        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
        )
    except Exception as e:
        logger.error(f"Export stock count excel error: {e}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  SEND STOCK COUNT EMAIL (from history/completed)
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/send-stock-count-email")
async def send_stock_count_email(
    audit_id: str = Form(...),
    to_email: str = Form(...),
    emp_id: str = Depends(get_current_user)
):
    """
    Send stock count report via email.
    Similar to checklist email but for stock count only.
    Called from History tab or after submission.
    """
    try:
        from bson import ObjectId
        
        # Get user name
        user = users.find_one({"email": emp_id})
        user_name = user.get("name", "Unknown") if user else "Unknown"
        
        # Get stock count data by audit_id
        audit_data = audit_data_collection.find_one({"user_id": emp_id, "_id": ObjectId(audit_id)})
        if not audit_data:
            return JSONResponse({"message": "Stock count not found", "success": False}, status_code=404)
        
        if not audit_data.get("stock_count_data"):
            return JSONResponse({"message": "No stock count data found", "success": False}, status_code=404)
        
        # Build Excel
        data = audit_data["stock_count_data"]
        df = pd.DataFrame(data)
        columns = ["sheet_name", "item_name", "item_code", "qty", "physical_amount", "remarks"]
        available_cols = [c for c in columns if c in df.columns]
        df = df[available_cols]
        df.rename(columns={
            "sheet_name": "Sheet Name",
            "item_name": "Item Name",
            "item_code": "Item Code",
            "qty": "Expected Qty",
            "physical_amount": "Physical Count",
            "remarks": "Remarks"
        }, inplace=True)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Stock Count')
        excel_bytes = output.getvalue()
        
        # Use original uploaded filename if available
        uploaded_filename = audit_data.get("uploaded_filename", "")
        if uploaded_filename:
            if not uploaded_filename.lower().endswith(('.xlsx', '.xls')):
                excel_name = f"{uploaded_filename}_outcome.xlsx"
            else:
                excel_name = uploaded_filename.rsplit('.', 1)[0] + "_outcome.xlsx"
        else:
            target_date = audit_data.get("date", datetime.now(timezone.utc).date().isoformat())
            excel_name = f"Stock_Count_{target_date}.xlsx"
        
        # Email content
        target_date = audit_data.get("date", datetime.now(timezone.utc).date().isoformat())
        subject = f"Stock Count Report – {target_date} – {user_name}"
        body = f"Dear Manager,\n\nPlease find the Stock Count report attached.\n\nDate: {target_date}\nTotal Items: {len(data)}\n\nRegards,\n{user_name}\nAudit App"
        
        # Create email
        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = os.getenv("MAIL_USERNAME")
        msg["To"] = to_email
        msg["Cc"] = emp_id
        msg.set_content(body)
        
        # Attach Excel
        msg.add_attachment(
            excel_bytes,
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=excel_name
        )
        
        mail_username = os.getenv("MAIL_USERNAME")
        mail_password = os.getenv("MAIL_PASSWORD")
        
        if not mail_username or not mail_password:
            return JSONResponse({
                "message": "Email credentials not configured. Please contact administrator.",
                "success": False
            }, status_code=500)

        try:
            await asyncio.to_thread(send_smtp_message_sync, msg, mail_username, mail_password)
        except smtplib.SMTPAuthenticationError as auth_err:
            logger.error(f"SMTP Authentication failed: {auth_err}")
            return JSONResponse({
                "message": "Email authentication failed. Please verify email credentials in settings.",
                "success": False
            }, status_code=500)
        
        return JSONResponse({"message": "Email sent successfully", "success": True}, status_code=200)
        
    except Exception as e:
        logger.error(f"Send stock count email error: {e}")
        return JSONResponse({"message": f"Failed to send email: {str(e)}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  ADMIN ENDPOINTS
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/check-admin")
async def check_admin(emp_id: str = Depends(get_current_user)):
    try:
        is_admin = admins_collection.find_one({"email": emp_id}) is not None
        print(is_admin)
        return JSONResponse({"message": "Admin check", "success": True, "data": {"is_admin": is_admin}}, status_code=200)
    except Exception as e:
        logger.error(f"Check admin error: {e}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)

@app.get("/api/admin/employees-stats")
async def get_employees_stats(emp_id: str = Depends(get_current_user)):
    try:
        if not admins_collection.find_one({"email": emp_id}):
            return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)
        
        # ULTRA-FAST: Just get email and name (no complex stats)
        all_users = list(users.find({}, {"_id": 0, "email": 1, "name": 1}).limit(100))
        
        return JSONResponse({
            "message": "Stats fetched", 
            "success": True, 
            "data": {"users": all_users, "total": len(all_users)}
        }, status_code=200)
    except Exception as e:
        logger.error(f"Stats error: {e}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)

@app.get("/api/admin/checklist-data")
async def admin_checklist_data(emp_id: str = Depends(get_current_user)):
    try:
        if not admins_collection.find_one({"email": emp_id}):
            return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)
        data = list(audit_data_collection.find({}, {"_id": 0}))
        for d in data:
            if "submitted_at" in d and isinstance(d["submitted_at"], datetime):
                d["submitted_at"] = d["submitted_at"].isoformat()
        return JSONResponse({"message": "Data fetched", "success": True, "data": {"checklists": data}}, status_code=200)
    except Exception as e:
        logger.error(f"Checklist error: {e}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)

def _serialize_mongo(obj):
    if isinstance(obj, datetime):
        return obj.isoformat()
    if isinstance(obj, dict):
        return {k: _serialize_mongo(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_serialize_mongo(v) for v in obj]
    return obj

@app.get("/api/admin/uploaded-history")
async def uploaded_history(emp_id: str = Depends(get_current_user)):
    try:
        if not admins_collection.find_one({"email": emp_id}):
            return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)
        
        # Fetch all uploads, return total_items field (renamed from total_count for frontend compat)
        history = list(upload_history_collection.find(
            {},
            {"_id": 0, "uploaded_by": 1, "uploaded_at": 1, "total_count": 1, "total_items": 1}
        ).sort("uploaded_at", -1))
        
        # Normalize field name and convert datetimes
        for h in history:
            if "uploaded_at" in h and isinstance(h["uploaded_at"], datetime):
                h["uploaded_at"] = h["uploaded_at"].isoformat()
            # Support both field names; always expose as total_items
            if "total_count" in h and "total_items" not in h:
                h["total_items"] = h.pop("total_count")
            elif "total_items" not in h:
                h["total_items"] = 0
        
        return JSONResponse({"message": "History fetched", "success": True, "data": {"history": history}}, status_code=200)
    except Exception as e:
        logger.error(f"Upload history error: {e}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  AUDIT COMPLETION DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────

CHECKLIST_SECTIONS = [
    "general_report", "stock_reconciliation",
    "observations_on_stacking", "observations_on_warehouse_operations",
    "observations_on_warehouse_record_keeping", "observations_on_wh_infrastructure",
    "observations_on_quality_operation", "checklist_wrt_exchange_circular_mentha_oil",
    "checklist_wrt_exchange_circular_metal", "checklist_wrt_exchange_circular_cotton_bales",
    "signature", "photo"
]

@app.get("/api/admin/audit-dashboard")
async def audit_dashboard(emp_id: str = Depends(get_current_user)):
    try:
        if not admins_collection.find_one({"email": emp_id}):
            return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)

        # Fetch ALL submitted audits with relevant fields
        projection = {
            "_id": 0, "user_id": 1, "date": 1, "submitted_at": 1,
            "completion_status": 1, "warehouse_name": 1,
            "sections.general_report.warehouse_name": 1,
            "stock_count_data": 1
        }
        submitted = list(audit_data_collection.find({}, projection).sort("submitted_at", -1))
        in_progress = list(temp_audit_data_collection.find({}, projection).sort("date", -1))

        # Collect all user emails for batch name lookup
        all_user_emails = set()
        for d in submitted + in_progress:
            if d.get("user_id"):
                all_user_emails.add(d["user_id"])
        user_name_map = get_user_names_cached(all_user_emails)

        total_checklist_sections = len(CHECKLIST_SECTIONS)

        def build_row(d, status_label):
            user_email = d.get("user_id", "")
            comp_status = d.get("completion_status") or {}
            completed = sum(1 for k in CHECKLIST_SECTIONS if comp_status.get(k, False))
            pct = round((completed / total_checklist_sections * 100) if total_checklist_sections > 0 else 0)
            # Resolve warehouse name
            wh_name = (d.get("warehouse_name")
                       or (d.get("sections") or {}).get("general_report", {}).get("warehouse_name")
                       or "—")
            sc_items = len(d.get("stock_count_data") or [])
            submitted_at = d.get("submitted_at", "")
            if isinstance(submitted_at, datetime):
                submitted_at = submitted_at.isoformat()
            sc_status = "Submitted" if (status_label == "Submitted" and sc_items > 0) else \
                        ("In Progress" if sc_items > 0 else "Pending")
            return {
                "user_id": user_email,
                "user_name": user_name_map.get(user_email, user_email),
                "date": d.get("date", ""),
                "warehouse_name": wh_name,
                "checklist_completed": completed,
                "checklist_total": total_checklist_sections,
                "checklist_pct": pct,
                "checklist_status": status_label,
                "stock_count_items": sc_items,
                "stock_count_status": sc_status,
                "status": status_label,
                "submitted_at": submitted_at
            }

        rows = [build_row(d, "Submitted") for d in submitted]
        rows += [build_row(d, "In Progress") for d in in_progress]

        return JSONResponse({"message": "Dashboard data fetched", "success": True,
                             "data": {"rows": rows}}, status_code=200)
    except Exception as e:
        logger.error(f"Audit dashboard error: {e}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  WAREHOUSE MASTER
# ─────────────────────────────────────────────────────────────────────────────

class WarehouseItem(BaseModel):
    warehouse_name: str
    warehouse_address: str

class WarehouseMasterPayload(BaseModel):
    warehouses: List[WarehouseItem]

@app.post("/api/admin/warehouse-master")
async def upload_warehouse_master(payload: WarehouseMasterPayload, emp_id: str = Depends(get_current_user)):
    try:
        if not admins_collection.find_one({"email": emp_id}):
            return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)
        if not payload.warehouses:
            return JSONResponse({"message": "No warehouse data provided", "success": False}, status_code=400)
        docs = [{"warehouse_name": w.warehouse_name.strip(),
                 "warehouse_address": w.warehouse_address.strip(),
                 "uploaded_by": emp_id,
                 "uploaded_at": datetime.now(timezone.utc)} for w in payload.warehouses
                if w.warehouse_name.strip()]
        warehouse_master_collection.delete_many({})
        warehouse_master_collection.insert_many(docs)
        return JSONResponse({"message": f"{len(docs)} warehouses uploaded successfully",
                             "success": True, "data": {"count": len(docs)}}, status_code=200)
    except Exception as e:
        logger.error(f"Warehouse master upload error: {e}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)

@app.get("/api/admin/warehouse-master")
async def get_warehouse_master_admin(emp_id: str = Depends(get_current_user)):
    try:
        if not admins_collection.find_one({"email": emp_id}):
            return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)
        warehouses = list(warehouse_master_collection.find({}, {"_id": 0, "warehouse_name": 1, "warehouse_address": 1}))
        return JSONResponse({"message": "Warehouses fetched", "success": True,
                             "data": {"warehouses": warehouses}}, status_code=200)
    except Exception as e:
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)

@app.get("/api/warehouses")
async def get_warehouses(emp_id: str = Depends(get_current_user)):
    """Public endpoint for users to fetch warehouse list for the dropdown."""
    try:
        warehouses = list(warehouse_master_collection.find({}, {"_id": 0, "warehouse_name": 1, "warehouse_address": 1}))
        return JSONResponse({"message": "Warehouses fetched", "success": True,
                             "data": {"warehouses": warehouses}}, status_code=200)
    except Exception as e:
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)

@app.get("/api/admin/analytics")
async def admin_analytics(start_date: str = None, end_date: str = None, emp_id: str = Depends(get_current_user)):
    """Analytics endpoint for admin dashboard with interactive graphs."""
    try:
        if not admins_collection.find_one({"email": emp_id}):
            return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)
        
        from datetime import datetime, timedelta
        from collections import defaultdict
        
        # Parse date range - default to last 7 days for performance
        if not start_date:
            start_date = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
        if not end_date:
            end_date = datetime.now().strftime("%Y-%m-%d")
        
        # Fetch audits in date range with LIMIT for performance
        query = {"date": {"$gte": start_date, "$lte": end_date}}
        temp_audits = list(temp_audit_data_collection.find(query).limit(100))
        submitted_audits = list(audit_data_collection.find(query).limit(100))
        all_audits = temp_audits + submitted_audits
        
        # Calculate metrics
        total_audits = len(all_audits)
        completed_audits = len([a for a in all_audits if a.get("submitted_at")])
        completion_rate = round((completed_audits / total_audits * 100) if total_audits > 0 else 0, 1)
        
        # Average sections completed
        total_sections = 0
        for audit in all_audits:
            comp_status = audit.get("completion_status", {})
            total_sections += sum(1 for v in comp_status.values() if v)
        avg_sections = total_sections / total_audits if total_audits > 0 else 0
        
        # Total stock items counted
        total_stock_items = sum(len(a.get("stock_count_data", [])) for a in all_audits)
        
        # Audits by date (timeline)
        audits_by_date = defaultdict(int)
        for audit in all_audits:
            audits_by_date[audit.get("date", "Unknown")] += 1
        audits_timeline = [{"date": k, "count": v} for k, v in sorted(audits_by_date.items())]
        
        # Completion by user
        user_stats = defaultdict(lambda: {"total": 0, "completed": 0})
        for audit in all_audits:
            user = audit.get("user_id", "Unknown")
            user_stats[user]["total"] += 1
            if audit.get("submitted_at"):
                user_stats[user]["completed"] += 1
        completion_by_user = [{"user": k, "total": v["total"], "completed": v["completed"]} 
                               for k, v in user_stats.items()]
        
        # Warehouse distribution
        warehouse_dist = defaultdict(int)
        for audit in all_audits:
            wh = audit.get("warehouse_name") or audit.get("general_report", {}).get("warehouse_name", "Unknown")
            warehouse_dist[wh] += 1
        warehouse_distribution = [{"warehouse": k, "count": v} for k, v in warehouse_dist.items()]
        
        # Section breakdown (completed, in progress, pending)
        section_stats = {"completed": 0, "in_progress": 0, "pending": 0}
        for audit in all_audits:
            comp_status = audit.get("completion_status", {})
            completed_count = sum(1 for v in comp_status.values() if v)
            total_count = len(comp_status)
            
            if completed_count == total_count and completed_count > 0:
                section_stats["completed"] += 1
            elif completed_count > 0:
                section_stats["in_progress"] += 1
            else:
                section_stats["pending"] += 1
        
        return JSONResponse({
            "message": "Analytics data fetched",
            "success": True,
            "data": {
                "total_audits": total_audits,
                "completion_rate": completion_rate,
                "avg_sections": avg_sections,
                "total_stock_items": total_stock_items,
                "audits_by_date": audits_timeline,
                "completion_by_user": completion_by_user,
                "warehouse_distribution": warehouse_distribution,
                "section_breakdown": section_stats
            }
        }, status_code=200)
    except Exception as e:
        logger.error(f"Analytics error: {e}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)

@app.get("/api/admin/warehouse-status")
async def admin_warehouse_status(date: str = None, emp_id: str = Depends(get_current_user)):
    """Get audit status for each warehouse - shows which warehouses have completed audits, in progress, or not started."""
    try:
        if not admins_collection.find_one({"email": emp_id}):
            return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)
        
        from datetime import datetime
        
        # Get all warehouses from warehouse master
        all_warehouses = list(warehouse_master_collection.find({}, {"_id": 0, "warehouse_name": 1, "warehouse_address": 1}))
        
        # Get audits for the specified date (both temp and submitted)
        # If no date specified, get RECENT audits only (last 30 days) for performance
        if date:
            temp_audits = list(temp_audit_data_collection.find({"date": date}).limit(100))
            submitted_audits = list(audit_data_collection.find({"date": date}).limit(100))
        else:
            # Get recent audits only - last 30 days
            from datetime import timedelta
            date_limit = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
            temp_audits = list(temp_audit_data_collection.find({"date": {"$gte": date_limit}}).limit(100))
            submitted_audits = list(audit_data_collection.find({"date": {"$gte": date_limit}}).limit(100))
        
        # Build warehouse status map
        warehouse_status_map = {}
        
        # Process submitted audits FIRST (higher priority)
        for audit in submitted_audits:
            sections = audit.get("sections", {})
            wh_name = sections.get("general_report", {}).get("warehouse_name") if sections else None
            if not wh_name:
                wh_name = audit.get("warehouse_name")
            if not wh_name:
                continue
            
            # If warehouse not in map yet, initialize it
            if wh_name not in warehouse_status_map:
                warehouse_status_map[wh_name] = {
                    "warehouse_name": wh_name,
                    "warehouse_address": "",
                    "status": "Completed",  # Submitted audit means completed
                    "assigned_users": [],
                    "progress_percentage": 100,
                    "last_updated": None,
                    "audit_id": None,
                    "has_stock_count": False
                }
            else:
                # If already exists but was "In Progress", upgrade to "Completed"
                warehouse_status_map[wh_name]["status"] = "Completed"
                warehouse_status_map[wh_name]["progress_percentage"] = 100
            
            # Update assigned users
            user_id = audit.get("user_id")
            if user_id and user_id not in warehouse_status_map[wh_name]["assigned_users"]:
                warehouse_status_map[wh_name]["assigned_users"].append(user_id)
            
            # Check for stock count
            if audit.get("stock_count_data") and len(audit.get("stock_count_data", [])) > 0:
                warehouse_status_map[wh_name]["has_stock_count"] = True
            
            warehouse_status_map[wh_name]["audit_id"] = str(audit.get("_id", ""))
            
            # Update last_updated - convert datetime to string
            updated_at = audit.get("submitted_at") or audit.get("date")
            if updated_at:
                # Convert datetime to string if needed
                if isinstance(updated_at, datetime):
                    updated_at = updated_at.strftime("%Y-%m-%d %H:%M:%S")
                warehouse_status_map[wh_name]["last_updated"] = updated_at
        
        # Process temp audits SECOND (lower priority - only if not already completed)
        for audit in temp_audits:
            sections = audit.get("sections", {})
            wh_name = sections.get("general_report", {}).get("warehouse_name") if sections else None
            if not wh_name:
                wh_name = audit.get("warehouse_name")
            if not wh_name:
                continue
            
            # If warehouse already marked as "Completed", skip updating it
            if wh_name in warehouse_status_map and warehouse_status_map[wh_name]["status"] == "Completed":
                # Just add user if not already added
                user_id = audit.get("user_id")
                if user_id and user_id not in warehouse_status_map[wh_name]["assigned_users"]:
                    warehouse_status_map[wh_name]["assigned_users"].append(user_id)
                continue
            
            # If warehouse not in map yet, initialize it
            if wh_name not in warehouse_status_map:
                warehouse_status_map[wh_name] = {
                    "warehouse_name": wh_name,
                    "warehouse_address": "",
                    "status": "Not Started",
                    "assigned_users": [],
                    "progress_percentage": 0,
                    "last_updated": None,
                    "audit_id": None,
                    "has_stock_count": False
                }
            
            # Update assigned users
            user_id = audit.get("user_id")
            if user_id and user_id not in warehouse_status_map[wh_name]["assigned_users"]:
                warehouse_status_map[wh_name]["assigned_users"].append(user_id)
            
            # Calculate progress
            comp_status = audit.get("completion_status", {})
            if comp_status:
                completed = sum(1 for v in comp_status.values() if v)
                total = len(comp_status)
                progress = round((completed / total * 100) if total > 0 else 0)
                
                # Update progress
                if progress > warehouse_status_map[wh_name]["progress_percentage"]:
                    warehouse_status_map[wh_name]["progress_percentage"] = progress
                    warehouse_status_map[wh_name]["audit_id"] = str(audit.get("_id", ""))
                
                # Set status to In Progress only if at least one section is completed
                if completed > 0:
                    warehouse_status_map[wh_name]["status"] = "In Progress"
            
            # Check for stock count
            if audit.get("stock_count_data") and len(audit.get("stock_count_data", [])) > 0:
                warehouse_status_map[wh_name]["has_stock_count"] = True
            
            warehouse_status_map[wh_name]["audit_id"] = str(audit.get("_id", ""))
            
            # Update last_updated - convert datetime to string
            updated_at = audit.get("submitted_at") or audit.get("date")
            if updated_at:
                # Convert datetime to string if needed
                if isinstance(updated_at, datetime):
                    updated_at = updated_at.strftime("%Y-%m-%d %H:%M:%S")
                if not warehouse_status_map[wh_name]["last_updated"] or updated_at > warehouse_status_map[wh_name]["last_updated"]:
                    warehouse_status_map[wh_name]["last_updated"] = updated_at
        
        # Process temp audits SECOND (lower priority - only if not already completed)
        for audit in temp_audits:
            sections = audit.get("sections", {})
            wh_name = sections.get("general_report", {}).get("warehouse_name") if sections else None
            if not wh_name:
                wh_name = audit.get("warehouse_name")
            if not wh_name:
                continue
            
            # If warehouse already marked as "Completed", skip updating it
            if wh_name in warehouse_status_map and warehouse_status_map[wh_name]["status"] == "Completed":
                # Just add user if not already added
                user_id = audit.get("user_id")
                if user_id and user_id not in warehouse_status_map[wh_name]["assigned_users"]:
                    warehouse_status_map[wh_name]["assigned_users"].append(user_id)
                continue
            
            # If warehouse not in map yet, initialize it
            if wh_name not in warehouse_status_map:
                warehouse_status_map[wh_name] = {
                    "warehouse_name": wh_name,
                    "warehouse_address": "",
                    "status": "Not Started",
                    "assigned_users": [],
                    "progress_percentage": 0,
                    "last_updated": None,
                    "audit_id": None,
                    "has_stock_count": False
                }
            
            # Update assigned users
            user_id = audit.get("user_id")
            if user_id and user_id not in warehouse_status_map[wh_name]["assigned_users"]:
                warehouse_status_map[wh_name]["assigned_users"].append(user_id)
            
            # Calculate progress
            comp_status = audit.get("completion_status", {})
            if comp_status:
                completed = sum(1 for v in comp_status.values() if v)
                total = len(comp_status)
                progress = round((completed / total * 100) if total > 0 else 0)
                
                # Update progress
                if progress > warehouse_status_map[wh_name]["progress_percentage"]:
                    warehouse_status_map[wh_name]["progress_percentage"] = progress
                    warehouse_status_map[wh_name]["audit_id"] = str(audit.get("_id", ""))
                
                # Set status to In Progress only if at least one section is completed
                if completed > 0:
                    warehouse_status_map[wh_name]["status"] = "In Progress"
            
            # Check for stock count
            if audit.get("stock_count_data") and len(audit.get("stock_count_data", [])) > 0:
                warehouse_status_map[wh_name]["has_stock_count"] = True
            
            # Update last_updated - convert datetime to string
            updated_at = audit.get("date")
            if updated_at:
                # Convert datetime to string if needed
                if isinstance(updated_at, datetime):
                    updated_at = updated_at.strftime("%Y-%m-%d %H:%M:%S")
                if not warehouse_status_map[wh_name]["last_updated"] or updated_at > warehouse_status_map[wh_name]["last_updated"]:
                    warehouse_status_map[wh_name]["last_updated"] = updated_at
        
        # Add warehouses from master that don't have audits
        for wh in all_warehouses:
            wh_name = wh.get("warehouse_name")
            if wh_name not in warehouse_status_map:
                warehouse_status_map[wh_name] = {
                    "warehouse_name": wh_name,
                    "warehouse_address": wh.get("warehouse_address", ""),
                    "status": "Not Started",
                    "assigned_users": [],
                    "progress_percentage": 0,
                    "last_updated": None,
                    "audit_id": None,
                    "has_stock_count": False
                }
            else:
                # Update address from master
                warehouse_status_map[wh_name]["warehouse_address"] = wh.get("warehouse_address", "")
        
        # Convert to list and sort by status priority (Completed, In Progress, Not Started)
        status_priority = {"Completed": 1, "In Progress": 2, "Not Started": 3}
        warehouse_list = sorted(
            warehouse_status_map.values(),
            key=lambda x: (status_priority.get(x["status"], 4), x["warehouse_name"])
        )
        
        return JSONResponse({
            "message": "Warehouse status fetched",
            "success": True,
            "data": {
                "warehouses": warehouse_list,
                "date": date
            }
        }, status_code=200)
    except Exception as e:
        logger.error(f"Warehouse status error: {e}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  STOCK COUNT RECONCILIATION
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/admin/audit-detail/{user_id}/{date}")
async def admin_audit_detail(user_id: str, date: str, emp_id: str = Depends(get_current_user)):
    """Get full audit detail for dashboard view button."""
    try:
        if not admins_collection.find_one({"email": emp_id}):
            return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)
        # Search submitted first, then in-progress
        audit = audit_data_collection.find_one({"user_id": user_id, "date": date})
        if not audit:
            audit = temp_audit_data_collection.find_one({"user_id": user_id, "date": date})
        if not audit:
            return JSONResponse({"message": "Audit not found", "success": False}, status_code=404)
        audit["_id"] = str(audit["_id"])
        if "submitted_at" in audit and isinstance(audit["submitted_at"], datetime):
            audit["submitted_at"] = audit["submitted_at"].isoformat()
        # Resolve warehouse name
        wh_name = (audit.get("warehouse_name")
                   or (audit.get("sections") or {}).get("general_report", {}).get("warehouse_name")
                   or "—")
        audit["warehouse_name"] = wh_name
        # Add user display name
        user_rec = users.find_one({"email": user_id}, {"_id": 0, "name": 1})
        audit["user_name"] = user_rec.get("name", user_id) if user_rec else user_id
        return JSONResponse({"message": "Audit detail fetched", "success": True, "data": audit}, status_code=200)
    except Exception as e:
        logger.error(f"Audit detail error: {e}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)


@app.get("/api/admin/export-audit/{user_id}/{date}")
async def admin_export_audit(user_id: str, date: str, type: str = "checklist", emp_id: str = Depends(get_current_user)):
    """Export audit Excel from admin panel."""
    try:
        if not admins_collection.find_one({"email": emp_id}):
            return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)
        # Find audit in submitted, then temp
        audit_data = audit_data_collection.find_one({"user_id": user_id, "date": date})
        if not audit_data:
            audit_data = temp_audit_data_collection.find_one({"user_id": user_id, "date": date})
        if not audit_data:
            return JSONResponse({"message": "Audit not found", "success": False}, status_code=404)
        if type == "stockcount":
            excel_bytes = generate_stock_count_excel_bytes(audit_data)
            filename = f"StockCount_{user_id}_{date}.xlsx"
        else:
            excel_bytes = await generate_checklist_excel_bytes(user_id, audit_data)
            filename = f"Checklist_{user_id}_{date}.xlsx"
        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
        )
    except Exception as e:
        logger.error(f"Admin export audit error: {e}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)


@app.get("/api/admin/stock-reconciliation")
async def admin_stock_reconciliation(
    from_date: str = None,
    to_date: str = None,
    date: str = None,
    warehouse: str = None,
    emp_id: str = Depends(get_current_user)
):
    """Get reconciliation report showing stock count data filled by users."""
    try:
        if not admins_collection.find_one({"email": emp_id}):
            return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)

        # Support both single date (legacy) and date range
        effective_from = from_date or date
        effective_to = to_date or date

        if not effective_from or not effective_to:
            return JSONResponse({
                "message": "Date range is required",
                "success": True,
                "data": {
                    "reconciliation": [],
                    "pagination": {"current_page": 1, "total_pages": 1, "total_items": 0, "items_per_page": 10},
                    "summary": {"total_items": 0, "matched": 0, "excess": 0, "shortage": 0, "match_rate": 0}
                }
            }, status_code=200)

        # Build date range query
        query = {"date": {"$gte": effective_from, "$lte": effective_to}}

        # Get audits in date range
        temp_audits = list(temp_audit_data_collection.find(query))
        submitted_audits = list(audit_data_collection.find(query))
        all_audits = submitted_audits + temp_audits
        
        reconciliation_data = []
        user_ids = set()
        
        for audit in all_audits:
            stock_count_data = audit.get("stock_count_data", [])
            if not stock_count_data:
                continue
            
            user_id = audit.get("user_id", "Unknown")
            user_ids.add(user_id)
            audit_status = "Submitted" if audit.get("submitted_at") else "In Progress"
            
            # Process each stock count item
            for item in stock_count_data:
                item_code = item.get("item_code", "")
                item_name = item.get("item_name", "")
                sheet_name = item.get("sheet_name", "")
                remarks = item.get("remarks", "")
                
                # Get physical quantity from what user filled
                physical_qty_raw = item.get("physical_amount", item.get("quantity", item.get("qty", 0)))
                try:
                    physical_qty = float(physical_qty_raw) if physical_qty_raw != "" else 0
                except Exception:
                    physical_qty = 0
                
                # Get system quantity from item master (if exists)
                system_qty = 0
                master_item = item_master_collection.find_one({"item_code": item_code, "sheet_name": sheet_name})
                if master_item:
                    system_qty_raw = master_item.get("qty", master_item.get("quantity", 0))
                    try:
                        system_qty = float(system_qty_raw) if system_qty_raw != "" else 0
                    except Exception:
                        system_qty = 0
                
                # Calculate variance
                variance = physical_qty - system_qty
                variance_pct = round((variance / system_qty * 100) if system_qty > 0 else 0, 2)
                
                # Categorize variance
                if variance == 0:
                    variance_status = "Match"
                elif variance > 0:
                    variance_status = "Excess"
                else:
                    variance_status = "Shortage"
                
                reconciliation_data.append({
                    "user_id": user_id,
                    "item_code": item_code,
                    "item_name": item_name,
                    "sheet_name": sheet_name,
                    "remarks": remarks,
                    "system_quantity": system_qty,
                    "physical_quantity": physical_qty,
                    "variance": variance,
                    "variance_percentage": variance_pct,
                    "variance_status": variance_status,
                    "audit_status": audit_status,
                    "date": date
                })
        
        # Use cached user name lookup for better performance
        user_name_map = get_user_names_cached(user_ids)
        
        # Add user names to records
        for record in reconciliation_data:
            record["auditor_name"] = user_name_map.get(record["user_id"], record["user_id"])
        
        # Apply warehouse filter if provided
        if warehouse:
            reconciliation_data = [
                r for r in reconciliation_data
                if warehouse.lower() in r.get("sheet_name", "").lower()
                or warehouse.lower() in r.get("remarks", "").lower()
            ]

        # Summary statistics (all data, before pagination)
        total_items = len(reconciliation_data)
        matched = sum(1 for r in reconciliation_data if r["variance_status"] == "Match")
        excess = sum(1 for r in reconciliation_data if r["variance_status"] == "Excess")
        shortage = sum(1 for r in reconciliation_data if r["variance_status"] == "Shortage")

        return JSONResponse({
            "message": "Reconciliation data fetched",
            "success": True,
            "data": {
                "reconciliation": reconciliation_data,
                "summary": {
                    "total_items": total_items,
                    "matched": matched,
                    "excess": excess,
                    "shortage": shortage,
                    "match_rate": round((matched / total_items * 100) if total_items > 0 else 0, 2)
                },
                "from_date": effective_from,
                "to_date": effective_to,
                "warehouse_filter": warehouse or "All"
            }
        }, status_code=200)
    except Exception as e:
        logger.error(f"Stock reconciliation error: {e}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  TASK ASSIGNMENT
# ─────────────────────────────────────────────────────────────────────────────

class TaskAssignment(BaseModel):
    warehouse_name: str
    assigned_to: List[str]  # List of user email IDs
    task_type: str  # "checklist" or "stock_count"
    due_date: str  # YYYY-MM-DD format
    notes: Optional[str] = ""

@app.post("/api/admin/assign-task")
async def assign_task(task: TaskAssignment, emp_id: str = Depends(get_current_user)):
    """Assign audit tasks to team members for specific warehouses."""
    try:
        if not admins_collection.find_one({"email": emp_id}):
            return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)
        
        # Create task assignment document
        task_doc = {
            "warehouse_name": task.warehouse_name,
            "assigned_to": task.assigned_to,
            "assigned_by": emp_id,
            "task_type": task.task_type,
            "due_date": task.due_date,
            "notes": task.notes,
            "status": "Assigned",
            "created_at": datetime.now(timezone.utc),
            "completed_at": None
        }
        
        # Check if task already exists for this warehouse and date
        existing = task_assignments_collection.find_one({
            "warehouse_name": task.warehouse_name,
            "due_date": task.due_date,
            "task_type": task.task_type
        })
        
        if existing:
            # Update existing task
            task_assignments_collection.update_one(
                {"_id": existing["_id"]},
                {"$set": {
                    "assigned_to": task.assigned_to,
                    "assigned_by": emp_id,
                    "notes": task.notes,
                    "updated_at": datetime.now(timezone.utc)
                }}
            )
            message = "Task assignment updated successfully"
        else:
            # Insert new task
            task_assignments_collection.insert_one(task_doc)
            message = "Task assigned successfully"
        
        # Send email notification to assigned users
        try:
            task_type_label = "Checklist Audit" if task.task_type == "checklist" else "Stock Count"
            email_subject = f"New Task Assignment: {task.warehouse_name}"
            email_body = f"""
            <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <h2 style="color: #4338ca;">New Task Assigned to You</h2>
                <p>You have been assigned a new audit task by {emp_id}.</p>
                
                <div style="background: #f3f4f6; padding: 15px; border-radius: 8px; margin: 20px 0;">
                    <h3 style="margin-top: 0;">Task Details:</h3>
                    <p><strong>Warehouse:</strong> {task.warehouse_name}</p>
                    <p><strong>Task Type:</strong> {task_type_label}</p>
                    <p><strong>Due Date:</strong> {task.due_date}</p>
                    {f'<p><strong>Notes:</strong> {task.notes}</p>' if task.notes else ''}
                </div>
                
                <p>Please complete this task by the due date.</p>
                <p style="color: #666; font-size: 12px; margin-top: 30px;">This is an automated message from the Audit Application.</p>
            </body>
            </html>
            """
            
            # Use helper function to send email to all assigned users
            success, message = send_email_notification(
                to_emails=task.assigned_to,
                subject=email_subject,
                body=email_body,
                attachments=None
            )
            
            if success:
                logger.info(f"Task assignment emails sent to {len(task.assigned_to)} user(s)")
            else:
                logger.error(f"Failed to send task assignment emails: {message}")
                
        except Exception as email_error:
            logger.error(f"Failed to send task assignment emails: {email_error}")
            # Don't fail the entire request if email fails
        
        # Convert datetime to string for JSON response
        task_response = {
            "warehouse_name": task.warehouse_name,
            "assigned_to": task.assigned_to,
            "assigned_by": emp_id,
            "task_type": task.task_type,
            "due_date": task.due_date,
            "notes": task.notes,
            "status": "Assigned"
        }
        
        return JSONResponse({
            "message": message,
            "success": True,
            "data": {"task": task_response}
        }, status_code=200)
    except Exception as e:
        logger.error(f"Task assignment error: {e}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)

@app.get("/api/admin/task-assignments")
async def get_task_assignments(date: str = None, warehouse: str = None, emp_id: str = Depends(get_current_user)):
    """Get all task assignments with optional filters."""
    try:
        if not admins_collection.find_one({"email": emp_id}):
            return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)
        
        # Build query
        query = {}
        if date:
            query["due_date"] = date
        if warehouse:
            query["warehouse_name"] = warehouse
        
        # Get all task assignments
        tasks = list(task_assignments_collection.find(query).sort("due_date", -1))
        
        # Convert ObjectId to string
        for task in tasks:
            task["_id"] = str(task["_id"])
            if "created_at" in task and isinstance(task["created_at"], datetime):
                task["created_at"] = task["created_at"].isoformat()
            if "completed_at" in task and isinstance(task["completed_at"], datetime):
                task["completed_at"] = task["completed_at"].isoformat()
        
        return JSONResponse({
            "message": "Task assignments fetched",
            "success": True,
            "data": {"tasks": tasks}
        }, status_code=200)
    except Exception as e:
        logger.error(f"Get task assignments error: {e}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)

@app.get("/api/user/my-tasks")
async def get_my_tasks(emp_id: str = Depends(get_current_user)):
    """Get tasks assigned to the current user."""
    try:
        # Get all tasks assigned to this user
        tasks = list(task_assignments_collection.find({
            "assigned_to": emp_id,
            "status": {"$ne": "Completed"}
        }).sort("due_date", 1))
        
        # Convert ObjectId to string
        for task in tasks:
            task["_id"] = str(task["_id"])
            if "created_at" in task and isinstance(task["created_at"], datetime):
                task["created_at"] = task["created_at"].isoformat()
        
        return JSONResponse({
            "message": "Your tasks fetched",
            "success": True,
            "data": {"tasks": tasks}
        }, status_code=200)
    except Exception as e:
        logger.error(f"Get my tasks error: {e}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)

@app.delete("/api/admin/task-assignments/{task_id}")
async def delete_task_assignment(task_id: str, emp_id: str = Depends(get_current_user)):
    """Delete a task assignment."""
    try:
        if not admins_collection.find_one({"email": emp_id}):
            return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)
        
        from bson import ObjectId
        
        # Delete the task
        result = task_assignments_collection.delete_one({"_id": ObjectId(task_id)})
        
        if result.deleted_count == 0:
            return JSONResponse({"message": "Task not found", "success": False}, status_code=404)
        
        return JSONResponse({
            "message": "Task deleted successfully",
            "success": True
        }, status_code=200)
    except Exception as e:
        logger.error(f"Delete task error: {e}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  USER DASHBOARD HISTORY
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/user/checklist-history")
async def user_checklist_history(emp_id: str = Depends(get_current_user)):
    """Get all submitted checklists for the current user."""
    try:
        history = list(audit_data_collection.find(
            {"user_id": emp_id},
            {"_id": 1, "date": 1, "submitted_at": 1, "sections": 1, "completion_status": 1}
        ).sort("date", -1))

        for h in history:
            h["audit_id"] = str(h.pop("_id"))
            if "submitted_at" in h and isinstance(h["submitted_at"], datetime):
                h["submitted_at"] = h["submitted_at"].isoformat()
            wh_name = (h.get("sections") or {}).get("general_report", {}).get("warehouse_name", "—")
            h["warehouse_name"] = wh_name
            cs = h.get("completion_status", {})
            h["sections_completed"] = sum(1 for s in CHECKLIST_SECTIONS if cs.get(s, False))
            h["sections_total"] = len(CHECKLIST_SECTIONS)
            # Keep completion_status in response so frontend can show per-section badges
            h["completion_status"] = cs
            # Strip heavy section data from list response (photos/signatures are large)
            h.pop("sections", None)

        return JSONResponse({"message": "History fetched", "success": True,
                             "data": {"history": history}}, status_code=200)
    except Exception as e:
        logger.error(f"Checklist history error: {e}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)


@app.get("/api/export-excel-by-date")
async def export_excel_by_date(date: str = Query(...), emp_id: str = Depends(get_current_user)):
    """Download the Excel for a historical audit by its date (from the History tab)."""
    try:
        audit_data = audit_data_collection.find_one({"user_id": emp_id, "date": date})
        if not audit_data:
            return JSONResponse({"message": f"No submitted audit found for {date}", "success": False}, status_code=404)
        excel_bytes = await generate_excel_bytes(emp_id, audit_data)
        filename = f"Audit_{date}.xlsx"
        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
        )
    except Exception as e:
        logger.error(f"export-excel-by-date error: {e}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)


@app.get("/api/export-excel-by-id")
async def export_excel_by_id(audit_id: str = Query(...), emp_id: str = Depends(get_current_user)):
    """Download the Excel for a historical checklist audit by its specific MongoDB ObjectId."""
    try:
        from bson import ObjectId
        audit_data = audit_data_collection.find_one({"user_id": emp_id, "_id": ObjectId(audit_id)})
        if not audit_data:
            return JSONResponse({"message": "No submitted audit found", "success": False}, status_code=404)
        excel_bytes = await generate_checklist_excel_bytes(emp_id, audit_data)
        date = audit_data.get("date", "Report")
        wh_name = (audit_data.get("sections") or {}).get("general_report", {}).get("warehouse_name", "Audit")
        safe_wh_name = re.sub(r'[\s/\\?*\[\]:]+', '_', wh_name)
        filename = f"Audit_{safe_wh_name}_{date}.xlsx"
        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
        )
    except Exception as e:
        logger.error(f"export-excel-by-id error: {e}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)


@app.get("/api/export-stock-count-excel-by-id")
async def export_stock_count_excel_by_id(audit_id: str = Query(...), emp_id: str = Depends(get_current_user)):
    """Download the Excel for a historical stock count by its specific MongoDB ObjectId."""
    try:
        from bson import ObjectId
        audit = audit_data_collection.find_one({"user_id": emp_id, "_id": ObjectId(audit_id)})
        if not audit or not audit.get("stock_count_data"):
            return JSONResponse({"message": "No stock count data found", "success": False}, status_code=404)
        
        data = audit["stock_count_data"]
        df = pd.DataFrame(data)
        columns = ["sheet_name", "item_name", "item_code", "qty", "physical_amount", "remarks"]
        available_cols = [c for c in columns if c in df.columns]
        df = df[available_cols]
        df.rename(columns={
            "sheet_name": "Sheet Name", 
            "item_name": "Item Name", 
            "item_code": "Item Code", 
            "qty": "Expected Qty", 
            "physical_amount": "Physical Count", 
            "remarks": "Remarks"
        }, inplace=True)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Stock Count')
        excel_bytes = output.getvalue()

        date = audit.get("date", "Report")
        wh_name = (audit.get("sections") or {}).get("general_report", {}).get("warehouse_name", "Audit")
        safe_wh_name = re.sub(r'[\s/\\?*\[\]:]+', '_', wh_name)
        filename = f"Stock_Count_{safe_wh_name}_{date}.xlsx"
        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
        )
    except Exception as e:
        logger.error(f"export-stock-count-excel-by-id error: {e}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)


@app.get("/api/get-historical-section/{audit_id}/{section_name}")
async def get_historical_section(audit_id: str, section_name: str, emp_id: str = Depends(get_current_user)):
    """Retrieve historical section data for a submitted audit."""
    try:
        from bson import ObjectId
        audit = audit_data_collection.find_one({"user_id": emp_id, "_id": ObjectId(audit_id)})
        if not audit:
            return JSONResponse({"message": "Historical audit not found", "success": False}, status_code=404)
        section_data = audit.get("sections", {}).get(section_name, {})
        response = base_response.copy()
        response.update({
            "message": f"Historical section {section_name} retrieved successfully",
            "success": True,
            "data": {"section_data": section_data},
            "status_code": status.HTTP_200_OK
        })
        return JSONResponse(content=response, status_code=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f"Error in get_historical_section: {str(e)}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)


@app.get("/api/get-historical-stock-count/{audit_id}")
async def get_historical_stock_count(audit_id: str, emp_id: str = Depends(get_current_user)):
    """Retrieve historical stock count data for a submitted audit."""
    try:
        from bson import ObjectId
        audit = audit_data_collection.find_one({"user_id": emp_id, "_id": ObjectId(audit_id)})
        if not audit:
            return JSONResponse({"message": "Historical audit not found", "success": False}, status_code=404)
        sc_data = audit.get("stock_count_data", [])
        return JSONResponse({
            "message": "Historical stock count retrieved successfully",
            "success": True,
            "data": {"stock_count_data": sc_data}
        }, status_code=200)
    except Exception as e:
        logger.error(f"Error in get_historical_stock_count: {str(e)}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)


@app.get("/api/user/stock-count-history")
async def user_stock_count_history(emp_id: str = Depends(get_current_user)):
    """Get all stock count records (pending, completed, history) for the current user."""
    try:
        today = datetime.now(timezone.utc).date().isoformat()
        
        # Current in-progress (temp collection) - only show if stock count not submitted
        current_temp = temp_audit_data_collection.find_one({
            "user_id": emp_id,
            "date": today,
            "completion_status.stock_count": {"$ne": True}
        })
        pending_items = []
        if current_temp:
            wh_name = (current_temp.get("sections") or {}).get("general_report", {}).get("warehouse_name", "—")
            sc_data = current_temp.get("stock_count_data", [])
            pending_items = [{
                "audit_id": str(current_temp["_id"]),
                "date": current_temp.get("date"),
                "warehouse_name": wh_name,
                "items_count": len(sc_data),
                "status": "Pending",
                "stock_count_data": sc_data
            }]
        
        # Completed (submitted today - in audit_data_collection)
        completed_today = audit_data_collection.find_one({
            "user_id": emp_id,
            "date": today,
            "completion_status.stock_count": True
        })
        completed_items = []
        if completed_today:
            wh_name = (completed_today.get("sections") or {}).get("general_report", {}).get("warehouse_name", "—")
            sc_data = completed_today.get("stock_count_data", [])
            submitted_at = completed_today.get("submitted_at", "")
            if isinstance(submitted_at, datetime):
                submitted_at = submitted_at.isoformat()
            completed_items = [{
                "audit_id": str(completed_today["_id"]),
                "date": completed_today.get("date"),
                "warehouse_name": wh_name,
                "items_count": len(sc_data),
                "status": "Completed",
                "submitted_at": submitted_at,
                "stock_count_data": sc_data
            }]
        
        # History (all past submitted audits with stock count filled)
        history = list(audit_data_collection.find(
            {"user_id": emp_id, "completion_status.stock_count": True},
            {"_id": 1, "date": 1, "submitted_at": 1, "sections": 1, "stock_count_data": 1}
        ).sort("date", -1))
        
        history_items = []
        for h in history:
            wh_name = (h.get("sections") or {}).get("general_report", {}).get("warehouse_name", "—")
            sc_data = h.get("stock_count_data", [])
            submitted_at = h.get("submitted_at", "")
            if isinstance(submitted_at, datetime):
                submitted_at = submitted_at.isoformat()
            history_items.append({
                "audit_id": str(h["_id"]),
                "date": h.get("date"),
                "warehouse_name": wh_name,
                "items_count": len(sc_data),
                "status": "Submitted",
                "submitted_at": submitted_at,
                "stock_count_data": sc_data
            })
        
        return JSONResponse({"message": "Stock count data fetched", "success": True,
                             "data": {
                                 "pending": pending_items,
                                 "completed": completed_items,
                                 "history": history_items
                             }}, status_code=200)
    except Exception as e:
        logger.error(f"Stock count history error: {e}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)

# ─────────────────────────────────────────────────────────────────────────────
#  STATIC HTML ROUTES
# ─────────────────────────────────────────────────────────────────────────────

def get_token(request: Request) -> str:
    token = request.cookies.get("access_token") 
    if token:
        return token
    auth = request.headers.get("Authorization")
    if auth and auth.startswith("Bearer "):
        return auth.split(" ")[1]
    raise HTTPException(status_code=401, detail="Could not validate credentials")


@app.get("/", response_class=FileResponse)
async def root(request: Request):
    try:
        token = get_token(request)
        get_current_user(token)
        return FileResponse(os.path.join(STATIC_DIR, "index.html"))
    except HTTPException:
        return RedirectResponse(url="/login")

@app.get("/login", response_class=FileResponse)
async def serve_login():
    return FileResponse(os.path.join(STATIC_DIR, "login.html"))

@app.get("/register", response_class=FileResponse)
async def serve_register():
    return FileResponse(os.path.join(STATIC_DIR, "register.html"))
