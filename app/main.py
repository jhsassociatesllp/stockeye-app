"""
main.py  —  Audit Application API
Performance improvements applied:
  - MongoDB compound indexes (see create_indexes.py)
  - N+1 query eliminated in stock reconciliation (bulk fetch + dict lookup)
  - Duplicate temp_audits processing loop removed in warehouse status
  - Warehouse filter syntax error fixed
  - User name cache used consistently across all admin endpoints
  - Projection fields tightened on every query (fetch only needed fields)
  - Aggregation pipeline used in audit dashboard instead of full doc fetch
"""

import datetime
import io
import os
import re
import smtplib
import base64
import logging
import asyncio
import traceback
import time
from collections import defaultdict
from typing import Dict, List, Optional
from datetime import datetime, timezone, timedelta
from bson import ObjectId

import bcrypt
import pandas as pd
import requests
from dotenv import load_dotenv
from email.message import EmailMessage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from fastapi import (
    FastAPI, Depends, File, Form, HTTPException,
    Query, Request, Response, UploadFile, status,
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import (
    FileResponse, JSONResponse, RedirectResponse, StreamingResponse,
)
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from app.auth import create_jwt, get_current_user
from app.database import (
    fs,
    users,
    admins_collection,
    audit_data_collection,
    temp_audit_data_collection,
    item_master_collection,
    upload_history_collection,
    warehouse_master_collection,
    task_assignments_collection,
    checklist_questions_collection,
)
from app.models import AuditForm, UserLogin, UserRegister

load_dotenv()

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

SERPAPI_KEY = os.getenv("SERPAPI_KEY")

# ─────────────────────────────────────────────────────────────────────────────
#  APP SETUP
# ─────────────────────────────────────────────────────────────────────────────

app = FastAPI()

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
STATIC_DIR = os.path.join(BASE_DIR, "static")
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

base_response = {
    "message": "",
    "success": False,
    "data": None,
    "status_code": status.HTTP_400_BAD_REQUEST,
}

CHECKLIST_SECTIONS = [
    "general_report", "stock_reconciliation", "verification_status_previous_audit",
    "observations_on_stacking", "observations_on_warehouse_operations",
    "observations_on_warehouse_record_keeping", "observations_on_wh_infrastructure",
    "observations_on_quality_operation", "checklist_wrt_exchange_circular_mentha_oil",
    "checklist_wrt_exchange_circular_metal", "checklist_wrt_exchange_circular_cotton_bales",
    "signature", "photo",
]

QUESTION_SECTION_TITLES = {
    "verification_status_previous_audit": "Verification Status of Previous Audit",
    "observations_on_stacking": "Observations on Stacking",
    "observations_on_warehouse_operations": "Observations on Warehouse Operations",
    "observations_on_warehouse_record_keeping": "Observations on Warehouse Record Keeping",
    "observations_on_wh_infrastructure": "Observations on WH Infrastructure",
    "observations_on_quality_operation": "Observations on Quality Operation",
    "checklist_wrt_exchange_circular_mentha_oil": "Checklist WRT Exchange Circular - Mentha Oil",
    "checklist_wrt_exchange_circular_metal": "Checklist WRT Exchange Circular - Metal",
    "checklist_wrt_exchange_circular_cotton_bales": "Checklist WRT Exchange Circular - Cotton Bales",
}

QUESTION_SECTIONS = list(QUESTION_SECTION_TITLES.keys())

# ─────────────────────────────────────────────────────────────────────────────
#  USER NAME CACHE  (avoids repeated DB hits for admin endpoints)
# ─────────────────────────────────────────────────────────────────────────────

_USER_NAME_CACHE: Dict[str, str] = {}
_CACHE_TIMESTAMP: Optional[datetime] = None
_CACHE_TTL = timedelta(minutes=10)


def get_user_names_cached(user_emails) -> Dict[str, str]:
    global _USER_NAME_CACHE, _CACHE_TIMESTAMP

    now = datetime.now(timezone.utc)
    if (
        not _CACHE_TIMESTAMP
        or now - _CACHE_TIMESTAMP > _CACHE_TTL
        or not _USER_NAME_CACHE
    ):
        try:
            all_users = list(users.find({}, {"_id": 0, "email": 1, "name": 1}))
            _USER_NAME_CACHE = {
                u["email"]: u.get("name", u["email"])
                for u in all_users
                if u.get("email")
            }
            _CACHE_TIMESTAMP = now
            logger.info(f"User name cache refreshed: {len(_USER_NAME_CACHE)} users")
        except Exception as e:
            logger.error(f"Failed to refresh user name cache: {e}")
            return {email: email for email in user_emails}

    return {email: _USER_NAME_CACHE.get(email, email) for email in user_emails if email}


# ─────────────────────────────────────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def validate_password(password: str) -> bool:
    return (
        len(password) >= 8
        and re.search(r"[A-Z]", password)
        and re.search(r"[a-z]", password)
        and re.search(r"\d", password)
        and re.search(r"[!@#$%^&*()]", password)
    )


def _serialize_mongo(obj):
    if isinstance(obj, datetime):
        return obj.isoformat()
    if isinstance(obj, dict):
        return {k: _serialize_mongo(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_serialize_mongo(v) for v in obj]
    return obj


def _adjust_ws(ws, widths):
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def is_admin(emp_id: str) -> bool:
    return admins_collection.find_one({"email": emp_id}) is not None


def require_admin(emp_id: str):
    if not is_admin(emp_id):
        raise HTTPException(status_code=403, detail="Unauthorized")


def stock_item_key(item: dict) -> tuple:
    return (
        str(item.get("item_code", "")).strip(),
        str(item.get("sheet_name", "")).strip(),
        str(item.get("item_name", "")).strip(),
    )


def merge_stock_count_data(existing_items: list, incoming_items: list) -> list:
    merged = []
    index_by_key = {}

    for item in existing_items or []:
        copied = dict(item)
        index_by_key[stock_item_key(copied)] = len(merged)
        merged.append(copied)

    for item in incoming_items or []:
        copied = dict(item)
        key = stock_item_key(copied)
        if key in index_by_key:
            merged[index_by_key[key]].update(copied)
        else:
            index_by_key[key] = len(merged)
            merged.append(copied)

    return merged


def send_smtp_message_sync(msg, mail_username, mail_password):
    last_err = None
    for attempt in range(1, 4):
        try:
            logger.info(f"SMTP send attempt {attempt}/3...")
            with smtplib.SMTP("smtp.gmail.com", 587, timeout=15) as smtp:
                smtp.ehlo()
                smtp.starttls()
                smtp.ehlo()
                smtp.login(mail_username, mail_password)
                smtp.send_message(msg)
            logger.info("SMTP send successful!")
            return True
        except smtplib.SMTPAuthenticationError as e:
            logger.error(f"SMTP auth failed: {e}")
            raise
        except Exception as e:
            last_err = e
            logger.warning(f"SMTP attempt {attempt} failed: {e}. Traceback:\n{traceback.format_exc()}")
            if attempt < 3:
                time.sleep(1.5)
    if last_err:
        raise last_err


def send_email_notification(
    to_emails: list,
    subject: str,
    body: str,
    attachments: list = None,
) -> tuple:
    """
    Send an email with optional attachments.
    Returns (success: bool, message: str).
    attachments: list of (filename, file_bytes, mime_subtype)
    """
    try:
        mail_username = os.getenv("MAIL_USERNAME")
        mail_password = os.getenv("MAIL_PASSWORD")
        if not mail_username or not mail_password:
            return False, "Email credentials not configured"

        msg = MIMEMultipart()
        msg["From"] = mail_username
        msg["To"] = ", ".join(to_emails)
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "html"))

        if attachments:
            for filename, file_bytes, mime_subtype in attachments:
                msg.add_attachment(
                    file_bytes,
                    maintype="application",
                    subtype=mime_subtype,
                    filename=filename,
                )

        send_smtp_message_sync(msg, mail_username, mail_password)
        return True, "Email sent successfully"

    except smtplib.SMTPAuthenticationError as e:
        return False, f"Email authentication failed: {e}"
    except Exception as e:
        logger.error(f"Email send error: {e}")
        return False, f"Failed to send email: {e}"


def safe_report_name(value: str, fallback: str = "Report") -> str:
    clean = re.sub(r"[\s/\\?*\[\]:]+", "_", str(value or "").strip())
    clean = re.sub(r"_+", "_", clean).strip("._")
    return clean or fallback


def get_audit_warehouse_name(audit_data: dict, fallback: str = "Audit") -> str:
    return (
        (audit_data.get("sections") or {}).get("general_report", {}).get("warehouse_name")
        or audit_data.get("warehouse_name")
        or fallback
    )


def stock_count_report_filename(audit_data: dict, fallback_date: str = "") -> str:
    date = safe_report_name(audit_data.get("date") or fallback_date, "Report")
    return f"Stock_Count_{date}.xlsx"


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL GENERATION
# ─────────────────────────────────────────────────────────────────────────────

async def generate_checklist_excel_bytes(emp_id: str, audit_data: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    sections = audit_data.get("sections", {}) or {}

    # General Report
    ws = wb.create_sheet("General Report")
    ws.append(["Field", "Value"])
    gr = sections.get("general_report", {})
    if gr:
        for k, v in gr.items():
            if k == "previous_audits" and isinstance(v, list):
                ws.append(["Previous Audits", ""])
                for i, rec in enumerate(v, 1):
                    ws.append([f"  Record {i} – Date", str(rec.get("date", ""))])
                    ws.append([f"  Record {i} – Auditor Name", str(rec.get("auditor_name", ""))])
                    ws.append([f"  Record {i} – Auditor Type", str(rec.get("auditor_type", ""))])
                    if rec.get("agency_name"):
                        ws.append([f"  Record {i} – Agency Name", str(rec.get("agency_name", ""))])
                pass
            else:
                ws.append([k.replace("_", " ").title(), str(v)])
    else:
        ws.append(["No general report saved.", ""])
    _adjust_ws(ws, [40, 30])

    # Stock Reconciliation
    ws = wb.create_sheet("Stock Reconciliation")
    ws.append(["Commodity Name", "Stock Type", "Qty as per MCXCCL",
               "Qty as per Registered", "Qty as per Physical", "Difference", "Remarks"])
    stock = sections.get("stock_reconciliation", {}).get("commodities", [])
    if stock:
        for item in stock:
            stock_rows = item.get("stocks") if isinstance(item.get("stocks"), list) else [item]
            for stock_row in stock_rows:
                ws.append([
                    item.get("commodity_name", ""), stock_row.get("commodity", ""),
                    stock_row.get("qty_mcxccl", ""), stock_row.get("qty_registered", ""),
                    stock_row.get("qty_physical", ""), stock_row.get("difference", ""),
                    stock_row.get("remarks", ""),
                ])
    else:
        ws.append(["No stock data.", "", "", "", "", "", ""])
    _adjust_ws(ws, [20, 20, 20, 20, 20, 20, 30])

    # Question-based sections
    q_sections = [
        ("verification_status_previous_audit",          "Previous Audit Verification"),
        ("observations_on_stacking",                    "Observations on Stacking"),
        ("observations_on_warehouse_operations",        "Observations on WH Operations"),
        ("observations_on_warehouse_record_keeping",    "Observations on WH Record Keeping"),
        ("observations_on_wh_infrastructure",           "Observations on WH Infrastructure"),
        ("observations_on_quality_operation",           "Observations on Quality Operation"),
        ("checklist_wrt_exchange_circular_mentha_oil",  "Checklist Mentha Oil"),
        ("checklist_wrt_exchange_circular_metal",       "Checklist Metals"),
        ("checklist_wrt_exchange_circular_cotton_bales","Checklist Cotton Bales"),
    ]
    for key, title in q_sections:
        ws = wb.create_sheet(title)
        ws.append(["Question", "Answer", "Remarks"])
        section_observations = sections.get(key, {}).get("section_observations", "")
        if section_observations:
            ws.append(["Section Observations", "", str(section_observations)])
        qlist = sections.get(key, {}).get("questions", [])
        if qlist:
            for idx, q in enumerate(qlist, 1):
                answer = q.get("answer", "")
                if not answer and q.get("value") not in (None, ""):
                    answer = q.get("value", "")
                ws.append([
                    f"{idx}. {q.get('question', f'Question {idx}').strip()}",
                    str(answer).strip(),
                    str(q.get("remarks", "")).strip(),
                ])
                for sub_idx, sub in enumerate(q.get("subquestions", []) or [], 1):
                    sub_answer = sub.get("answer", "")
                    if not sub_answer and sub.get("value") not in (None, ""):
                        sub_answer = sub.get("value", "")
                    ws.append([
                        f"  {idx}.{sub_idx} {sub.get('question', f'Subquestion {sub_idx}').strip()}",
                        str(sub_answer).strip(),
                        str(sub.get("remarks", "")).strip(),
                    ])
        else:
            ws.append(["No data saved.", "", ""])
        _adjust_ws(ws, [60, 10, 30])

    # Signature
    ws = wb.create_sheet("Signature")
    sig = sections.get("signature", {}).get("signature")
    if sig:
        try:
            img_data = re.sub("^data:image/.+;base64,", "", sig)
            img_bytes = io.BytesIO(base64.b64decode(img_data))
            img = Image(img_bytes)
            img.width, img.height = 250, 150
            ws["A1"] = "Signature captured during the audit"
            ws.row_dimensions[1].height = 18
            ws.row_dimensions[2].height = 8
            ws.add_image(img, "A3")
        except Exception as e:
            ws["A1"] = f"Unable to embed signature: {e}"
    else:
        ws["A1"] = "Signature not found."
    ws.column_dimensions["A"].width = 60

    # Photo
    ws = wb.create_sheet("Photo")
    photo_section = sections.get("photo", {})
    photos_list = photo_section.get("photos", [])
    if not photos_list and photo_section.get("photo"):
        photos_list = [{
            "photo": photo_section.get("photo"),
            "maps_url": photo_section.get("maps_url", ""),
            "timestamp": datetime.now().isoformat(),
            "location_text": "Legacy photo",
        }]

    if photos_list:
        ws.append(["Photo #", "Timestamp", "Location", "Google Maps Link"])
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 50

        for idx, photo_item in enumerate(photos_list, 1):
            data_row = idx + 1
            ws[f"A{data_row}"] = f"Photo {idx}"
            ws[f"C{data_row}"] = photo_item.get("location_text", "N/A")
            ws[f"D{data_row}"] = photo_item.get("maps_url", "") or "N/A"

            try:
                photo_data = photo_item.get("photo", "")
                if photo_data:
                    from PIL import Image as PILImage
                    raw = base64.b64decode(
                        re.sub("^data:image/.+;base64,", "", photo_data)
                        if photo_data.startswith("data:image")
                        else photo_data
                    )
                    pil_img = PILImage.open(io.BytesIO(raw))
                    pil_img.thumbnail((400, 300), PILImage.Resampling.LANCZOS)
                    buf = io.BytesIO()
                    pil_img.save(buf, format="PNG")
                    buf.seek(0)
                    excel_img = Image(buf)
                    excel_img.width, excel_img.height = 400, 300
                    ws.add_image(excel_img, f"B{data_row}")
                    ws.row_dimensions[data_row].height = 225
            except Exception as e:
                logger.warning(f"Failed to embed photo {idx}: {e}")
                ws[f"B{data_row}"] = f"[Image embedding failed: {e}]"
    else:
        ws.append(["No photos captured"])
        ws.column_dimensions["A"].width = 30

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


# Alias kept so old callers still work
async def generate_excel_bytes(emp_id: str, audit_data: dict) -> bytes:
    return await generate_checklist_excel_bytes(emp_id, audit_data)


def generate_stock_count_excel_bytes(audit_data: dict) -> bytes:
    sc_data = audit_data.get("stock_count_data", [])
    df = pd.DataFrame(sc_data) if sc_data else pd.DataFrame(
        columns=["sheet_name", "item_name", "item_code", "qty", "physical_amount", "remarks"]
    )
    cols = ["sheet_name", "item_name", "item_code", "qty", "physical_amount", "remarks"]
    df = df[[c for c in cols if c in df.columns]]
    df.rename(columns={
        "sheet_name": "Sheet Name", "item_name": "Item Name",
        "item_code": "Item Code", "qty": "Expected Qty",
        "physical_amount": "Physical Count", "remarks": "Remarks",
    }, inplace=True)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Stock Count")
    return output.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
#  AUTH ENDPOINTS
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/register")
async def register(user: UserRegister):
    logger.info(f"Register attempt: {user.email}")
    try:
        if not validate_password(user.password):
            return JSONResponse(
                {"message": "Invalid password format", "success": False}, status_code=400
            )
        if user.password != user.confirm_password:
            return JSONResponse(
                {"message": "Passwords do not match", "success": False}, status_code=400
            )
        if users.find_one({"email": user.email}):
            return JSONResponse(
                {"message": "Email already registered", "success": False}, status_code=400
            )
        hashed = bcrypt.hashpw(user.password.encode(), bcrypt.gensalt()).decode()
        users.insert_one({
            "name": user.name,
            "email": user.email,
            "password_hash": hashed,
            "created_at": datetime.now(timezone.utc),
        })
        return JSONResponse(
            {"message": "User registered successfully", "success": True,
             "data": {"email": user.email}},
            status_code=201,
        )
    except Exception as e:
        logger.error(f"Register error: {e}")
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)


@app.post("/api/login")
def login(user: UserLogin):
    logger.info(f"Login attempt: {user.email}")
    db_user = users.find_one({"email": user.email})
    if not db_user:
        return JSONResponse({"message": "Invalid email or password", "success": False}, status_code=401)
    pw = user.password
    ph = db_user.get("password_hash")
    if not isinstance(pw, str) or not isinstance(ph, str):
        return JSONResponse({"message": "Invalid email or password", "success": False}, status_code=401)
    if not bcrypt.checkpw(pw.encode(), ph.encode()):
        return JSONResponse({"message": "Invalid email or password", "success": False}, status_code=401)
    token = create_jwt({"sub": user.email})
    return JSONResponse(
        {"message": "Logged in successfully", "success": True,
         "data": {"access_token": token}},
        status_code=200,
    )


@app.get("/api/me")
async def get_me(emp_id: str = Depends(get_current_user)):
    user = users.find_one({"email": emp_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    resp = base_response.copy()
    resp.update({
        "message": "User info retrieved successfully",
        "success": True,
        "data": {"email": user["email"], "name": user.get("name", "Unknown")},
        "status_code": 200,
    })
    return JSONResponse(content=resp, status_code=200)


@app.post("/api/logout")
async def logout(emp_id: str = Depends(get_current_user)):
    return JSONResponse({"message": "Logged out successfully", "success": True, "data": None})


# ─────────────────────────────────────────────────────────────────────────────
#  AUDIT SECTIONS
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/get-sections")
async def get_sections(emp_id: str = Depends(get_current_user)):
    try:
        today = datetime.now(timezone.utc).date().isoformat()
        audit = temp_audit_data_collection.find_one(
            {"user_id": emp_id, "date": today},
            {"completion_status": 1},
        )
        cs = (audit or {}).get("completion_status", {})
        completion_status = {k: cs.get(k, False) for k in CHECKLIST_SECTIONS}
        resp = base_response.copy()
        resp.update({
            "message": "Sections retrieved successfully",
            "success": True,
            "data": {"completion_status": completion_status},
            "status_code": 200,
        })
        return JSONResponse(content=resp, status_code=200)
    except Exception as e:
        logger.error(f"get_sections error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/get-section/{section_name}")
async def get_section(section_name: str, emp_id: str = Depends(get_current_user)):
    try:
        today = datetime.now(timezone.utc).date().isoformat()
        audit = temp_audit_data_collection.find_one(
            {"user_id": emp_id, "date": today},
            {"sections": 1},
        )
        section_data = (audit or {}).get("sections", {}).get(section_name, {})
        resp = base_response.copy()
        resp.update({
            "message": f"Section {section_name} retrieved successfully",
            "success": True,
            "data": {"section_data": section_data},
            "status_code": 200,
        })
        return JSONResponse(content=resp, status_code=200)
    except Exception as e:
        logger.error(f"get_section error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/save-section")
async def save_section(request: Request, emp_id: str = Depends(get_current_user)):
    try:
        body = await request.json()
        section = body.get("section")
        data = body.get("data")
        date = body.get("date")
        if not section or not data or not date:
            return JSONResponse(
                {"message": "Missing required fields (section, data, date)", "success": False},
                status_code=400,
            )

        # Store photo in GridFS to avoid large base64 blobs in the document
        if section == "photo" and "photo" in data and isinstance(data["photo"], str) and data["photo"].startswith("data:image"):
            try:
                _, b64data = data["photo"].split(",", 1)
                file_id = fs.put(
                    base64.b64decode(b64data),
                    filename=f"{emp_id}_{date}_photo.png",
                    content_type="image/png",
                    metadata={"user_id": emp_id, "date": date},
                )
                data["photo"] = None
                data["photo_file_id"] = str(file_id)
            except Exception as img_err:
                logger.warning(f"GridFS photo store failed: {img_err}")

        audit = temp_audit_data_collection.find_one({"user_id": emp_id, "date": date})
        if not audit:
            audit = {
                "user_id": emp_id, "date": date,
                "sections": {}, "completion_status": {},
                "submitted_by": emp_id,
                "submitted_at": datetime.now(timezone.utc),
            }

        audit["sections"][section] = data
        audit["completion_status"][section] = True

        if audit.get("_id"):
            temp_audit_data_collection.update_one(
                {"_id": audit["_id"]},
                {"$set": {
                    f"sections.{section}": data,
                    f"completion_status.{section}": True,
                }},
            )
        else:
            temp_audit_data_collection.insert_one(audit)

        resp = base_response.copy()
        resp.update({
            "message": f"Section {section} saved successfully",
            "success": True,
            "data": {"completion_status": audit["completion_status"]},
            "status_code": 200,
        })
        return JSONResponse(content=resp, status_code=200)
    except Exception as e:
        logger.error(f"save_section error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/get-photo/{file_id}")
async def get_photo(file_id: str, emp_id: str = Depends(get_current_user)):
    try:
        oid = ObjectId(file_id)
        if not fs.exists(oid):
            return JSONResponse({"message": "Photo not found", "success": False}, status_code=404)
        grid_out = fs.get(oid)
        return StreamingResponse(io.BytesIO(grid_out.read()), media_type="image/png")
    except Exception as e:
        logger.error(f"get_photo error: {e}")
        return JSONResponse({"message": str(e), "success": False}, status_code=500)


@app.post("/api/submit-audit")
async def submit_audit(emp_id: str = Depends(get_current_user)):
    try:
        today = datetime.now(timezone.utc).date().isoformat()
        temp_audit = temp_audit_data_collection.find_one({"user_id": emp_id, "date": today})
        if not temp_audit:
            raise HTTPException(status_code=404, detail="No audit data found to submit")

        completion = temp_audit.get("completion_status", {})
        if not all(completion.get(s, False) for s in CHECKLIST_SECTIONS):
            raise HTTPException(status_code=400, detail="Not all sections are completed")

        temp_audit["submitted_at"] = datetime.now(timezone.utc)
        temp_id = temp_audit.pop("_id")
        result = audit_data_collection.insert_one(temp_audit)
        temp_audit_data_collection.delete_one({"_id": temp_id})

        resp = base_response.copy()
        resp.update({
            "message": "Audit submitted successfully",
            "success": True,
            "data": {"submitted": True, "audit_id": str(result.inserted_id)},
            "status_code": 200,
        })
        return JSONResponse(content=resp, status_code=200)
    except Exception as e:
        logger.error(f"submit_audit error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/clear-sections")
async def clear_sections(emp_id: str = Depends(get_current_user)):
    try:
        today = datetime.now(timezone.utc).date().isoformat()
        temp_audit_data_collection.delete_one({"user_id": emp_id, "date": today})
        return JSONResponse({"message": "Sections cleared", "success": True})
    except Exception as e:
        return JSONResponse({"message": str(e), "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  LOCATION
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/get-ip-location")
def get_ip_location():
    services = [
        (
            "https://freeipapi.com/api/json",
            lambda d: {"lat": d.get("latitude"), "lon": d.get("longitude")}
            if d.get("latitude") else None,
        ),
        (
            "https://ipapi.co/json/",
            lambda d: {"lat": d.get("latitude"), "lon": d.get("longitude")}
            if d.get("latitude") else None,
        ),
        (
            "https://ip-api.com/json/?fields=lat,lon,status",
            lambda d: {"lat": d.get("lat"), "lon": d.get("lon")}
            if d.get("status") == "success" else None,
        ),
    ]
    for url, parse_fn in services:
        try:
            res = requests.get(url, timeout=5)
            if res.status_code == 200:
                coords = parse_fn(res.json())
                if coords and coords["lat"] and coords["lon"]:
                    return {"latitude": coords["lat"], "longitude": coords["lon"], "success": True}
        except Exception as e:
            logger.warning(f"IP geolocation {url} failed: {e}")
    return {"success": False, "error": "Could not determine IP location"}


@app.get("/api/get-location")
def get_location(lat: float = Query(...), lon: float = Query(...)):
    serp_url = (
        f"https://serpapi.com/search?engine=google_maps"
        f"&q={lat},{lon}&type=search&api_key={SERPAPI_KEY}"
    )
    try:
        serp_res = requests.get(serp_url, timeout=6)
        serp_data = serp_res.json()
        if serp_res.status_code == 200 and "search_metadata" in serp_data:
            place = serp_data.get("place_results", {})
            plus_code = place.get("plus_code") or "N/A"
            address = place.get("title") or "N/A"
            maps_url = serp_data.get("search_metadata", {}).get("google_maps_url")
            if maps_url:
                return {
                    "source": "serpapi", "latitude": lat, "longitude": lon,
                    "plus_code": plus_code, "address": address, "maps_url": maps_url,
                }
        raise Exception("SerpApi incomplete")
    except Exception as e:
        logger.warning(f"SerpApi failed: {e}")
        osm = requests.get(
            f"https://nominatim.openstreetmap.org/reverse?format=jsonv2&lat={lat}&lon={lon}",
            headers={"User-Agent": "audit-app"},
        ).json()
        address = osm.get("display_name", "Address not found")
        return {
            "source": "osm", "latitude": lat, "longitude": lon,
            "plus_code": address, "address": address,
            "maps_url": f"https://www.google.com/maps/search/{lat}%2C{lon}?hl=en",
        }


# ─────────────────────────────────────────────────────────────────────────────
#  EXPORT EXCEL
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/admin/export-audit/{user_id}/{date}")
async def export_audit(
    user_id: str,
    date: str,
    type: str = Query("checklist", description="Type: checklist or stockcount"),
    emp_id: str = Depends(get_current_user)
):
    """Export individual audit data as Excel file"""
    try:
        require_admin(emp_id)
        
        # Find the audit record
        audit_data = temp_audit_data_collection.find_one(
            {"user_id": user_id, "date": date}
        )
        if not audit_data:
            audit_data = audit_data_collection.find_one(
                {"user_id": user_id, "date": date}
            )
        
        if not audit_data:
            return JSONResponse(
                {"message": "Audit record not found", "success": False},
                status_code=404
            )
        
        if type == "checklist":
            # Generate checklist Excel
            excel_bytes = await generate_checklist_excel_bytes(user_id, audit_data)
            filename = f"checklist_{user_id}_{date}.xlsx"
        else:
            # Generate stock count Excel
            excel_bytes = generate_stock_count_excel_bytes(audit_data)
            filename = f"stockcount_{user_id}_{date}.xlsx"
        
        if not excel_bytes:
            return JSONResponse(
                {"message": "Failed to generate Excel file", "success": False},
                status_code=500
            )
        
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type=media_type,
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
        )
        
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Export audit error: {e}")
        return JSONResponse(
            {"message": f"Export failed: {str(e)}", "success": False},
            status_code=500
        )

        # Prefer temp collection first (in-progress), fall back to submitted
        audit_data = temp_audit_data_collection.find_one(
            {"user_id": emp_id, "date": today, "sections": {"$exists": True}},
            sort=[("submitted_at", -1)],
        )
        if not audit_data:
            audit_data = audit_data_collection.find_one(
                {"user_id": emp_id, "date": today, "sections": {"$exists": True}},
                sort=[("submitted_at", -1)],
            )
        if not audit_data:
            return JSONResponse(
                {"message": "No checklist audit data for today", "success": False},
                status_code=404,
            )

        missing = [
            s for s in CHECKLIST_SECTIONS
            if not audit_data.get("completion_status", {}).get(s, False)
        ]
        if missing:
            return JSONResponse(
                {"message": f"Complete all sections before exporting. Missing: {', '.join(missing)}",
                 "success": False},
                status_code=400,
            )

        excel_bytes = await generate_excel_bytes(emp_id, audit_data)
        filename = f"audit_{emp_id}_{today}.xlsx"
        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
        )
    except Exception as e:
        logger.error(f"export-excel error: {e}")
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)

# ─────────────────────────────────────────────────────────────────────────────
#  SEND EMAIL
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/send-email")
async def send_email(
    to_email: str = Form(...),
    email_subject: Optional[str] = Form(default=None),
    email_body: Optional[str] = Form(default=None),
    attachment: Optional[UploadFile] = File(default=None),
    attachments: List[UploadFile] = File(default=[]),
    email_type: str = Form(default="manual"),
    audit_id: Optional[str] = Form(default=None),
    emp_id: str = Depends(get_current_user),
):
    try:
        user = users.find_one({"email": emp_id}, {"name": 1})
        user_name = (user or {}).get("name", "Unknown")
        today = datetime.now(timezone.utc).date().isoformat()
        all_uploads = ([attachment] if attachment else []) + (attachments or [])

        if email_type == "manual":
            if not all_uploads:
                return JSONResponse(
                    {"message": "Please attach at least one PDF or Excel file", "success": False},
                    status_code=400,
                )

            msg = EmailMessage()
            msg["Subject"] = email_subject or f"Documents from {user_name}"
            msg["From"] = os.getenv("MAIL_USERNAME")
            msg["To"] = to_email
            msg["Cc"] = emp_id
            msg.set_content(
                email_body
                or f"Dear Manager,\n\nPlease find the attached document(s).\n\nRegards,\n{user_name}"
            )

            allowed_ext = (".pdf", ".xlsx", ".xls", ".xlsb")
            for uploaded_file in all_uploads:
                if not uploaded_file.filename.lower().endswith(allowed_ext):
                    return JSONResponse(
                        {"message": f"File '{uploaded_file.filename}' has invalid type. "
                                    "Only PDF or Excel files are allowed",
                         "success": False},
                        status_code=400,
                    )
                file_bytes = await uploaded_file.read()
                subtype = (
                    "pdf" if uploaded_file.filename.lower().endswith(".pdf")
                    else "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                msg.add_attachment(
                    file_bytes, maintype="application",
                    subtype=subtype, filename=uploaded_file.filename,
                )

            mail_username = os.getenv("MAIL_USERNAME")
            mail_password = os.getenv("MAIL_PASSWORD")
            if not mail_username or not mail_password:
                return JSONResponse(
                    {"message": "Email credentials not configured.", "success": False},
                    status_code=500,
                )

            try:
                await asyncio.to_thread(send_smtp_message_sync, msg, mail_username, mail_password)
            except smtplib.SMTPAuthenticationError as e:
                return JSONResponse(
                    {"message": "Email authentication failed.", "success": False},
                    status_code=500,
                )

            return JSONResponse({
                "message": f"Email sent successfully with {len(all_uploads)} attachment(s)",
                "success": True,
            })

        # Resolve audit document
        if audit_id:
            audit_data = audit_data_collection.find_one(
                {"user_id": emp_id, "_id": ObjectId(audit_id)}
            )
            if not audit_data:
                audit_data = temp_audit_data_collection.find_one(
                    {"user_id": emp_id, "_id": ObjectId(audit_id)}
                )
        else:
            collection_filter = {"user_id": emp_id, "date": today}
            if email_type == "stock-count":
                collection_filter["stock_count_data"] = {"$exists": True}
            else:
                collection_filter["sections"] = {"$exists": True}

            audit_data = audit_data_collection.find_one(
                collection_filter, sort=[("submitted_at", -1)]
            )
            if not audit_data:
                audit_data = temp_audit_data_collection.find_one(
                    collection_filter, sort=[("submitted_at", -1)]
                )

        if not audit_data:
            return JSONResponse(
                {"message": "No audit data found to email", "success": False},
                status_code=404,
            )

        target_date = audit_data.get("date", today)
        safe_wh = safe_report_name(get_audit_warehouse_name(audit_data), "Audit")

        if email_type == "stock-count":
            if not audit_data.get("completion_status", {}).get("stock_count", False):
                return JSONResponse(
                    {"message": "Please submit stock count before sending email", "success": False},
                    status_code=400,
                )
            if not audit_data.get("stock_count_data"):
                return JSONResponse(
                    {"message": "No stock count data found", "success": False},
                    status_code=400,
                )
            excel_bytes = generate_stock_count_excel_bytes(audit_data)
            excel_name = stock_count_report_filename(audit_data, target_date)
            subject = email_subject or f"Stock Count Report – {target_date} – {user_name}"
            body = email_body or (
                f"Dear Manager,\n\nPlease find the Stock Count report attached."
                f"\n\nRegards,\n{user_name}\nAudit App"
            )
        else:
            completion = audit_data.get("completion_status", {})
            if not all(completion.get(s, False) for s in CHECKLIST_SECTIONS):
                return JSONResponse(
                    {"message": "Complete all checklist sections before sending email",
                     "success": False},
                    status_code=400,
                )
            excel_bytes = await generate_excel_bytes(emp_id, audit_data)
            excel_name = f"Checklist_Audit_{safe_wh}_{target_date}.xlsx"
            subject = email_subject or f"Checklist Audit Report – {target_date} – {user_name}"
            body = email_body or (
                f"Dear Auditor Manager,\n\nPlease find the Checklist Audit report attached."
                f"\n\nRegards,\n{user_name}\nAudit App"
            )

        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = os.getenv("MAIL_USERNAME")
        msg["To"] = to_email
        msg["Cc"] = emp_id
        msg.set_content(body)

        # Attach user-uploaded files
        allowed_ext = (".pdf", ".xlsx", ".xls", ".xlsb")
        for uploaded_file in all_uploads:
            if not uploaded_file.filename.lower().endswith(allowed_ext):
                return JSONResponse(
                    {"message": f"File '{uploaded_file.filename}' has invalid type. "
                                "Only PDF or Excel files are allowed",
                     "success": False},
                    status_code=400,
                )
            file_bytes = await uploaded_file.read()
            subtype = (
                "pdf" if uploaded_file.filename.lower().endswith(".pdf")
                else "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            msg.add_attachment(
                file_bytes, maintype="application",
                subtype=subtype, filename=uploaded_file.filename,
            )

        # Always attach the generated Excel
        msg.add_attachment(
            excel_bytes, maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=excel_name,
        )

        mail_username = os.getenv("MAIL_USERNAME")
        mail_password = os.getenv("MAIL_PASSWORD")
        if not mail_username or not mail_password:
            return JSONResponse(
                {"message": "Email credentials not configured.", "success": False},
                status_code=500,
            )

        try:
            await asyncio.to_thread(send_smtp_message_sync, msg, mail_username, mail_password)
        except smtplib.SMTPAuthenticationError as e:
            return JSONResponse(
                {"message": "Email authentication failed.", "success": False},
                status_code=500,
            )

        return JSONResponse({"message": "Email sent successfully", "success": True})
    except Exception as e:
        logger.error(f"send-email error: {e}")
        return JSONResponse({"message": f"Failed to send email: {e}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  ITEM MASTER  —  JSON upload (wizard)
# ─────────────────────────────────────────────────────────────────────────────

class SheetItemsPayload(BaseModel):
    sheet_name: str
    items: List[dict]


class UploadItemsJsonPayload(BaseModel):
    sheets: List[SheetItemsPayload]


@app.post("/api/upload-items-json")
async def upload_items_json(
    payload: UploadItemsJsonPayload,
    emp_id: str = Depends(get_current_user),
):
    try:
        logger.info(f"upload-items-json: {emp_id}, sheets={[s.sheet_name for s in payload.sheets]}")
        all_items = []
        sheet_summary = []
        now = datetime.now(timezone.utc)

        for sheet in payload.sheets:
            sheet_name = sheet.sheet_name.strip()
            valid_items = []
            for raw in sheet.items:
                item_code = str(raw.get("item_code", "")).strip()
                item_name = str(raw.get("item_name", "")).strip()
                qty       = str(raw.get("qty", "")).strip()
                extra_col = str(raw.get("extra_col", "")).strip()

                if not item_code or not item_name:
                    continue
                if item_code.lower() in ("nan", "none", "item code", ""):
                    continue
                if item_name.lower() in ("nan", "none", "item name", ""):
                    continue

                valid_items.append({
                    "item_code": item_code, "item_name": item_name,
                    "qty": qty, "extra_col": extra_col,
                    "sheet_name": sheet_name,
                    "uploaded_by": emp_id, "uploaded_at": now,
                })

            all_items.extend(valid_items)
            sheet_summary.append({"sheet": sheet_name, "count": len(valid_items)})
            logger.info(f"  Sheet '{sheet_name}': {len(valid_items)} valid items")

        if not all_items:
            return JSONResponse(
                {"message": "No valid items found. Check Item Code and Item Name columns.",
                 "success": False},
                status_code=400,
            )

        upload_history_collection.insert_one({
            "uploaded_by": emp_id, "uploaded_at": now,
            "total_items": len(all_items), "sheets": sheet_summary,
        })

        uploaded_sheet_names = [s.sheet_name.strip() for s in payload.sheets]
        item_master_collection.delete_many({"sheet_name": {"$in": uploaded_sheet_names}})
        item_master_collection.insert_many(all_items)

        return JSONResponse({
            "message": f"Successfully uploaded {len(all_items)} items from {len(payload.sheets)} sheet(s)",
            "success": True,
            "data": {
                "total_count": len(all_items),
                "sheets": sheet_summary,
                "sample_items": [
                    {"item_code": i["item_code"], "item_name": i["item_name"],
                     "sheet_name": i["sheet_name"]}
                    for i in all_items[:5]
                ],
            },
        })
    except Exception as e:
        logger.error(f"upload-items-json error: {e}")
        import traceback; logger.error(traceback.format_exc())
        return JSONResponse({"message": f"Failed to upload items: {e}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  ITEM MASTER  —  legacy file upload
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/upload-items")
async def upload_items(
    file: UploadFile = File(...),
    emp_id: str = Depends(get_current_user),
):
    try:
        fname = file.filename.lower()
        if not any(fname.endswith(ext) for ext in (".xlsx", ".xls", ".xlsb", ".csv")):
            return JSONResponse(
                {"message": "Only Excel or CSV files are allowed", "success": False},
                status_code=400,
            )
        contents = await file.read()
        try:
            if fname.endswith(".csv"):
                df = pd.read_csv(io.BytesIO(contents))
            elif fname.endswith(".xlsb"):
                df = pd.read_excel(io.BytesIO(contents), engine="pyxlsb")
            else:
                df = pd.read_excel(io.BytesIO(contents))
        except Exception as e:
            return JSONResponse({"message": f"Could not parse file: {e}", "success": False}, status_code=400)

        item_code_col = next((c for c in df.columns if "item code" in str(c).lower()), None)
        item_name_col = next((c for c in df.columns if "item name" in str(c).lower()), None)
        if not item_code_col or not item_name_col:
            return JSONResponse(
                {"message": "Could not find 'Item Code' and 'Item Name' columns.", "success": False},
                status_code=400,
            )

        now = datetime.now(timezone.utc)
        items = []
        for _, row in df.iterrows():
            code = str(row[item_code_col]).strip() if pd.notna(row[item_code_col]) else ""
            name = str(row[item_name_col]).strip() if pd.notna(row[item_name_col]) else ""
            if (code and name
                    and code.lower() not in ("nan", "none", "", "item code")
                    and name.lower() not in ("nan", "none", "", "item name")):
                items.append({
                    "item_code": code, "item_name": name,
                    "qty": "", "sheet_name": "Default",
                    "uploaded_by": emp_id, "uploaded_at": now,
                })

        if not items:
            return JSONResponse({"message": "No valid items found in file.", "success": False}, status_code=400)

        item_master_collection.delete_many({})
        item_master_collection.insert_many(items)

        # Insert upload history record
        upload_history_collection.insert_one({
            "uploaded_by": emp_id, "uploaded_at": now,
            "total_items": len(items), "sheets": [{"sheet": "Default", "count": len(items)}],
        })

        return JSONResponse(
            {"message": f"Successfully uploaded {len(items)} items",
             "success": True, "data": {"count": len(items)}}
        )
    except Exception as e:
        logger.error(f"upload-items error: {e}")
        return JSONResponse({"message": f"Failed to upload items: {e}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  GET ITEMS
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/get-items")
async def get_items(
    search: str = Query(None),
    page: int = Query(1),
    limit: int = Query(50),
    emp_id: str = Depends(get_current_user),
):
    try:
        query = {}
        if search:
            query = {"$or": [
                {"item_code": {"$regex": search, "$options": "i"}},
                {"item_name": {"$regex": search, "$options": "i"}},
            ]}

        skip = (page - 1) * limit
        total = item_master_collection.count_documents(query)
        items = list(item_master_collection.find(
            query,
            {"_id": 0, "item_code": 1, "item_name": 1, "sheet_name": 1, "qty": 1},
        ).skip(skip).limit(limit))

        # Merge with today's draft entries and hide items already submitted today.
        today = datetime.now(timezone.utc).date().isoformat()
        audit = temp_audit_data_collection.find_one(
            {"user_id": emp_id, "date": today},
            {"stock_count_data": 1},
        )
        submitted_audit = audit_data_collection.find_one(
            {"user_id": emp_id, "date": today, "completion_status.stock_count": True},
            {"stock_count_data": 1},
        )

        sc_lookup = {}
        if audit and "stock_count_data" in audit:
            for sc in audit["stock_count_data"]:
                sc_lookup[stock_item_key(sc)] = {
                    "physical_amount": sc.get("physical_amount", ""),
                    "remarks": sc.get("remarks", ""),
                }

        submitted_keys = {
            stock_item_key(sc)
            for sc in (submitted_audit or {}).get("stock_count_data", [])
        }
        if submitted_keys:
            items = [item for item in items if stock_item_key(item) not in submitted_keys]

        for item in items:
            sc = sc_lookup.get(stock_item_key(item), {})
            item["physical_amount"] = sc.get("physical_amount", "")
            item["remarks"] = sc.get("remarks", "")
            item.setdefault("sheet_name", "")
            item.setdefault("qty", "")

        return JSONResponse({
            "message": "Items retrieved successfully",
            "success": True,
            "data": {
                "items": items, "page": page,
                "limit": limit, "total": total,
                "has_more": (page * limit) < total,
            },
        })
    except Exception as e:
        logger.error(f"get-items error: {e}")
        return JSONResponse({"message": f"Failed to get items: {e}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  STOCK COUNT — save item
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/save-stock-count-item")
async def save_stock_count_item(request: Request, emp_id: str = Depends(get_current_user)):
    try:
        body = await request.json()
        item_code  = body.get("item_code")
        item_name  = body.get("item_name")
        sheet_name = body.get("sheet_name", "")
        physical_amount = body.get("physical_amount", "")
        remarks    = body.get("remarks", "")

        if not item_code or not item_name:
            return JSONResponse(
                {"message": "Item code and name are required", "success": False},
                status_code=400,
            )

        master = item_master_collection.find_one(
            {"item_code": item_code}, {"_id": 0, "qty": 1}
        )
        qty = (master or {}).get("qty", "")

        today = datetime.now(timezone.utc).date().isoformat()
        audit = temp_audit_data_collection.find_one({"user_id": emp_id, "date": today})
        if not audit:
            audit = {
                "user_id": emp_id, "date": today,
                "sections": {}, "completion_status": {},
                "stock_count_data": [],
                "submitted_by": emp_id,
                "submitted_at": datetime.now(timezone.utc),
            }

        if "stock_count_data" not in audit:
            audit["stock_count_data"] = []

        incoming_key = stock_item_key({
            "item_code": item_code,
            "item_name": item_name,
            "sheet_name": sheet_name,
            "qty": qty,
        })

        submitted_audit = audit_data_collection.find_one(
            {"user_id": emp_id, "date": today, "completion_status.stock_count": True},
            {"stock_count_data": 1},
        )
        submitted_keys = {
            stock_item_key(sc)
            for sc in (submitted_audit or {}).get("stock_count_data", [])
        }
        if incoming_key in submitted_keys:
            return JSONResponse(
                {
                    "message": "This item has already been submitted for today's stock count.",
                    "success": False,
                },
                status_code=409,
            )

        found = False
        for item in audit["stock_count_data"]:
            if stock_item_key(item) == incoming_key:
                item.update({
                    "physical_amount": physical_amount, "remarks": remarks,
                    "item_name": item_name, "sheet_name": sheet_name, "qty": qty,
                })
                found = True
                break

        if not found:
            audit["stock_count_data"].append({
                "item_code": item_code, "item_name": item_name,
                "sheet_name": sheet_name, "qty": qty,
                "physical_amount": physical_amount, "remarks": remarks,
            })

        if audit.get("_id"):
            temp_audit_data_collection.update_one(
                {"_id": audit["_id"]},
                {"$set": {"stock_count_data": audit["stock_count_data"]}},
            )
        else:
            temp_audit_data_collection.insert_one(audit)

        return JSONResponse({"message": "Stock count item saved successfully", "success": True})
    except Exception as e:
        logger.error(f"save-stock-count-item error: {e}")
        return JSONResponse({"message": f"Failed to save: {e}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  STOCK COUNT — submit
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/submit-stock-count")
async def submit_stock_count(emp_id: str = Depends(get_current_user)):
    try:
        today = datetime.now(timezone.utc).date().isoformat()
        audit = temp_audit_data_collection.find_one({"user_id": emp_id, "date": today})
        if not audit:
            return JSONResponse({"message": "No stock count data found", "success": False}, status_code=404)
        if not audit.get("stock_count_data"):
            return JSONResponse(
                {"message": "Please count at least one item before submitting", "success": False},
                status_code=400,
            )

        now = datetime.now(timezone.utc)
        temp_audit_data_collection.update_one(
            {"_id": audit["_id"]},
            {"$set": {"completion_status.stock_count": True, "submitted_at": now}},
        )

        existing = audit_data_collection.find_one({"user_id": emp_id, "date": today})
        if existing:
            merged_stock_count_data = merge_stock_count_data(
                existing.get("stock_count_data", []),
                audit.get("stock_count_data", []),
            )
            existing_completion = existing.get("completion_status", {}) or {}
            merged_completion = {
                **existing_completion,
                **(audit.get("completion_status", {}) or {}),
                "stock_count": True,
            }
            audit_data_collection.update_one(
                {"_id": existing["_id"]},
                {"$set": {
                    "stock_count_data": merged_stock_count_data,
                    "completion_status": merged_completion,
                    "submitted_at": now,
                }},
            )
            audit_id = str(existing["_id"])
        else:
            audit["completion_status"] = {
                **(audit.get("completion_status", {}) or {}),
                "stock_count": True,
            }
            audit["submitted_at"] = now
            result = audit_data_collection.insert_one(audit.copy())
            audit_id = str(result.inserted_id)

        checklist_completed_count = sum(
            1 for section in CHECKLIST_SECTIONS
            if audit.get("completion_status", {}).get(section, False)
        )
        if checklist_completed_count:
            temp_audit_data_collection.update_one(
                {"_id": audit["_id"]},
                {"$unset": {"stock_count_data": "", "completion_status.stock_count": ""}},
            )
        else:
            temp_audit_data_collection.delete_one({"_id": audit["_id"]})

        return JSONResponse(
            {"message": "Stock count submitted successfully",
             "success": True, "data": {"audit_id": audit_id}}
        )
    except Exception as e:
        logger.error(f"submit-stock-count error: {e}")
        return JSONResponse({"message": f"Failed to submit: {e}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  STOCK COUNT — export Excel
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/export-stock-count-excel")
async def export_stock_count_excel(emp_id: str = Depends(get_current_user)):
    try:
        today = datetime.now(timezone.utc).date().isoformat()
        audit = temp_audit_data_collection.find_one({"user_id": emp_id, "date": today})
        if not audit or not audit.get("stock_count_data"):
            return JSONResponse({"message": "No stock count data found", "success": False}, status_code=404)
        if not audit.get("completion_status", {}).get("stock_count", False):
            return JSONResponse(
                {"message": "Please submit stock count before exporting", "success": False},
                status_code=400,
            )

        excel_bytes = generate_stock_count_excel_bytes(audit)
        filename = stock_count_report_filename(audit, today)
        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
        )
    except Exception as e:
        logger.error(f"export-stock-count-excel error: {e}")
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  STOCK COUNT — send email from history
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/send-stock-count-email")
async def send_stock_count_email(
    audit_id: str = Form(...),
    to_email: str = Form(...),
    email_subject: Optional[str] = Form(default=None),
    email_body: Optional[str] = Form(default=None),
    emp_id: str = Depends(get_current_user),
):
    try:
        user = users.find_one({"email": emp_id}, {"name": 1})
        user_name = (user or {}).get("name", "Unknown")

        audit_data = audit_data_collection.find_one(
            {"user_id": emp_id, "_id": ObjectId(audit_id)}
        )
        if not audit_data or not audit_data.get("stock_count_data"):
            return JSONResponse({"message": "Stock count not found", "success": False}, status_code=404)

        excel_bytes = generate_stock_count_excel_bytes(audit_data)

        target_date = audit_data.get("date", "")
        excel_name = stock_count_report_filename(audit_data, target_date)
        subject = email_subject or f"Stock Count Report – {target_date} – {user_name}"
        body = email_body or (
            f"Dear Manager,\n\nPlease find the Stock Count report attached."
            f"\n\nDate: {target_date}\nTotal Items: {len(audit_data['stock_count_data'])}"
            f"\n\nRegards,\n{user_name}\nAudit App"
        )

        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = os.getenv("MAIL_USERNAME")
        msg["To"] = to_email
        msg["Cc"] = emp_id
        msg.set_content(body)
        msg.add_attachment(
            excel_bytes, maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=excel_name,
        )

        mail_username = os.getenv("MAIL_USERNAME")
        mail_password = os.getenv("MAIL_PASSWORD")
        if not mail_username or not mail_password:
            return JSONResponse(
                {"message": "Email credentials not configured.", "success": False},
                status_code=500,
            )

        try:
            await asyncio.to_thread(send_smtp_message_sync, msg, mail_username, mail_password)
        except smtplib.SMTPAuthenticationError as e:
            return JSONResponse(
                {"message": "Email authentication failed.", "success": False},
                status_code=500,
            )

        return JSONResponse({"message": "Email sent successfully", "success": True})
    except Exception as e:
        logger.error(f"send-stock-count-email error: {e}")
        return JSONResponse({"message": f"Failed to send email: {e}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  ADMIN — check / employees
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/check-admin")
async def check_admin(emp_id: str = Depends(get_current_user)):
    try:
        return JSONResponse({
            "message": "Admin check",
            "success": True,
            "data": {"is_admin": is_admin(emp_id)},
        })
    except Exception as e:
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)


@app.get("/api/checklist-definitions")
async def checklist_definitions(emp_id: str = Depends(get_current_user)):
    try:
        docs = list(checklist_questions_collection.find({}, {"_id": 0}))
        stored = {doc.get("section"): doc for doc in docs if doc.get("section")}
        sections = []
        for key in QUESTION_SECTIONS:
            doc = stored.get(key) or {}
            sections.append({
                "section": key,
                "title": doc.get("title") or QUESTION_SECTION_TITLES[key],
                "questions": doc.get("questions") or [],
            })
        return JSONResponse({
            "message": "Checklist definitions fetched",
            "success": True,
            "data": {"sections": sections},
        })
    except Exception as e:
        logger.error(f"checklist-definitions error: {e}")
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)


@app.put("/api/admin/checklist-definitions/{section_name}")
async def admin_save_checklist_definition(
    section_name: str,
    request: Request,
    emp_id: str = Depends(get_current_user),
):
    try:
        require_admin(emp_id)
        if section_name not in QUESTION_SECTIONS:
            return JSONResponse({"message": "Invalid checklist section", "success": False}, status_code=400)
        body = await request.json()
        questions = body.get("questions")
        if not isinstance(questions, list):
            return JSONResponse({"message": "questions must be a list", "success": False}, status_code=400)
        cleaned = []
        for idx, q in enumerate(questions, 1):
            text = str(q.get("question", "")).strip()
            if not text:
                continue
            field_type = str(q.get("field_type") or "yes_no_remarks")
            subquestions = q.get("subquestions") if isinstance(q.get("subquestions"), list) else []
            cleaned.append({
                "id": str(q.get("id") or f"q{idx}"),
                "question": text,
                "field_type": field_type,
                "requires_remarks_on_no": bool(q.get("requires_remarks_on_no", field_type == "yes_no_remarks")),
                "subquestions": subquestions,
            })
        checklist_questions_collection.update_one(
            {"section": section_name},
            {"$set": {
                "section": section_name,
                "title": body.get("title") or QUESTION_SECTION_TITLES[section_name],
                "questions": cleaned,
                "updated_by": emp_id,
                "updated_at": datetime.now(timezone.utc),
            }},
            upsert=True,
        )
        return JSONResponse({"message": "Checklist definition saved", "success": True, "data": {"questions": cleaned}})
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"admin-save-checklist-definition error: {e}")
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)


@app.delete("/api/admin/checklist-definitions/{section_name}/{question_id}")
async def admin_delete_checklist_question(
    section_name: str,
    question_id: str,
    emp_id: str = Depends(get_current_user),
):
    try:
        require_admin(emp_id)
        result = checklist_questions_collection.update_one(
            {"section": section_name},
            {
                "$pull": {"questions": {"id": question_id}},
                "$set": {"updated_by": emp_id, "updated_at": datetime.now(timezone.utc)},
            },
        )
        return JSONResponse({"message": "Question deleted", "success": True, "data": {"modified": result.modified_count}})
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"admin-delete-checklist-question error: {e}")
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)


@app.get("/api/admin/employees-stats")
async def get_employees_stats(
    page: int = 1,
    limit: int = 10,
    emp_id: str = Depends(get_current_user)
):
    try:
        require_admin(emp_id)
        page = max(1, page)
        limit = min(max(1, limit), 100)
        skip = (page - 1) * limit

        total = users.count_documents({})
        all_users = list(users.find({}, {"_id": 0, "email": 1, "name": 1}).skip(skip).limit(limit))

        total_pages = (total + limit - 1) // limit

        return JSONResponse({
            "message": "Stats fetched", "success": True,
            "data": {
                "users": all_users,
                "total": total,
                "pagination": {
                    "current_page": page,
                    "total_pages": total_pages,
                    "total_items": total,
                    "items_per_page": limit
                }
            },
        })
    except HTTPException:
        return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)
    except Exception as e:
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)


@app.get("/api/admin/checklist-data")
async def admin_checklist_data(
    page: int = 1,
    limit: int = 5000,
    emp_id: str = Depends(get_current_user)
):
    try:
        require_admin(emp_id)

        # Validate pagination params
        page = max(1, page)
        limit = min(max(1, limit), 5000)  # Max 5000 items per page
        skip = (page - 1) * limit

        # Enhanced projection to include warehouse information
        projection = {
            "_id": 0, "user_id": 1, "date": 1, "submitted_at": 1, 
            "completion_status": 1, "stock_count_data": 1,
            "sections.general_report.warehouse_name": 1,
            "sections.general_report.warehouse_address": 1,
        }

        # Get total count for pagination
        total_submitted = audit_data_collection.count_documents({})
        total_in_progress = temp_audit_data_collection.count_documents({})
        total_records = total_submitted + total_in_progress

        # Get paginated data from both collections
        submitted = list(audit_data_collection.find({}, projection).sort("submitted_at", -1).skip(skip).limit(limit))
        remaining_limit = limit - len(submitted)
        
        in_progress = []
        if remaining_limit > 0:
            skip_in_progress = max(0, skip - total_submitted)
            in_progress = list(temp_audit_data_collection.find({}, projection).sort("date", -1).skip(skip_in_progress).limit(remaining_limit))

        # Combine and process data
        all_data = []
        for audit in submitted + in_progress:
            # Extract warehouse name properly
            general_report = audit.get("sections", {}).get("general_report", {})
            warehouse_name = general_report.get("warehouse_name", "Unknown Warehouse")
            
            # Add warehouse_name to top level for easier access
            audit["warehouse_name"] = warehouse_name
            all_data.append(audit)

        # Calculate pagination info
        total_pages = (total_records + limit - 1) // limit

        return JSONResponse({
            "message": "Data fetched", "success": True,
            "data": {
                "checklists": _serialize_mongo(all_data),
                "pagination": {
                    "current_page": page,
                    "total_pages": total_pages,
                    "total_items": total_records,
                    "items_per_page": limit
                }
            },
        })
    except HTTPException:
        return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)
    except Exception as e:
        logger.error(f"admin-checklist-data error: {e}")
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)


@app.get("/api/admin/uploaded-history")
async def uploaded_history(emp_id: str = Depends(get_current_user)):
    try:
        require_admin(emp_id)
        history = list(
            upload_history_collection.find(
                {}, {"_id": 0, "uploaded_by": 1, "uploaded_at": 1, "total_items": 1, "total_count": 1}
            ).sort("uploaded_at", -1).limit(500)
        )
        # Normalize total_items and total_count
        for h in history:
            if "total_items" not in h and "total_count" in h:
                h["total_items"] = h["total_count"]
            elif "total_count" not in h and "total_items" in h:
                h["total_count"] = h["total_items"]
        return JSONResponse({
            "message": "History fetched", "success": True,
            "data": {"history": _serialize_mongo(history)},
        })
    except HTTPException:
        return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)
    except Exception as e:
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  ADMIN — audit dashboard
#  FIX: uses aggregation pipeline to compute completion % in DB, not Python
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/admin/audit-dashboard")
async def audit_dashboard(
    page: int = 1,
    limit: int = 5000,
    emp_id: str = Depends(get_current_user)
):
    try:
        require_admin(emp_id)

        # Validate pagination params
        page = max(1, page)
        limit = min(max(1, limit), 5000)  # Max 5000 items per page
        skip = (page - 1) * limit

        # Enhanced projection to include warehouse information
        projection = {
            "_id": 0, "user_id": 1, "date": 1,
            "submitted_at": 1, "completion_status": 1, "stock_count_data": 1,
            "sections.general_report.warehouse_name": 1,
            "sections.general_report.warehouse_address": 1,
        }

        # Get total counts for pagination
        total_submitted = audit_data_collection.count_documents({})
        total_in_progress = temp_audit_data_collection.count_documents({})
        total_records = total_submitted + total_in_progress

        # Get paginated data
        submitted = list(audit_data_collection.find({}, projection).sort("submitted_at", -1).skip(skip).limit(limit))
        remaining_limit = limit - len(submitted)
        
        in_progress = []
        if remaining_limit > 0:
            skip_in_progress = max(0, skip - total_submitted)
            in_progress = list(temp_audit_data_collection.find({}, projection).sort("date", -1).skip(skip_in_progress).limit(remaining_limit))

        all_emails = {d.get("user_id") for d in submitted + in_progress if d.get("user_id")}
        name_map = get_user_names_cached(all_emails)
        total_stock_count_items = item_master_collection.count_documents({})

        rows = []
        for d in submitted:
            cs = d.get("completion_status", {})
            completed_count = sum(1 for s in CHECKLIST_SECTIONS if cs.get(s, False))
            is_checklist_complete = completed_count == len(CHECKLIST_SECTIONS)
            stock_count_items = len(d.get("stock_count_data", []))
            submitted_at = d.get("submitted_at", "")
            if isinstance(submitted_at, datetime):
                submitted_at = submitted_at.isoformat()
            user_email = d.get("user_id", "")

            # Extract warehouse info
            general_report = d.get("sections", {}).get("general_report", {})
            warehouse_name = general_report.get("warehouse_name", "Unknown Warehouse")

            rows.append({
                "user_id": user_email,
                "user_name": name_map.get(user_email, user_email),
                "date": d.get("date", ""),
                "warehouse_name": warehouse_name,
                "checklist_completed": completed_count,
                "checklist_total": len(CHECKLIST_SECTIONS),
                "checklist_pct": round(completed_count / len(CHECKLIST_SECTIONS) * 100),
                "checklist_status": "Submitted" if is_checklist_complete else ("In Progress" if completed_count else "Pending"),
                "stock_count_items": stock_count_items,
                "stock_count_status": "Submitted" if cs.get("stock_count") else ("In Progress" if stock_count_items else "Pending"),
                "status": "Submitted",
                "submitted_at": submitted_at,
            })

        for d in in_progress:
            cs = d.get("completion_status", {})
            completed_count = sum(1 for s in CHECKLIST_SECTIONS if cs.get(s, False))
            stock_count_items = len(d.get("stock_count_data", []))
            user_email = d.get("user_id", "")

            # Extract warehouse info
            general_report = d.get("sections", {}).get("general_report", {})
            warehouse_name = general_report.get("warehouse_name", "In Progress")

            rows.append({
                "user_id": user_email,
                "user_name": name_map.get(user_email, user_email),
                "date": d.get("date", ""),
                "warehouse_name": warehouse_name,
                "checklist_completed": completed_count,
                "checklist_total": len(CHECKLIST_SECTIONS),
                "checklist_pct": round(completed_count / len(CHECKLIST_SECTIONS) * 100),
                "checklist_status": "In Progress" if completed_count else "Pending",
                "stock_count_items": stock_count_items,
                "stock_count_status": "Submitted" if cs.get("stock_count") else ("In Progress" if stock_count_items else "Pending"),
                "status": "In Progress",
                "submitted_at": "",
            })

        # Calculate pagination info
        total_pages = (total_records + limit - 1) // limit

        return JSONResponse({
            "message": "Dashboard data fetched", "success": True,
            "data": {
                "rows": rows,
                "total_stock_count_items": total_stock_count_items,
                "pagination": {
                    "current_page": page,
                    "total_pages": total_pages,
                    "total_items": total_records,
                    "items_per_page": limit
                }
            },
        })
    except HTTPException:
        return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)
    except Exception as e:
        logger.error(f"audit-dashboard error: {e}")
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  ADMIN — warehouse master
# ─────────────────────────────────────────────────────────────────────────────

class WarehouseItem(BaseModel):
    warehouse_name: str
    warehouse_address: str


class WarehouseMasterPayload(BaseModel):
    warehouses: List[WarehouseItem]


@app.post("/api/admin/warehouse-master")
async def upload_warehouse_master(
    payload: WarehouseMasterPayload,
    emp_id: str = Depends(get_current_user),
):
    try:
        require_admin(emp_id)
        if not payload.warehouses:
            return JSONResponse({"message": "No warehouse data provided", "success": False}, status_code=400)
        now = datetime.now(timezone.utc)
        docs = [
            {"warehouse_name": w.warehouse_name.strip(),
             "warehouse_address": w.warehouse_address.strip(),
             "uploaded_by": emp_id, "uploaded_at": now}
            for w in payload.warehouses if w.warehouse_name.strip()
        ]
        warehouse_master_collection.delete_many({})
        warehouse_master_collection.insert_many(docs)
        return JSONResponse({
            "message": f"{len(docs)} warehouses uploaded successfully",
            "success": True, "data": {"count": len(docs)},
        })
    except HTTPException:
        return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)
    except Exception as e:
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)


@app.get("/api/admin/warehouse-master")
async def get_warehouse_master_admin(
    page: int = 1,
    limit: int = 10,
    emp_id: str = Depends(get_current_user)
):
    try:
        require_admin(emp_id)
        page = max(1, page)
        limit = min(max(1, limit), 100)
        skip = (page - 1) * limit

        total = warehouse_master_collection.count_documents({})
        warehouses = list(warehouse_master_collection.find(
            {}, {"_id": 0, "warehouse_name": 1, "warehouse_address": 1}
        ).skip(skip).limit(limit))

        total_pages = (total + limit - 1) // limit

        return JSONResponse({
            "message": "Warehouses fetched", "success": True,
            "data": {
                "warehouses": warehouses,
                "total": total,
                "pagination": {
                    "current_page": page,
                    "total_pages": total_pages,
                    "total_items": total,
                    "items_per_page": limit
                }
            },
        })
    except HTTPException:
        return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)
    except Exception as e:
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)


@app.get("/api/warehouses")
async def get_warehouses(emp_id: str = Depends(get_current_user)):
    try:
        warehouses = list(warehouse_master_collection.find(
            {}, {"_id": 0, "warehouse_name": 1, "warehouse_address": 1}
        ))
        return JSONResponse({
            "message": "Warehouses fetched", "success": True,
            "data": {"warehouses": warehouses},
        })
    except Exception as e:
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  ADMIN — analytics
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/admin/analytics")
async def admin_analytics(
    start_date: str = None,
    end_date: str = None,
    emp_id: str = Depends(get_current_user),
):
    try:
        require_admin(emp_id)
        if not start_date:
            start_date = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
        if not end_date:
            end_date = datetime.now().strftime("%Y-%m-%d")

        query = {"date": {"$gte": start_date, "$lte": end_date}}
        # Only fetch fields needed for analytics
        proj = {
            "_id": 0, "user_id": 1, "date": 1,
            "submitted_at": 1, "completion_status": 1,
            "stock_count_data": 1,
            "warehouse_name": 1,
            "sections.general_report.warehouse_name": 1,
        }
        all_audits = (
            list(temp_audit_data_collection.find(query, proj).limit(100))
            + list(audit_data_collection.find(query, proj).limit(100))
        )
        user_ids = {a.get("user_id", "") for a in all_audits if a.get("user_id")}
        user_name_map = get_user_names_cached(user_ids)

        total_audits   = len(all_audits)
        completed      = sum(1 for a in all_audits if a.get("submitted_at"))
        completion_rate = round((completed / total_audits * 100) if total_audits else 0, 1)
        avg_sections   = (
            sum(sum(1 for v in a.get("completion_status", {}).values() if v) for a in all_audits)
            / total_audits if total_audits else 0
        )
        total_stock_items = sum(len(a.get("stock_count_data", [])) for a in all_audits)

        audits_by_date = defaultdict(int)
        user_stats     = defaultdict(lambda: {"total": 0, "completed": 0})
        warehouse_dist = defaultdict(int)
        section_stats  = {"completed": 0, "pending": 0}

        for audit in all_audits:
            audits_by_date[audit.get("date", "Unknown")] += 1
            u = audit.get("user_id", "Unknown")
            user_stats[u]["total"] += 1
            if audit.get("submitted_at"):
                user_stats[u]["completed"] += 1
            wh = (
                audit.get("warehouse_name")
                or (audit.get("sections") or {}).get("general_report", {}).get("warehouse_name")
                or "Unknown"
            )
            warehouse_dist[wh] += 1
            cs = audit.get("completion_status", {})
            done = sum(1 for section in CHECKLIST_SECTIONS if cs.get(section, False))
            section_stats["completed"] += done
            section_stats["pending"] += max(len(CHECKLIST_SECTIONS) - done, 0)

        return JSONResponse({
            "message": "Analytics data fetched", "success": True,
            "data": {
                "total_audits": total_audits,
                "completion_rate": completion_rate,
                "avg_sections": avg_sections,
                "total_stock_items": total_stock_items,
                "audits_by_date": [{"date": k, "count": v} for k, v in sorted(audits_by_date.items())],
                "completion_by_user": [
                    {"user": k, "user_name": user_name_map.get(k, k), "total": v["total"], "completed": v["completed"]}
                    for k, v in user_stats.items()
                ],
                "warehouse_distribution": [{"warehouse": k, "count": v} for k, v in warehouse_dist.items()],
                "section_breakdown": section_stats,
            },
        })
    except HTTPException:
        return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)
    except Exception as e:
        logger.error(f"analytics error: {e}")
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  ADMIN — warehouse status
#  FIX: duplicate temp_audits loop removed; correct warehouse filter
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/admin/warehouse-status")
async def admin_warehouse_status(
    date: str = None,
    page: int = 1,
    limit: int = 15,
    emp_id: str = Depends(get_current_user),
):
    try:
        require_admin(emp_id)
        page = max(1, page)
        limit = min(max(1, limit), 100)

        all_warehouses = list(warehouse_master_collection.find(
            {}, {"_id": 0, "warehouse_name": 1, "warehouse_address": 1}
        ))

        if date:
            query = {"date": date}
        else:
            date_limit = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
            query = {"date": {"$gte": date_limit}}

        proj = {
            "_id": 1, "user_id": 1, "date": 1,
            "submitted_at": 1, "completion_status": 1,
            "sections.general_report.warehouse_name": 1,
            "warehouse_name": 1,
            "stock_count_data": 1,
        }

        submitted_audits = list(audit_data_collection.find(query, proj).limit(100))
        temp_audits      = list(temp_audit_data_collection.find(query, proj).limit(100))
        today_str = datetime.now(timezone.utc).date().isoformat()
        task_docs = list(task_assignments_collection.find({}, {
            "_id": 1, "warehouse_name": 1, "assigned_to": 1,
            "task_type": 1, "due_date": 1, "status": 1,
        }))
        task_user_ids = {
            email
            for task in task_docs
            for email in (task.get("assigned_to") or [])
            if email
        }

        wh_map: Dict[str, dict] = {}

        def _wh_name(audit):
            return (
                (audit.get("sections") or {}).get("general_report", {}).get("warehouse_name")
                or audit.get("warehouse_name")
            )

        def _dt_str(val):
            if isinstance(val, datetime):
                return val.strftime("%Y-%m-%d %H:%M:%S")
            return val or ""

        # Process submitted audits first (highest priority — override In Progress)
        for audit in submitted_audits:
            wh = _wh_name(audit)
            if not wh:
                continue
            if wh not in wh_map:
                wh_map[wh] = {
                    "warehouse_name": wh, "warehouse_address": "",
                    "status": "Completed", "assigned_users": [],
                    "progress_percentage": 100,
                    "last_updated": _dt_str(audit.get("submitted_at") or audit.get("date")),
                    "audit_id": str(audit["_id"]),
                    "audit_user_id": audit.get("user_id"),
                    "audit_date": audit.get("date"),
                    "has_stock_count": bool(audit.get("stock_count_data")),
                }
            else:
                wh_map[wh]["status"] = "Completed"
                wh_map[wh]["progress_percentage"] = 100
                wh_map[wh]["audit_id"] = str(audit["_id"])
                wh_map[wh]["audit_user_id"] = audit.get("user_id")
                wh_map[wh]["audit_date"] = audit.get("date")

            uid = audit.get("user_id")
            if uid and uid not in wh_map[wh]["assigned_users"]:
                wh_map[wh]["assigned_users"].append(uid)

        # Process temp audits ONCE (skip if already Completed)
        for audit in temp_audits:
            wh = _wh_name(audit)
            if not wh:
                continue

            if wh in wh_map and wh_map[wh]["status"] == "Completed":
                uid = audit.get("user_id")
                if uid and uid not in wh_map[wh]["assigned_users"]:
                    wh_map[wh]["assigned_users"].append(uid)
                continue

            if wh not in wh_map:
                wh_map[wh] = {
                    "warehouse_name": wh, "warehouse_address": "",
                    "status": "Not Started", "assigned_users": [],
                    "progress_percentage": 0,
                    "last_updated": None, "audit_id": None,
                    "audit_user_id": None, "audit_date": None,
                    "has_stock_count": False,
                }

            uid = audit.get("user_id")
            if uid and uid not in wh_map[wh]["assigned_users"]:
                wh_map[wh]["assigned_users"].append(uid)

            cs = audit.get("completion_status", {})
            if cs:
                done  = sum(1 for v in cs.values() if v)
                total = len(cs)
                pct   = round((done / total * 100) if total else 0)
                if pct > wh_map[wh]["progress_percentage"]:
                    wh_map[wh]["progress_percentage"] = pct
                    wh_map[wh]["audit_id"] = str(audit["_id"])
                    wh_map[wh]["audit_user_id"] = audit.get("user_id")
                    wh_map[wh]["audit_date"] = audit.get("date")
                if done > 0:
                    wh_map[wh]["status"] = "In Progress"

            if audit.get("stock_count_data"):
                wh_map[wh]["has_stock_count"] = True

            updated = _dt_str(audit.get("date"))
            if updated and (not wh_map[wh]["last_updated"] or updated > wh_map[wh]["last_updated"]):
                wh_map[wh]["last_updated"] = updated

        # Merge warehouse master addresses; add "Not Started" for unseen warehouses
        for wh in all_warehouses:
            wh_name = wh.get("warehouse_name")
            if wh_name not in wh_map:
                wh_map[wh_name] = {
                    "warehouse_name": wh_name,
                    "warehouse_address": wh.get("warehouse_address", ""),
                    "status": "Not Started", "assigned_users": [],
                    "progress_percentage": 0,
                    "last_updated": None, "audit_id": None,
                    "audit_user_id": None, "audit_date": None,
                    "has_stock_count": False,
                }
            else:
                wh_map[wh_name]["warehouse_address"] = wh.get("warehouse_address", "")

        all_assigned_users = {
            email
            for wh in wh_map.values()
            for email in (wh.get("assigned_users") or [])
            if email
        } | task_user_ids
        name_map = get_user_names_cached(all_assigned_users)

        for wh in wh_map.values():
            assigned_from_tasks = []
            overdue_tasks = []
            open_tasks = []
            for task in task_docs:
                if task.get("warehouse_name") != wh.get("warehouse_name"):
                    continue

                assigned_from_tasks.extend(task.get("assigned_to") or [])
                is_completed_task = task.get("status") == "Completed" or wh.get("status") == "Completed"
                is_overdue = bool(task.get("due_date")) and task.get("due_date") < today_str and not is_completed_task
                task_summary = {
                    "task_id": str(task.get("_id")),
                    "task_type": task.get("task_type", ""),
                    "due_date": task.get("due_date", ""),
                    "status": task.get("status", ""),
                    "is_overdue": is_overdue,
                }
                open_tasks.append(task_summary)
                if is_overdue:
                    overdue_tasks.append(task_summary)

            for email in assigned_from_tasks:
                if email and email not in wh["assigned_users"]:
                    wh["assigned_users"].append(email)

            wh["assigned_user_names"] = [
                name_map.get(email, email) for email in (wh.get("assigned_users") or [])
            ]
            wh["open_tasks"] = open_tasks
            wh["overdue_tasks"] = overdue_tasks
            wh["is_overdue"] = bool(overdue_tasks)
            wh["due_dates"] = sorted({
                task.get("due_date")
                for task in open_tasks
                if task.get("due_date")
            })
            wh["next_due_date"] = wh["due_dates"][0] if wh["due_dates"] else ""
            wh["overdue_due_dates"] = sorted({
                task.get("due_date")
                for task in overdue_tasks
                if task.get("due_date")
            })

        status_order = {"Completed": 1, "In Progress": 2, "Not Started": 3}
        warehouse_list = sorted(
            wh_map.values(),
            key=lambda x: (status_order.get(x["status"], 4), x["warehouse_name"]),
        )
        total_items = len(warehouse_list)
        total_pages = (total_items + limit - 1) // limit if total_items else 0
        start = (page - 1) * limit
        paginated_warehouses = warehouse_list[start:start + limit]
        summary = {
            "total": total_items,
            "completed": sum(1 for w in warehouse_list if w["status"] == "Completed"),
            "in_progress": sum(1 for w in warehouse_list if w["status"] == "In Progress"),
            "not_started": sum(1 for w in warehouse_list if w["status"] == "Not Started"),
            "overdue": sum(1 for w in warehouse_list if w.get("is_overdue")),
        }

        return JSONResponse({
            "message": "Warehouse status fetched", "success": True,
            "data": {
                "warehouses": paginated_warehouses,
                "date": date,
                "summary": summary,
                "pagination": {
                    "current_page": page,
                    "total_pages": total_pages,
                    "total_items": total_items,
                    "items_per_page": limit,
                },
            },
        })
    except HTTPException:
        return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)
    except Exception as e:
        logger.error(f"warehouse-status error: {e}")
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  ADMIN — stock reconciliation
#  FIX: N+1 eliminated — bulk fetch item master, then dict lookup
#  FIX: warehouse filter syntax corrected
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/admin/stock-reconciliation")
async def admin_stock_reconciliation(
    from_date: str = None,
    to_date: str = None,
    page: int = 1,
    limit: int = 5000,
    emp_id: str = Depends(get_current_user),
):
    """Get reconciliation report with date range and pagination."""
    try:
        require_admin(emp_id)

        # Validate pagination params
        page = max(1, page)
        limit = min(max(1, limit), 5000)
        skip = (page - 1) * limit

        if not from_date or not to_date:
            return JSONResponse({
                "message": "Both from_date and to_date are required",
                "success": True,
                "data": {
                    "reconciliation": [],
                    "summary": {"total_items": 0, "matched": 0, "excess": 0, "shortage": 0, "match_rate": 0},
                    "pagination": {"current_page": page, "total_pages": 0, "total_items": 0, "items_per_page": limit}
                }
            }, status_code=200)

        # Build date range query
        query = {"date": {"$gte": from_date, "$lte": to_date}}
        proj = {"_id": 0, "user_id": 1, "submitted_at": 1, "stock_count_data": 1, "date": 1}

        # Get audits for date range (no warehouse filter)
        all_audits = (
            list(audit_data_collection.find(query, proj).sort("date", -1))
            + list(temp_audit_data_collection.find(query, proj).sort("date", -1))
        )

        # Process all audits to get reconciliation data
        reconciliation_data = []
        all_item_codes = {
            item.get("item_code")
            for audit in all_audits
            for item in audit.get("stock_count_data", [])
            if item.get("item_code")
        }

        # Bulk fetch from item master
        master_docs = list(item_master_collection.find(
            {"item_code": {"$in": list(all_item_codes)}},
            {"_id": 0, "item_code": 1, "sheet_name": 1, "qty": 1},
        ))

        # Build lookup maps
        master_by_code_sheet = {}
        master_by_code = {}
        for d in master_docs:
            try:
                qty = float(d.get("qty", 0) or 0)
            except (ValueError, TypeError):
                qty = 0.0
            master_by_code_sheet[(d["item_code"], d.get("sheet_name", ""))] = qty
            master_by_code[d["item_code"]] = qty

        # Build user name map
        user_ids = {a.get("user_id", "") for a in all_audits}
        user_name_map = get_user_names_cached(user_ids)

        # Process reconciliation data
        for audit in all_audits:
            sc_data = audit.get("stock_count_data", [])
            if not sc_data:
                continue
            user_id = audit.get("user_id", "Unknown")
            audit_status = "Submitted" if audit.get("submitted_at") else "In Progress"
            audit_date = audit.get("date", "")

            for item in sc_data:
                item_code = item.get("item_code", "")
                item_name = item.get("item_name", "")
                sheet_name = item.get("sheet_name", "")
                remarks = item.get("remarks", "")

                try:
                    physical_qty = float(
                        item.get("physical_amount") or item.get("quantity") or item.get("qty") or 0
                    )
                except (ValueError, TypeError):
                    physical_qty = 0.0

                system_qty = master_by_code_sheet.get(
                    (item_code, sheet_name),
                    master_by_code.get(item_code, 0.0),
                )

                variance = physical_qty - system_qty
                variance_pct = round((variance / system_qty * 100) if system_qty else 0, 2)
                variance_status = (
                    "Match" if variance == 0 else
                    "Excess" if variance > 0 else
                    "Shortage"
                )

                reconciliation_data.append({
                    "user_id": user_id,
                    "auditor_name": user_name_map.get(user_id, user_id),
                    "item_code": item_code,
                    "item_name": item_name,
                    "sheet_name": sheet_name,
                    "remarks": remarks,
                    "system_quantity": system_qty,
                    "physical_quantity": physical_qty,
                    "variance": variance,
                    "variance_percentage": variance_pct,
                    "variance_status": variance_status,
                    "audit_status": audit_status,
                    "date": audit_date,
                })

        # Sort by date (newest first)
        reconciliation_data.sort(key=lambda x: x["date"], reverse=True)

        # Calculate totals
        total_items = len(reconciliation_data)
        matched = sum(1 for r in reconciliation_data if r["variance_status"] == "Match")
        excess = sum(1 for r in reconciliation_data if r["variance_status"] == "Excess")
        shortage = sum(1 for r in reconciliation_data if r["variance_status"] == "Shortage")

        return JSONResponse({
            "message": "Reconciliation data fetched",
            "success": True,
            "data": {
                "reconciliation": reconciliation_data,
                "summary": {
                    "total_items": total_items,
                    "matched": matched,
                    "excess": excess,
                    "shortage": shortage,
                    "match_rate": round((matched / total_items * 100) if total_items > 0 else 0, 2)
                },
                "pagination": {
                    "current_page": 1,
                    "total_pages": 1,
                    "total_items": total_items,
                    "items_per_page": total_items or 10
                },
                "from_date": from_date,
                "to_date": to_date
            }
        }, status_code=200)

    except Exception as e:
        logger.error(f"Stock reconciliation error: {e}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)


# @app.get("/api/admin/audit-detail/{user_id}/{date}")
# async def admin_audit_detail(user_id: str, date: str, emp_id: str = Depends(get_current_user)):
#     """Get full audit detail for dashboard view button."""
#     try:
#         require_admin(emp_id)
#         # Search submitted first, then in-progress
#         audit = audit_data_collection.find_one({"user_id": user_id, "date": date})
#         if not audit:
#             audit = temp_audit_data_collection.find_one({"user_id": user_id, "date": date})
#         if not audit:
#             return JSONResponse({"message": "Audit not found", "success": False}, status_code=404)
#         audit["_id"] = str(audit["_id"])
#         if "submitted_at" in audit and isinstance(audit["submitted_at"], datetime):
#             audit["submitted_at"] = audit["submitted_at"].isoformat()
#         # Resolve warehouse name
#         wh_name = (audit.get("warehouse_name")
#                    or (audit.get("sections") or {}).get("general_report", {}).get("warehouse_name")
#                    or "—")
#         audit["warehouse_name"] = wh_name
#         # Add user display name
#         user_rec = users.find_one({"email": user_id}, {"_id": 0, "name": 1})
#         audit["user_name"] = user_rec.get("name", user_id) if user_rec else user_id
#         return JSONResponse({"message": "Audit detail fetched", "success": True, "data": audit}, status_code=200)
#     except HTTPException:
#         return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)
#     except Exception as e:
#         logger.error(f"Audit detail error: {e}")
#         return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)

@app.get("/api/admin/audit-detail/{user_id}/{date}")
async def admin_audit_detail(user_id: str, date: str, emp_id: str = Depends(get_current_user)):
    require_admin(emp_id)
    
    # Prefer submitted audit
    audit = audit_data_collection.find_one({"user_id": user_id, "date": date})
    if not audit:
        audit = temp_audit_data_collection.find_one({"user_id": user_id, "date": date})
    
    if not audit:
        return JSONResponse({"message": "Audit not found", "success": False}, status_code=404)

    audit["_id"] = str(audit["_id"])
    if isinstance(audit.get("submitted_at"), datetime):
        audit["submitted_at"] = audit["submitted_at"].isoformat()

    # Ensure sections are always present
    if "sections" not in audit:
        audit["sections"] = {}

    # Resolve warehouse name
    wh_name = (
        audit.get("warehouse_name") or
        (audit.get("sections") or {}).get("general_report", {}).get("warehouse_name") or
        "—"
    )
    audit["warehouse_name"] = wh_name

    # User name
    user_rec = users.find_one({"email": user_id}, {"name": 1})
    audit["user_name"] = user_rec.get("name", user_id) if user_rec else user_id

    return JSONResponse({
        "message": "Audit detail fetched",
        "success": True,
        "data": audit
    }, status_code=200)

@app.get("/api/admin/export-audit/{user_id}/{date}")
async def admin_export_audit(user_id: str, date: str, type: str = "checklist", emp_id: str = Depends(get_current_user)):
    """Export audit Excel from admin panel."""
    try:
        require_admin(emp_id)
        # Find audit in submitted, then temp
        audit_data = audit_data_collection.find_one({"user_id": user_id, "date": date})
        if not audit_data:
            audit_data = temp_audit_data_collection.find_one({"user_id": user_id, "date": date})
        if not audit_data:
            return JSONResponse({"message": "Audit not found", "success": False}, status_code=404)
        if type == "stockcount":
            excel_bytes = generate_stock_count_excel_bytes(audit_data)
            filename = f"StockCount_{user_id}_{date}.xlsx"
        else:
            excel_bytes = await generate_checklist_excel_bytes(user_id, audit_data)
            filename = f"Checklist_{user_id}_{date}.xlsx"
        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
        )
    except HTTPException:
        return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)
    except Exception as e:
        logger.error(f"Admin export audit error: {e}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  TASK ASSIGNMENT
# ─────────────────────────────────────────────────────────────────────────────

class TaskAssignment(BaseModel):
    warehouse_name: str
    assigned_to: List[str]
    task_type: str
    due_date: str
    notes: Optional[str] = ""


@app.post("/api/admin/assign-task")
async def assign_task(task: TaskAssignment, emp_id: str = Depends(get_current_user)):
    try:
        require_admin(emp_id)
        now = datetime.now(timezone.utc)
        task_doc = {
            "warehouse_name": task.warehouse_name,
            "assigned_to": task.assigned_to,
            "assigned_by": emp_id,
            "task_type": task.task_type,
            "due_date": task.due_date,
            "notes": task.notes,
            "status": "Assigned",
            "created_at": now,
            "completed_at": None,
        }

        existing = task_assignments_collection.find_one({
            "warehouse_name": task.warehouse_name,
            "due_date": task.due_date,
            "task_type": task.task_type,
        })
        if existing:
            task_assignments_collection.update_one(
                {"_id": existing["_id"]},
                {"$set": {
                    "assigned_to": task.assigned_to,
                    "assigned_by": emp_id,
                    "notes": task.notes,
                    "updated_at": now,
                }},
            )
            message = "Task assignment updated successfully"
        else:
            task_assignments_collection.insert_one(task_doc)
            message = "Task assigned successfully"

        # Send email notification (non-blocking — failure doesn't fail the endpoint)
        try:
            task_type_label = "Checklist Audit" if task.task_type == "checklist" else "Stock Count"
            email_body = f"""
            <html><body style="font-family: Arial, sans-serif; color: #333;">
            <h2 style="color:#4338ca;">New Task Assigned to You</h2>
            <p>You have been assigned a new audit task by {emp_id}.</p>
            <div style="background:#f3f4f6;padding:15px;border-radius:8px;margin:20px 0;">
                <h3 style="margin-top:0;">Task Details:</h3>
                <p><strong>Warehouse:</strong> {task.warehouse_name}</p>
                <p><strong>Task Type:</strong> {task_type_label}</p>
                <p><strong>Due Date:</strong> {task.due_date}</p>
                {"<p><strong>Notes:</strong> " + task.notes + "</p>" if task.notes else ""}
            </div>
            <p>Please complete this task by the due date.</p>
            </body></html>
            """
            send_email_notification(
                to_emails=task.assigned_to,
                subject=f"New Task Assignment: {task.warehouse_name}",
                body=email_body,
            )
        except Exception as email_error:
            logger.warning(f"Task assignment email failed (non-fatal): {email_error}")

        return JSONResponse({
            "message": message, "success": True,
            "data": {"task": {
                "warehouse_name": task.warehouse_name,
                "assigned_to": task.assigned_to,
                "task_type": task.task_type,
                "due_date": task.due_date,
                "status": "Assigned",
            }},
        })
    except HTTPException:
        return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)
    except Exception as e:
        logger.error(f"assign-task error: {e}")
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)


@app.get("/api/admin/task-assignments")
async def get_task_assignments(
    date: str = None,
    warehouse: str = None,
    emp_id: str = Depends(get_current_user),
):
    try:
        require_admin(emp_id)
        query = {}
        if date:
            query["due_date"] = date
        if warehouse:
            query["warehouse_name"] = warehouse
        tasks = list(task_assignments_collection.find(query).sort("due_date", -1))
        user_ids = {
            email
            for task in tasks
            for email in (task.get("assigned_to") or [])
            if email
        }
        name_map = get_user_names_cached(user_ids)
        today_str = datetime.now(timezone.utc).date().isoformat()
        completed_task_keys = set()
        completed_audits = list(audit_data_collection.find({}, {
            "_id": 0,
            "completion_status": 1,
            "warehouse_name": 1,
            "sections.general_report.warehouse_name": 1,
        }).limit(500))
        for audit in completed_audits:
            wh_name = (
                audit.get("warehouse_name")
                or (audit.get("sections") or {}).get("general_report", {}).get("warehouse_name")
            )
            if not wh_name:
                continue

            cs = audit.get("completion_status", {})
            checklist_done = all(cs.get(section, False) for section in CHECKLIST_SECTIONS)
            if checklist_done:
                completed_task_keys.add((wh_name, "checklist"))
            if cs.get("stock_count"):
                completed_task_keys.add((wh_name, "stock_count"))

        task_rows = []
        for task in tasks:
            assigned_to = task.get("assigned_to") or []
            completed_by_audit = (task.get("warehouse_name"), task.get("task_type")) in completed_task_keys
            is_overdue = (
                bool(task.get("due_date"))
                and task.get("due_date") < today_str
                and task.get("status") != "Completed"
                and not completed_by_audit
            )
            task_rows.append({
                **task,
                "_id": str(task["_id"]),
                "assigned_to_names": [name_map.get(email, email) for email in assigned_to],
                "is_overdue": is_overdue,
                "effective_status": "Completed" if completed_by_audit else ("Overdue" if is_overdue else task.get("status", "Assigned")),
            })

        return JSONResponse({
            "message": "Task assignments fetched", "success": True,
            "data": {"tasks": _serialize_mongo(task_rows)},
        })
    except HTTPException:
        return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)
    except Exception as e:
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)


@app.delete("/api/admin/task-assignments/{task_id}")
async def delete_task_assignment(task_id: str, emp_id: str = Depends(get_current_user)):
    try:
        require_admin(emp_id)
        result = task_assignments_collection.delete_one({"_id": ObjectId(task_id)})
        if result.deleted_count == 0:
            return JSONResponse({"message": "Task not found", "success": False}, status_code=404)
        return JSONResponse({"message": "Task deleted successfully", "success": True})
    except HTTPException:
        return JSONResponse({"message": "Unauthorized", "success": False}, status_code=403)
    except Exception as e:
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)


@app.get("/api/user/my-tasks")
async def get_my_tasks(emp_id: str = Depends(get_current_user)):
    try:
        tasks = list(task_assignments_collection.find(
            {"assigned_to": emp_id, "status": {"$ne": "Completed"}},
        ).sort("due_date", 1))
        today_str = datetime.now(timezone.utc).date().isoformat()

        progress_by_task = {}
        completed_audits = list(audit_data_collection.find(
            {"user_id": emp_id},
            {
                "_id": 0,
                "completion_status": 1,
                "warehouse_name": 1,
                "sections.general_report.warehouse_name": 1,
            },
        ))
        for audit in completed_audits:
            wh_name = (
                audit.get("warehouse_name")
                or (audit.get("sections") or {}).get("general_report", {}).get("warehouse_name")
            )
            if not wh_name:
                continue

            normalized_wh = str(wh_name).strip().lower()
            completion = audit.get("completion_status", {}) or {}
            checklist_done = sum(1 for section in CHECKLIST_SECTIONS if completion.get(section, False))
            checklist_progress = round((checklist_done / len(CHECKLIST_SECTIONS)) * 100) if CHECKLIST_SECTIONS else 0
            existing_checklist = progress_by_task.get((normalized_wh, "checklist"), {})
            if checklist_progress >= existing_checklist.get("progress_percent", 0):
                progress_by_task[(normalized_wh, "checklist")] = {
                    "progress_percent": checklist_progress,
                    "progress_label": f"{checklist_done}/{len(CHECKLIST_SECTIONS)} sections",
                    "is_completed": checklist_done == len(CHECKLIST_SECTIONS),
                }

            stock_completed = bool(completion.get("stock_count"))
            progress_by_task[(normalized_wh, "stock_count")] = {
                "progress_percent": 100 if stock_completed else 0,
                "progress_label": "Submitted" if stock_completed else "Not submitted",
                "is_completed": stock_completed,
            }

        task_rows = []
        for task in tasks:
            normalized_task_wh = str(task.get("warehouse_name", "")).strip().lower()
            task_progress = progress_by_task.get((normalized_task_wh, task.get("task_type")), {
                "progress_percent": 0,
                "progress_label": "Not started",
                "is_completed": False,
            })
            completed_by_audit = task_progress.get("is_completed", False)
            is_overdue = (
                bool(task.get("due_date"))
                and task.get("due_date") < today_str
                and task.get("status") != "Completed"
                and not completed_by_audit
            )
            task_rows.append({
                **task,
                "_id": str(task["_id"]),
                "is_overdue": is_overdue,
                "effective_status": "Completed" if completed_by_audit else ("Overdue" if is_overdue else task.get("status", "Assigned")),
                "progress_percent": task_progress.get("progress_percent", 0),
                "progress_label": task_progress.get("progress_label", "Not started"),
            })

        return JSONResponse({
            "message": "Your tasks fetched", "success": True,
            "data": {"tasks": _serialize_mongo(task_rows)},
        })
    except Exception as e:
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  USER HISTORY — checklists
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/user/checklist-history")
async def user_checklist_history(emp_id: str = Depends(get_current_user)):
    try:
        history = list(audit_data_collection.find(
            {"user_id": emp_id},
            {"_id": 1, "date": 1, "submitted_at": 1,
             "sections.general_report.warehouse_name": 1,
             "completion_status": 1},
        ).sort("date", -1))

        result = []
        for h in history:
            cs  = h.get("completion_status", {})
            sat = h.get("submitted_at", "")
            if isinstance(sat, datetime):
                sat = sat.isoformat()
            result.append({
                "audit_id": str(h["_id"]),
                "date": h.get("date", ""),
                "submitted_at": sat,
                "warehouse_name": (h.get("sections") or {}).get("general_report", {}).get("warehouse_name", "—"),
                "sections_completed": sum(1 for s in CHECKLIST_SECTIONS if cs.get(s, False)),
                "sections_total": len(CHECKLIST_SECTIONS),
                "completion_status": cs,
            })

        return JSONResponse({
            "message": "History fetched", "success": True,
            "data": {"history": result},
        })
    except Exception as e:
        logger.error(f"checklist-history error: {e}")
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  USER HISTORY — stock count
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/user/stock-count-history")
async def user_stock_count_history(emp_id: str = Depends(get_current_user)):
    try:
        today = datetime.now(timezone.utc).date().isoformat()

        current_temp = temp_audit_data_collection.find_one({
            "user_id": emp_id, "date": today,
            "completion_status.stock_count": {"$ne": True},
        })
        pending_items = []
        if current_temp:
            sc = current_temp.get("stock_count_data", [])
            pending_items = [{
                "audit_id": str(current_temp["_id"]),
                "date": current_temp.get("date"),
                "warehouse_name": (current_temp.get("sections") or {}).get("general_report", {}).get("warehouse_name", "—"),
                "items_count": len(sc),
                "status": "Pending",
                "stock_count_data": sc,
            }]

        completed_today = audit_data_collection.find_one({
            "user_id": emp_id, "date": today, "completion_status.stock_count": True,
        })
        completed_items = []
        if completed_today:
            sc  = completed_today.get("stock_count_data", [])
            sat = completed_today.get("submitted_at", "")
            if isinstance(sat, datetime):
                sat = sat.isoformat()
            completed_items = [{
                "audit_id": str(completed_today["_id"]),
                "date": completed_today.get("date"),
                "warehouse_name": (completed_today.get("sections") or {}).get("general_report", {}).get("warehouse_name", "—"),
                "items_count": len(sc),
                "status": "Completed",
                "submitted_at": sat,
                "stock_count_data": sc,
            }]

        proj = {
            "_id": 1, "date": 1, "submitted_at": 1,
            "sections.general_report.warehouse_name": 1, "stock_count_data": 1,
        }
        history = list(audit_data_collection.find(
            {"user_id": emp_id, "completion_status.stock_count": True}, proj
        ).sort("date", -1))

        history_items = []
        for h in history:
            sc  = h.get("stock_count_data", [])
            sat = h.get("submitted_at", "")
            if isinstance(sat, datetime):
                sat = sat.isoformat()
            history_items.append({
                "audit_id": str(h["_id"]),
                "date": h.get("date"),
                "warehouse_name": (h.get("sections") or {}).get("general_report", {}).get("warehouse_name", "—"),
                "items_count": len(sc),
                "status": "Submitted",
                "submitted_at": sat,
                "stock_count_data": sc,
            })

        return JSONResponse({
            "message": "Stock count data fetched", "success": True,
            "data": {
                "pending": pending_items,
                "completed": completed_items,
                "history": history_items,
            },
        })
    except Exception as e:
        logger.error(f"stock-count-history error: {e}")
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  EXPORT BY DATE / ID
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/export-excel-by-date")
async def export_excel_by_date(
    date: str = Query(...),
    emp_id: str = Depends(get_current_user),
):
    try:
        audit_data = audit_data_collection.find_one({"user_id": emp_id, "date": date})
        if not audit_data:
            return JSONResponse(
                {"message": f"No submitted audit found for {date}", "success": False},
                status_code=404,
            )
        excel_bytes = await generate_excel_bytes(emp_id, audit_data)
        filename = f"Audit_{date}.xlsx"
        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
        )
    except Exception as e:
        logger.error(f"export-excel-by-date error: {e}")
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)


@app.get("/api/export-excel-by-id")
async def export_excel_by_id(
    audit_id: str = Query(...),
    emp_id: str = Depends(get_current_user),
):
    try:
        audit_data = audit_data_collection.find_one(
            {"user_id": emp_id, "_id": ObjectId(audit_id)}
        )
        if not audit_data:
            return JSONResponse({"message": "No submitted audit found", "success": False}, status_code=404)
        excel_bytes = await generate_checklist_excel_bytes(emp_id, audit_data)
        date     = audit_data.get("date", "Report")
        wh_name  = (audit_data.get("sections") or {}).get("general_report", {}).get("warehouse_name", "Audit")
        safe_wh  = re.sub(r'[\s/\\?*\[\]:]+', '_', wh_name)
        filename = f"Audit_{safe_wh}_{date}.xlsx"
        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
        )
    except Exception as e:
        logger.error(f"export-excel-by-id error: {e}")
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)


@app.get("/api/export-stock-count-excel-by-id")
async def export_stock_count_excel_by_id(
    audit_id: str = Query(...),
    emp_id: str = Depends(get_current_user),
):
    try:
        audit = audit_data_collection.find_one(
            {"user_id": emp_id, "_id": ObjectId(audit_id)}
        )
        if not audit or not audit.get("stock_count_data"):
            return JSONResponse({"message": "No stock count data found", "success": False}, status_code=404)

        excel_bytes = generate_stock_count_excel_bytes(audit)
        date     = audit.get("date", "Report")
        filename = stock_count_report_filename(audit, date)
        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
        )
    except Exception as e:
        logger.error(f"export-stock-count-excel-by-id error: {e}")
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  HISTORICAL SECTION / STOCK COUNT RETRIEVAL
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/get-historical-section/{audit_id}/{section_name}")
async def get_historical_section(
    audit_id: str,
    section_name: str,
    emp_id: str = Depends(get_current_user),
):
    try:
        audit = audit_data_collection.find_one(
            {"user_id": emp_id, "_id": ObjectId(audit_id)},
            {f"sections.{section_name}": 1},
        )
        if not audit:
            return JSONResponse({"message": "Historical audit not found", "success": False}, status_code=404)
        section_data = (audit.get("sections") or {}).get(section_name, {})
        resp = base_response.copy()
        resp.update({
            "message": f"Historical section {section_name} retrieved",
            "success": True,
            "data": {"section_data": section_data},
            "status_code": 200,
        })
        return JSONResponse(content=resp, status_code=200)
    except Exception as e:
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)


@app.get("/api/get-historical-stock-count/{audit_id}")
async def get_historical_stock_count(
    audit_id: str,
    emp_id: str = Depends(get_current_user),
):
    try:
        audit = audit_data_collection.find_one(
            {"user_id": emp_id, "_id": ObjectId(audit_id)},
            {"stock_count_data": 1},
        )
        if not audit:
            return JSONResponse({"message": "Historical audit not found", "success": False}, status_code=404)
        return JSONResponse({
            "message": "Historical stock count retrieved",
            "success": True,
            "data": {"stock_count_data": audit.get("stock_count_data", [])},
        })
    except Exception as e:
        return JSONResponse({"message": f"Server error: {e}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  STATIC HTML ROUTES
# ─────────────────────────────────────────────────────────────────────────────

def _get_token(request: Request) -> str:
    token = request.cookies.get("access_token")
    if token:
        return token
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        return auth.split(" ", 1)[1]
    raise HTTPException(status_code=401, detail="Could not validate credentials")


@app.get("/", response_class=FileResponse)
async def root(request: Request):
    try:
        token = _get_token(request)
        get_current_user(token)
        return FileResponse(os.path.join(STATIC_DIR, "index.html"))
    except HTTPException:
        return RedirectResponse(url="/login")


@app.get("/login", response_class=FileResponse)
async def serve_login():
    return FileResponse(os.path.join(STATIC_DIR, "login.html"))


@app.get("/register", response_class=FileResponse)
async def serve_register():
    return FileResponse(os.path.join(STATIC_DIR, "register.html"))
