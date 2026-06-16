import datetime
import os
import re
from typing import Dict, Optional, List
from datetime import timezone
from dotenv import load_dotenv
from fastapi import FastAPI, Depends, HTTPException, Request, Response, status, Query, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import JSONResponse
import logging
import requests
from app.auth import *
from app.database import *
from app.database import fs
from app.database import fs, warehouse_master_collection
from app.models import AuditForm, UserLogin, UserRegister
from fastapi.responses import JSONResponse, StreamingResponse, FileResponse
import io
from docx import Document
from docx.shared import Pt
import base64, re
from docx.shared import Inches
import smtplib
from email.message import EmailMessage
import tempfile
from fastapi import UploadFile, Form
from fastapi import FastAPI, Depends, Form, File, UploadFile, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse, RedirectResponse
from datetime import datetime
import io
import os
import smtplib
from email.message import EmailMessage
import base64
import re
import logging
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import smtplib
from email.message import EmailMessage
from fastapi import UploadFile, Form, File, Depends
from typing import Optional
from datetime import datetime, timezone
from fastapi.responses import JSONResponse
import os, io, base64, re
from docx import Document
from docx.shared import Pt, Inches
import bcrypt
import pandas as pd
from bson import ObjectId
from pydantic import BaseModel

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

load_dotenv()

SERPAPI_KEY = os.getenv("SERPAPI_KEY")

app = FastAPI()

# Get the absolute path to the static folder
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
STATIC_DIR = os.path.join(BASE_DIR, "static")

# Mount static files for serving HTML/CSS/JS
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

# CORS configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Base response format
base_response = {
    "message": "",
    "success": False,
    "data": None,
    "status_code": status.HTTP_400_BAD_REQUEST
}

def validate_password(password: str) -> bool:
    if len(password) < 8:
        return False
    if not re.search(r"[A-Z]", password):
        return False
    if not re.search(r"[a-z]", password):
        return False
    if not re.search(r"\d", password):
        return False
    if not re.search(r"[!@#$%^&*()]", password):
        return False
    return True


# ─────────────────────────────────────────────────────────────────────────────
#  AUTH ENDPOINTS
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/register")
async def register(user: UserRegister):
    logger.info(f"Register attempt for email: {user.email}")
    try:
        if not validate_password(user.password):
            return JSONResponse(
                content={"message": "Invalid password format", "success": False},
                status_code=status.HTTP_400_BAD_REQUEST
            )
        if user.password != user.confirm_password:
            return JSONResponse(
                content={"message": "Passwords do not match", "success": False},
                status_code=status.HTTP_400_BAD_REQUEST
            )
        existing_user = users.find_one({"email": user.email})
        if existing_user:
            return JSONResponse(
                content={"message": "Email already registered", "success": False},
                status_code=status.HTTP_400_BAD_REQUEST
            )
        hashed_password = bcrypt.hashpw(
            user.password.encode("utf-8"), bcrypt.gensalt()
        ).decode("utf-8")
        users.insert_one({
            "name": user.name,
            "email": user.email,
            "password_hash": hashed_password,
            "created_at": datetime.now(timezone.utc),
        })
        return JSONResponse(
            content={"message": "User registered successfully", "success": True, "data": {"email": user.email}},
            status_code=status.HTTP_201_CREATED
        )
    except Exception as e:
        logger.error(f"Error in register: {str(e)}")
        return JSONResponse(
            content={"message": f"Server error: {str(e)}", "success": False},
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@app.post("/api/login")
def login(user: UserLogin):
    logger.info(f"Login attempt for email: {user.email}")
    db_user = users.find_one({"email": user.email})
    if not db_user:
        return JSONResponse({"message": "Invalid email or password", "success": False}, status_code=401)
    password = user.password
    password_hash = db_user.get("password_hash")
    if not isinstance(password, str) or not isinstance(password_hash, str):
        return JSONResponse({"message": "Invalid email or password", "success": False}, status_code=401)
    if not bcrypt.checkpw(password.encode("utf-8"), password_hash.encode("utf-8")):
        return JSONResponse({"message": "Invalid email or password", "success": False}, status_code=401)
    token = create_jwt({"sub": user.email})
    return JSONResponse(
        {"message": "Logged in successfully", "success": True, "data": {"access_token": token}},
        status_code=200
    )


@app.get("/api/me")
async def get_me(emp_id: str = Depends(get_current_user)):
    try:
        user = users.find_one({"email": emp_id})
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        response = base_response.copy()
        response.update({
            "message": "User info retrieved successfully",
            "success": True,
            "data": {"email": user["email"], "name": user.get("name", "Unknown")},
            "status_code": status.HTTP_200_OK
        })
        return JSONResponse(content=response, status_code=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f"Error in get_me: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/logout")
async def logout(emp_id: str = Depends(get_current_user)):
    return JSONResponse(
        content={"message": "Logged out successfully", "success": True, "data": None},
        status_code=200
    )


# ─────────────────────────────────────────────────────────────────────────────
#  AUDIT SECTIONS
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/get-sections")
async def get_sections(emp_id: str = Depends(get_current_user)):
    try:
        today = datetime.now(timezone.utc).date().isoformat()
        audit = temp_audit_data_collection.find_one({"user_id": emp_id, "date": today})
        section_keys = [
            "general_report", "stock_reconciliation",
            "observations_on_stacking", "observations_on_warehouse_operations",
            "observations_on_warehouse_record_keeping", "observations_on_wh_infrastructure",
            "observations_on_quality_operation", "checklist_wrt_exchange_circular_mentha_oil",
            "checklist_wrt_exchange_circular_metal", "checklist_wrt_exchange_circular_cotton_bales",
            "signature", "photo"
        ]
        completion_status = {
            k: (audit["completion_status"].get(k, False) if audit and "completion_status" in audit else False)
            for k in section_keys
        }
        response = base_response.copy()
        response.update({
            "message": "Sections retrieved successfully",
            "success": True,
            "data": {"completion_status": completion_status},
            "status_code": status.HTTP_200_OK
        })
        return JSONResponse(content=response, status_code=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f"Error in get_sections: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/get-section/{section_name}")
async def get_section(section_name: str, emp_id: str = Depends(get_current_user)):
    try:
        today = datetime.now(timezone.utc).date().isoformat()
        audit = temp_audit_data_collection.find_one({"user_id": emp_id, "date": today})
        section_data = audit["sections"].get(section_name, {}) if audit and "sections" in audit else {}
        response = base_response.copy()
        response.update({
            "message": f"Section {section_name} retrieved successfully",
            "success": True,
            "data": {"section_data": section_data},
            "status_code": status.HTTP_200_OK
        })
        return JSONResponse(content=response, status_code=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f"Error in get_section: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/save-section")
async def save_section(request: Request, emp_id: str = Depends(get_current_user)):
    try:
        body = await request.json()
        section = body.get("section")
        data = body.get("data")
        date = body.get("date")
        if not section or not data or not date:
            return JSONResponse(
                content={"message": "Missing required fields (section, data, date)", "success": False},
                status_code=400
            )

        # --- GridFS: extract photo and store separately ---
        if section == "photo" and "photo" in data and data["photo"].startswith("data:image"):
            try:
                header, b64data = data["photo"].split(",", 1)
                img_bytes = base64.b64decode(b64data)
                file_id = fs.put(
                    img_bytes,
                    filename=f"{emp_id}_{date}_photo.png",
                    content_type="image/png",
                    metadata={"user_id": emp_id, "date": date}
                )
                data["photo"] = None          # don't store base64 in document
                data["photo_file_id"] = str(file_id)  # store GridFS reference
            except Exception as img_err:
                logger.warning(f"GridFS photo store failed, falling back to base64: {img_err}")

        audit = temp_audit_data_collection.find_one({"user_id": emp_id, "date": date})
        if not audit:
            audit = {
                "user_id": emp_id, "date": date, "sections": {},
                "completion_status": {}, "submitted_by": emp_id,
                "submitted_at": datetime.now(timezone.utc)
            }
        audit["sections"][section] = data
        audit["completion_status"][section] = True
        if audit.get("_id"):
            temp_audit_data_collection.update_one(
                {"_id": audit["_id"]},
                {"$set": {"sections": audit["sections"], "completion_status": audit["completion_status"]}}
            )
        else:
            temp_audit_data_collection.insert_one(audit)
        response = base_response.copy()
        response.update({
            "message": f"Section {section} saved successfully",
            "success": True,
            "data": {"completion_status": audit["completion_status"]},
            "status_code": 200
        })
        return JSONResponse(content=response, status_code=200)
    except Exception as e:
        logger.error(f"Error in save_section: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/get-photo/{file_id}")
async def get_photo(file_id: str, emp_id: str = Depends(get_current_user)):
    """Serve a photo stored in GridFS by its file ID."""
    try:
        oid = ObjectId(file_id)
        if not fs.exists(oid):
            return JSONResponse({"message": "Photo not found", "success": False}, status_code=404)
        grid_out = fs.get(oid)
        data = grid_out.read()
        return StreamingResponse(io.BytesIO(data), media_type="image/png")
    except Exception as e:
        logger.error(f"get_photo error: {e}")
        return JSONResponse({"message": str(e), "success": False}, status_code=500)


@app.post("/api/submit-audit")
async def submit_audit(emp_id: str = Depends(get_current_user)):
    try:
        today = datetime.now(timezone.utc).date().isoformat()
        temp_audit = temp_audit_data_collection.find_one({"user_id": emp_id, "date": today})
        if not temp_audit:
            raise HTTPException(status_code=404, detail="No audit data found to submit")
        if not all(temp_audit.get("completion_status", {}).values()):
            raise HTTPException(status_code=400, detail="Not all sections are completed")
        
        # Update submitted_at timestamp to the actual submission time
        temp_audit["submitted_at"] = datetime.now(timezone.utc)
        
        result = audit_data_collection.insert_one(temp_audit)
        temp_audit_data_collection.delete_one({"_id": temp_audit["_id"]})
        response = base_response.copy()
        response.update({
            "message": "Audit submitted successfully", "success": True,
            "data": {"submitted": True, "audit_id": str(result.inserted_id)},
            "status_code": 200
        })
        return JSONResponse(content=response, status_code=200)
    except Exception as e:
        logger.error(f"Error in submit_audit: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/clear-sections")
async def clear_sections(emp_id: str = Depends(get_current_user)):
    try:
        today = datetime.now(timezone.utc).date().isoformat()
        temp_audit_data_collection.delete_one({"user_id": emp_id, "date": today})
        return JSONResponse({"message": "Sections cleared", "success": True}, status_code=200)
    except Exception as e:
        return JSONResponse({"message": str(e), "success": False}, status_code=500)

# ═════════════════════════════════════════════════════════════════════════════
# ADD THIS ENDPOINT TO YOUR main.py (before the existing /api/get-location)
# ═════════════════════════════════════════════════════════════════════════════

@app.get("/api/get-ip-location")
def get_ip_location():
    """
    Fetch user's location via IP address from backend (avoids CORS issues).
    This solves the CORS problem when geolocation API is not available in production.
    """
    services = [
        ("https://freeipapi.com/api/json", lambda d: {"lat": d.get("latitude"), "lon": d.get("longitude")} if d.get("latitude") else None),
        ("https://ipapi.co/json/", lambda d: {"lat": d.get("latitude"), "lon": d.get("longitude")} if d.get("latitude") else None),
        ("https://ip-api.com/json/?fields=lat,lon,status", lambda d: {"lat": d.get("lat"), "lon": d.get("lon")} if d.get("status") == "success" else None),
    ]
    
    for service_url, parse_fn in services:
        try:
            res = requests.get(service_url, timeout=5)
            if res.status_code == 200:
                data = res.json()
                coords = parse_fn(data)
                if coords and coords["lat"] and coords["lon"]:
                    logger.info(f"IP geolocation via {service_url}: {coords}")
                    return {"latitude": coords["lat"], "longitude": coords["lon"], "success": True}
        except Exception as e:
            logger.warning(f"IP geolocation service {service_url} failed: {e}")
    
    logger.error("All IP geolocation services failed")
    return {"success": False, "error": "Could not determine IP location"}

# ─────────────────────────────────────────────────────────────────────────────
#  LOCATION
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/get-location")
def get_location(lat: float = Query(...), lon: float = Query(...)):
    serp_url = f"https://serpapi.com/search?engine=google_maps&q={lat},{lon}&type=search&api_key={SERPAPI_KEY}"
    try:
        serp_res = requests.get(serp_url, timeout=6)
        serp_data = serp_res.json()
        if serp_res.status_code == 200 and "search_metadata" in serp_data:
            place_result = serp_data.get("place_results", {})
            plus_code = place_result.get("plus_code") or "N/A"
            address = place_result.get("title") or "N/A"
            maps_url = serp_data.get("search_metadata", {}).get("google_maps_url")
            if maps_url:
                return {"source": "serpapi", "latitude": lat, "longitude": lon, "plus_code": plus_code, "address": address, "maps_url": maps_url}
        raise Exception("SerpApi failed or incomplete")
    except Exception as e:
        print(f"⚠️ SerpApi failed: {e}")
        osm_url = f"https://nominatim.openstreetmap.org/reverse?format=jsonv2&lat={lat}&lon={lon}"
        osm_res = requests.get(osm_url, headers={"User-Agent": "audit-app"})
        osm_data = osm_res.json()
        address = osm_data.get("display_name", "Address not found")
        plus_code = address
        maps_url = f"https://www.google.com/maps/search/{lat}%2C{lon}?hl=en"
        return {"source": "osm", "latitude": lat, "longitude": lon, "plus_code": plus_code, "address": address, "maps_url": maps_url}


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL GENERATION HELPER
# ─────────────────────────────────────────────────────────────────────────────

def _adjust_ws(ws, widths):
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


async def generate_checklist_excel_bytes(emp_id: str, audit_data: dict) -> bytes:
    """Checklist-only Excel – NO Stock Count sheet."""
    wb = Workbook()
    wb.remove(wb.active)
    sections = audit_data.get("sections", {})

    # General Report
    ws = wb.create_sheet("General Report")
    ws.append(["Field", "Value"])
    gr = sections.get("general_report", {})
    if gr:
        for k, v in gr.items():
            if k == "previous_audits" and isinstance(v, list):
                ws.append(["Previous Audits", ""])
                for i, rec in enumerate(v, start=1):
                    ws.append([f"  Record {i} – Date", str(rec.get("date", ""))])
                    ws.append([f"  Record {i} – Auditor Name", str(rec.get("auditor_name", ""))])
                    ws.append([f"  Record {i} – Auditor Type", str(rec.get("auditor_type", ""))])
                    if rec.get("agency_name"):
                        ws.append([f"  Record {i} – Agency Name", str(rec.get("agency_name", ""))])
            else:
                ws.append([k.replace("_", " ").title(), str(v)])
    else:
        ws.append(["No general report saved.", ""])
    _adjust_ws(ws, [40, 30])

    # Stock Reconciliation
    ws = wb.create_sheet("Stock Reconciliation")
    ws.append(["Commodity Name", "Stock Type", "Qty as per MCXCCL", "Qty as per Registered", "Qty as per Physical", "Difference", "Remarks"])
    stock = sections.get("stock_reconciliation", {}).get("commodities", [])
    if stock:
        for item in stock:
            ws.append([item.get("commodity_name",""), item.get("commodity",""), item.get("qty_mcxccl",""), item.get("qty_registered",""), item.get("qty_physical",""), item.get("difference",""), item.get("remarks","")])
    else:
        ws.append(["No stock data.", "", "", "", "", "", ""])
    _adjust_ws(ws, [20, 20, 20, 20, 20, 20, 30])

    # Question-based sections
    q_sections = [
        ("observations_on_stacking", "Observations on Stacking"),
        ("observations_on_warehouse_operations", "Observations on WH Operations"),
        ("observations_on_warehouse_record_keeping", "Observations on WH Record Keeping"),
        ("observations_on_wh_infrastructure", "Observations on WH Infrastructure"),
        ("observations_on_quality_operation", "Observations on Quality Operation"),
        ("checklist_wrt_exchange_circular_mentha_oil", "Checklist Mentha Oil"),
        ("checklist_wrt_exchange_circular_metal", "Checklist Metals"),
        ("checklist_wrt_exchange_circular_cotton_bales", "Checklist Cotton Bales"),
    ]
    for key, title in q_sections:
        ws = wb.create_sheet(title)
        ws.append(["Question", "Yes/No", "Remarks"])
        qlist = sections.get(key, {}).get("questions", [])
        if qlist:
            for idx, q in enumerate(qlist, start=1):
                ws.append([f"{idx}. {q.get('question', f'Question {idx}').strip()}", q.get("answer","").strip(), q.get("remarks","").strip()])
        else:
            ws.append(["No data saved.", "", ""])
        _adjust_ws(ws, [60, 10, 30])

    # Signature
    ws = wb.create_sheet("Signature")
    sig = sections.get("signature", {}).get("signature")
    if sig:
        try:
            img_data = re.sub("^data:image/.+;base64,", "", sig)
            img_bytes = io.BytesIO(base64.b64decode(img_data))
            img = Image(img_bytes)
            img.width = 250; img.height = 150
            ws["A1"] = "Signature captured during the audit"
            ws.row_dimensions[1].height = 18
            ws.row_dimensions[2].height = 8
            ws.add_image(img, "A3")
        except Exception as e:
            ws["A1"] = f"Unable to embed signature: {e}"
    else:
        ws["A1"] = "Signature not found."
    ws.column_dimensions["A"].width = 60

    # Photo
    # Photo section - support both single photo (legacy) and multiple photos
    ws = wb.create_sheet("Photo")
    photo_section = sections.get("photo", {})
    
    # Check if it's the new multi-photo format
    photos_list = photo_section.get("photos", [])
    
    # Handle legacy single photo format
    if not photos_list and photo_section.get("photo"):
        photos_list = [{
            "photo": photo_section.get("photo"),
            "maps_url": photo_section.get("maps_url", ""),
            "timestamp": datetime.now().isoformat(),
            "location_text": "Legacy photo"
        }]
    
    if photos_list:
        # Multi-photo format
        ws.append(["Photo #", "Timestamp", "Location", "Google Maps Link"])
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 60  # Wide column for images
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 50
        
        for idx, photo_item in enumerate(photos_list, 1):
            timestamp = photo_item.get("timestamp", "N/A")
            location = photo_item.get("location_text", "N/A")
            maps_url = photo_item.get("maps_url", "")
            
            # Add data row
            data_row = idx + 1
            ws[f"A{data_row}"] = f"Photo {idx}"
            ws[f"C{data_row}"] = location
            ws[f"D{data_row}"] = maps_url if maps_url else "N/A"
            
            # Try to embed the image
            try:
                photo_data = photo_item.get("photo", "")
                if photo_data:
                    # Handle base64 encoded image
                    if photo_data.startswith("data:image"):
                        img_data = re.sub("^data:image/.+;base64,", "", photo_data)
                        photo_bytes = base64.b64decode(img_data)
                    else:
                        photo_bytes = base64.b64decode(photo_data)
                    
                    # Create PIL image and resize if needed
                    from PIL import Image as PILImage
                    pil_img = PILImage.open(io.BytesIO(photo_bytes))
                    
                    # Resize to reasonable size for Excel
                    max_size = (400, 300)
                    pil_img.thumbnail(max_size, PILImage.Resampling.LANCZOS)
                    
                    # Save to bytes
                    img_buf = io.BytesIO()
                    pil_img.save(img_buf, format='PNG')
                    img_buf.seek(0)
                    
                    # Create Excel image and add to worksheet
                    excel_img = Image(img_buf)
                    excel_img.width = 400
                    excel_img.height = 300
                    
                    # Position in column B
                    ws.add_image(excel_img, f"B{data_row}")
                    ws.row_dimensions[data_row].height = 225  # Adjust row height for image
                    
            except Exception as e:
                logger.warning(f"Failed to embed photo {idx}: {e}")
                ws[f"B{data_row}"] = f"[Image embedding failed: {str(e)}]"
    else:
        ws.append(["No photos captured"])
        ws.column_dimensions["A"].width = 30

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def generate_stock_count_excel_bytes(audit_data: dict) -> bytes:
    """Stock-count-only Excel – no checklist sheets."""
    sc_data = audit_data.get("stock_count_data", [])
    df = pd.DataFrame(sc_data) if sc_data else pd.DataFrame(columns=["sheet_name", "item_name", "item_code", "qty", "physical_amount", "remarks"])
    columns = ["sheet_name", "item_name", "item_code", "qty", "physical_amount", "remarks"]
    available_cols = [c for c in columns if c in df.columns]
    df = df[available_cols]
    df.rename(columns={"sheet_name": "Sheet Name", "item_name": "Item Name",
                       "item_code": "Item Code", "qty": "Expected Qty",
                       "physical_amount": "Physical Count", "remarks": "Remarks"}, inplace=True)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Stock Count')
    return output.getvalue()


# keep old name as alias so existing callers still work
async def generate_excel_bytes(emp_id: str, audit_data: dict) -> bytes:
    return await generate_checklist_excel_bytes(emp_id, audit_data)



# ─────────────────────────────────────────────────────────────────────────────
#  EXPORT EXCEL
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/export-excel")
async def export_excel(emp_id: str = Depends(get_current_user)):
    try:
        today = datetime.now(timezone.utc).date().isoformat()
        # Check temp collection first (audit in progress), then submitted collection
        audit_data = temp_audit_data_collection.find_one({"user_id": emp_id, "date": today})
        if not audit_data:
            audit_data = audit_data_collection.find_one({"user_id": emp_id, "date": today})
        if not audit_data:
            return JSONResponse({"message": "No audit data for today", "success": False}, status_code=404)
        completion = audit_data.get("completion_status", {})
        expected = [
            "general_report", "stock_reconciliation",
            "observations_on_stacking", "observations_on_warehouse_operations",
            "observations_on_warehouse_record_keeping", "observations_on_wh_infrastructure",
            "observations_on_quality_operation", "checklist_wrt_exchange_circular_mentha_oil",
            "checklist_wrt_exchange_circular_metal", "checklist_wrt_exchange_circular_cotton_bales",
            "signature", "photo"
        ]
        if not all(completion.get(s, False) for s in expected):
            return JSONResponse({"message": "Complete all sections before exporting", "success": False}, status_code=400)
        excel_bytes = await generate_excel_bytes(emp_id, audit_data)
        filename = f"audit_{emp_id}_{today}.xlsx"
        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename*=UTF-8\'\'{filename}'}
        )
    except Exception as e:
        logger.error(f"Export-excel error: {e}")
        return JSONResponse({"message": f"Server error: {str(e)}", "success": False}, status_code=500)


# ─────────────────────────────────────────────────────────────────────────────
#  SEND EMAIL
# ─────────────────────────────────────────────────────────────────────────────

